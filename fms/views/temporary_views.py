import calendar
import datetime
import os
import openpyxl
import traceback
from openpyxl.styles import PatternFill
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from fms.models import Budget, CsManage, SubmissionDocument
from fms.models import Progress, DepartmentMaster, Log, DailyConstruction
from django.utils.timezone import make_aware
from socket import gethostname
from fms.views.common_def_views import output_log_exception
from fms.views.notice_mail_views import step_notice


# 届出チェック進捗レコード作成
@login_required
@require_POST
def progress_cs_record_add(request):
    try:
        DIFF_JST_FROM_UTC = 9
        # JST = timezone(timedelta(hours=+9), 'JST')

        # now = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)
        now_naive = datetime.datetime.now()
        now = make_aware(now_naive)

        operator = request.user.username

        business_year = int(request.POST['business_year'])  # 年度情報取得・・・データからでもOK

        this_step = int(request.POST['this_step'])  # 現状ステップ情報取得・・・データからでもOK

        target_budget_data_list = Budget.objects.filter(business_year=business_year, lost_flag=0).all()

        for target_budget_data_list in target_budget_data_list:
            target_budget_id = target_budget_data_list.budget_id
            department = target_budget_data_list.budget_main_department.department_cd
            next_operator = target_budget_data_list.budget_department_charge_person

            # ↓↓↓↓↓↓↓↓　ここから追加　↓↓↓↓↓↓↓↓

            # ユーザー権限に登録されている場合の処理･･･普通はされているはず→次作業者、部署、部門データ取得
            # if user_attribute_id > 0:
            #     user_attribute_data = UserAttribute.objects.get(id=user_attribute_id, lost_flag=0)
            #     next_person = user_attribute_data.username
            #     next_division = user_attribute_data.division
             #    next_department = user_attribute_data.department

            # cs_data_num = CsGeneralAffairs.objects.all().count()
            cs_data_num = CsManage.objects.all().count()

            # 届け出チェックシートのレコードがない時の処理･･･チェックシートid=1 とする
            if cs_data_num == 0:
                this_cs_no = 1
            # 予算のレコードがある時の処理･･･最終の予算idを取得し、予算id=最終の予算id+1 とする
            else:
                # last_cs_data = CsGeneralAffairs.objects.all().order_by('-cs_no')[0]
                last_cs_data = CsManage.objects.all().order_by('-cs_no')[0]
                # 今回のCSidを設定(=最終のCSid+1)
                this_cs_no = last_cs_data.cs_no + 1
            # 設定した予算idでレコードを抽出し、あれば呼出、なければ新規作成･･･ないはずなので、新規作成
            cs_manage_data_num = CsManage.objects.filter(cs_no=this_cs_no).count()
            cs_manage_data, created = CsManage.objects.get_or_create(cs_no=this_cs_no)

            cs_manage_data.budget_id = target_budget_id
            # cs_manage_data.work_id = work_id
            cs_manage_data.cs_rev_no = 0
            cs_manage_data.lost_flag = 0
            cs_manage_data.entry_on_progress_flag = 1
            if cs_manage_data_num == 0:
                cs_manage_data.entry_datetime = now
                cs_manage_data.entry_operator = operator
            else:
                cs_manage_data.update_datetime = now
                cs_manage_data.update_operator = operator

            cs_manage_data.save()

            # ↑↑↑↑↑↑↑↑　ここまで追加　↑↑↑↑↑↑↑↑

            # 届出チェック状況を対象(notification_check)と予算idで抽出･･･あれば呼び出し、なければ新規登録
            progress_data, created = Progress.objects.get_or_create(target="cs", target_id=target_budget_id)
            # 各項目を設定
            progress_data.present_step = 134001001
            progress_data.present_operator = next_operator
            progress_data.present_department = department
            department_data = DepartmentMaster.objects.get(department_cd=department)
            progress_data.present_division = department_data.division_cd
            progress_data.last_operation_step = this_step
            progress_data.last_operator = operator
            progress_data.last_operation_datetime = now

            progress_data.save()

            # 進捗通知機能
            if this_step != 134001001:
                step_notice(progress_data)

    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# 提出書類レコード移行作成
@login_required
@require_POST
def submission_document_record_add(request):
    try:
        DIFF_JST_FROM_UTC = 9
        # JST = timezone(timedelta(hours=+9), 'JST')

        # now = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)
        now_naive = datetime.datetime.now()
        now = make_aware(now_naive)

        operator = request.user.username

        work_id = int(request.POST["work_id"])

        # 提出書類マスタデータを取得

        document_lists_at_plane = SubmissionDocument.objects.filter(work_id=work_id, lost_flag=0, entry_class='計画').all().order_by('display_order')
        for document_lists_at_plane in document_lists_at_plane:
            work_id = document_lists_at_plane.work_id
            rev_no = document_lists_at_plane.rev_no
            document_name = document_lists_at_plane.document_name
            number_of_copies = document_lists_at_plane.number_of_copies
            submission_deadline = document_lists_at_plane.submission_deadline

            document_lists_at_execute_num = SubmissionDocument.objects.filter(work_id=work_id, lost_flag=0, entry_class='実行', document_name=document_name).count()

            document_lists_at_execute, created = SubmissionDocument.objects.get_or_create(work_id=work_id, lost_flag=0, entry_class='実行', document_name=document_name)

            document_lists_at_execute.work_id = work_id
            document_lists_at_execute.rev_no = rev_no
            document_lists_at_execute.document_name = document_name
            document_lists_at_execute.number_of_copies = number_of_copies
            document_lists_at_execute.submission_deadline = submission_deadline
            document_lists_at_execute.entry_class = '実行'
            document_lists_at_execute.lost_flag = 0
            document_lists_at_execute.entry_on_progress_flag = 0

            if document_lists_at_execute_num > 0:
                document_lists_at_execute.update_datetime = now
                document_lists_at_execute.update_operator = operator

            else:
                document_lists_at_execute.entry_datetime = now
                document_lists_at_execute.entry_operator = operator
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

# ログリスト表示処理
@require_POST
def get_cs_log_lists(request):
    try:
        # JSからのPOST引数を取得・・・数値項目は数値に変換(int関数)
        target_id = request.POST['target_id']
        target = 'cs'
        log_ga_lists_num = 0
        log_ev_lists_num = 0
        log_sh_lists_num = 0
        log_en_lists_num = 0
        log_bd_lists_num = 0

        sql = """ SELECT fms_log.*, fms_stepmaster.step_name, fms_actionmaster.action_name, """
        sql = sql + """ [isk_tools_base].[dbo].[auth_user].first_name, [isk_tools_base].[dbo].[auth_user].last_name, """
        sql = sql + """ fms_targetmaster.target_name """
        sql = sql + """ FROM (((fms_log """
        sql = sql + """ LEFT JOIN fms_stepmaster ON fms_log.step=fms_stepmaster.step_id) """
        sql = sql + """ LEFT JOIN fms_actionmaster ON fms_log.action=fms_actionmaster.action_cd) """
        sql = sql + """ LEFT JOIN [isk_tools_base].[dbo].[auth_user] ON fms_log.operator=[isk_tools_base].[dbo].[auth_user].username) """
        sql = sql + """ LEFT JOIN fms_targetmaster ON fms_log.target=fms_targetmaster.target_cd"""
        sql = sql + """ WHERE fms_log.target='cs' AND fms_log.operator_division='SOU&ROU' AND fms_log.target_id=""" + target_id
        sql = sql + """ ORDER BY fms_log.operation_datetime DESC """

        log_ga_lists_num = Log.objects.filter(target='cs', target_id=target_id, operator_division='SOU&ROU').count()
        log_ga_lists = Log.objects.all().raw(sql)

        sql = """ SELECT fms_log.*, fms_stepmaster.step_name, fms_actionmaster.action_name, """
        sql = sql + """ [isk_tools_base].[dbo].[auth_user].first_name, [isk_tools_base].[dbo].[auth_user].last_name, """
        sql = sql + """ fms_targetmaster.target_name """
        sql = sql + """ FROM (((fms_log """
        sql = sql + """ LEFT JOIN fms_stepmaster ON fms_log.step=fms_stepmaster.step_id) """
        sql = sql + """ LEFT JOIN fms_actionmaster ON fms_log.action=fms_actionmaster.action_cd) """
        sql = sql + """ LEFT JOIN [isk_tools_base].[dbo].[auth_user] ON fms_log.operator=[isk_tools_base].[dbo].[auth_user].username) """
        sql = sql + """ LEFT JOIN fms_targetmaster ON fms_log.target=fms_targetmaster.target_cd"""
        sql = sql + """ WHERE fms_log.target='cs' AND (fms_log.operator_department='KA' OR fms_log.operator_department='KA&A') AND fms_log.target_id=""" + target_id
        sql = sql + """ ORDER BY fms_log.operation_datetime DESC """

        log_ev_lists_num = Log.objects.filter(target='cs', target_id=target_id, operator_department='KA').count()
        log_ev_lists = Log.objects.all().raw(sql)

        sql = """ SELECT fms_log.*, fms_stepmaster.step_name, fms_actionmaster.action_name, """
        sql = sql + """ [isk_tools_base].[dbo].[auth_user].first_name, [isk_tools_base].[dbo].[auth_user].last_name, """
        sql = sql + """ fms_targetmaster.target_name """
        sql = sql + """ FROM (((fms_log """
        sql = sql + """ LEFT JOIN fms_stepmaster ON fms_log.step=fms_stepmaster.step_id) """
        sql = sql + """ LEFT JOIN fms_actionmaster ON fms_log.action=fms_actionmaster.action_cd) """
        sql = sql + """ LEFT JOIN [isk_tools_base].[dbo].[auth_user] ON fms_log.operator=[isk_tools_base].[dbo].[auth_user].username) """
        sql = sql + """ LEFT JOIN fms_targetmaster ON fms_log.target=fms_targetmaster.target_cd"""
        sql = sql + """ WHERE fms_log.target='cs' AND (fms_log.operator_department='A' OR fms_log.operator_department='KA&A')AND fms_log.target_id=""" + target_id
        sql = sql + """ ORDER BY fms_log.operation_datetime DESC """

        log_sh_lists_num = Log.objects.filter(target='cs', target_id=target_id, operator_department='A').count()
        log_sh_lists = Log.objects.all().raw(sql)

        sql = """ SELECT fms_log.*, fms_stepmaster.step_name, fms_actionmaster.action_name, """
        sql = sql + """ [isk_tools_base].[dbo].[auth_user].first_name, [isk_tools_base].[dbo].[auth_user].last_name, """
        sql = sql + """ fms_targetmaster.target_name """
        sql = sql + """ FROM (((fms_log """
        sql = sql + """ LEFT JOIN fms_stepmaster ON fms_log.step=fms_stepmaster.step_id) """
        sql = sql + """ LEFT JOIN fms_actionmaster ON fms_log.action=fms_actionmaster.action_cd) """
        sql = sql + """ LEFT JOIN [isk_tools_base].[dbo].[auth_user] ON fms_log.operator=[isk_tools_base].[dbo].[auth_user].username) """
        sql = sql + """ LEFT JOIN fms_targetmaster ON fms_log.target=fms_targetmaster.target_cd"""
        sql = sql + """ WHERE fms_log.target='cs' AND fms_log.operator_division='KOUMU' AND fms_log.target_id=""" + target_id
        sql = sql + """ ORDER BY fms_log.operation_datetime DESC """

        log_en_lists_num = Log.objects.filter(target='cs', target_id=target_id, operator_division='KOUMU').count()
        log_en_lists = Log.objects.all().raw(sql)

        sql = """ SELECT fms_log.*, fms_stepmaster.step_name, fms_actionmaster.action_name, """
        sql = sql + """ [isk_tools_base].[dbo].[auth_user].first_name, [isk_tools_base].[dbo].[auth_user].last_name, """
        sql = sql + """ fms_targetmaster.target_name """
        sql = sql + """ FROM (((fms_log """
        sql = sql + """ LEFT JOIN fms_stepmaster ON fms_log.step=fms_stepmaster.step_id) """
        sql = sql + """ LEFT JOIN fms_actionmaster ON fms_log.action=fms_actionmaster.action_cd) """
        sql = sql + """ LEFT JOIN [isk_tools_base].[dbo].[auth_user] ON fms_log.operator=[isk_tools_base].[dbo].[auth_user].username) """
        sql = sql + """ LEFT JOIN fms_targetmaster ON fms_log.target=fms_targetmaster.target_cd"""
        sql = sql + """ WHERE fms_log.target='cs' AND (fms_log.step=134001001 OR fms_log.step=134001011 OR fms_log.step=134004001 OR fms_log.step=134008001 OR fms_log.step=134008011) AND fms_log.target_id=""" + target_id
        sql = sql + """ ORDER BY fms_log.operation_datetime DESC """

        log_bd_lists_num = Log.objects.filter(target='cs', target_id=target_id, operator_division='KOUMU').count()
        log_bd_lists = Log.objects.all().raw(sql)

        data = {
            'log_ga_lists': log_ga_lists,
            'log_ev_lists': log_ev_lists,
            'log_sh_lists': log_sh_lists,
            'log_en_lists': log_en_lists,
            'log_bd_lists': log_bd_lists,
            'log_ga_lists_num': log_ga_lists_num,
            'log_ev_lists_num': log_ev_lists_num,
            'log_sh_lists_num': log_sh_lists_num,
            'log_en_lists_num': log_en_lists_num,
            'log_bd_lists_num': log_bd_lists_num,
        }

        return render(request, 'fms/parts/check_sheet/cs_log.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


def get_construction_list(start_date, end_date, fire_flag, high_place_flag, entering_the_tank_flag, heavy_equipment_flag, brake_off_flag, contamination_flag, blockage_flag):

    condition = {
            'construction_date__gte': start_date,
            'construction_date__lte': end_date,
            'lost_flag': 0,
            }

    if fire_flag == 1:
        condition['fires_flag'] = fire_flag
    if high_place_flag == 1:
        condition['high_place_flag'] = high_place_flag
    if entering_the_tank_flag == 1:
        condition['entering_the_tank_flag'] = entering_the_tank_flag
    if heavy_equipment_flag == 1:
        condition['heavy_equipment_flag'] = heavy_equipment_flag
    if brake_off_flag == 1:
        condition['brake_off_flag'] = brake_off_flag
    if contamination_flag == 1:
        condition['contamination_flag'] = contamination_flag
    if blockage_flag == 1:
        condition['blockage_flag'] = blockage_flag

    construction_list_raw = DailyConstruction.objects.filter(**condition)

    i = 1

    for construction_list_raw in construction_list_raw:
        construction_list_raw.display_number = i
        construction_list_raw.save()
        i += 1

    construction_list = DailyConstruction.objects.filter(**condition)

    return (construction_list)


def get_blocking_construction_list(start_date, end_date, fire_flag, high_place_flag, entering_the_tank_flag, heavy_equipment_flag, brake_off_flag, contamination_flag, blockage_flag):

    condition = {
            'construction_date__gte': start_date,
            'construction_date__lte': end_date,
            'lost_flag': 0,
            'blockage_flag': 1,
            }

    if fire_flag == 1:
        condition['fires_flag'] = fire_flag
    if high_place_flag == 1:
        condition['high_place_flag'] = high_place_flag
    if entering_the_tank_flag == 1:
        condition['entering_the_tank_flag'] = entering_the_tank_flag
    if heavy_equipment_flag == 1:
        condition['heavy_equipment_flag'] = heavy_equipment_flag
    if brake_off_flag == 1:
        condition['brake_off_flag'] = brake_off_flag
    if contamination_flag == 1:
        condition['contamination_flag'] = contamination_flag
    if blockage_flag == 1:
        condition['blockage_flag'] = blockage_flag

    blocking_construction_list = DailyConstruction.objects.filter(**condition)

    return (blocking_construction_list)


@require_POST
def daily_construction_display(request):
    try:
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']

        fire_flag = int(request.POST['sel_fire_flag'])
        high_place_flag = int(request.POST['sel_high_place_flag'])
        entering_the_tank_flag = int(request.POST['sel_entering_the_tank_flag'])
        heavy_equipment_flag = int(request.POST['sel_heavy_equipment_flag'])
        brake_off_flag = int(request.POST['sel_brake_off_flag'])
        contamination_flag = int(request.POST['sel_contamination_flag'])
        blockage_flag = int(request.POST['sel_blockage_flag'])

        d_today = datetime.date.today()

        if start_date == '':
            start_date = d_today
        else:
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        if end_date == '':
            end_date = d_today
        else:
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')

        construction_list = get_construction_list(start_date, end_date, fire_flag, high_place_flag, entering_the_tank_flag, heavy_equipment_flag, brake_off_flag, contamination_flag, blockage_flag)

        blocking_construction_list = get_blocking_construction_list(start_date, end_date, fire_flag, high_place_flag, entering_the_tank_flag, heavy_equipment_flag, brake_off_flag, contamination_flag, blockage_flag)

        data = {
            'construction_list': construction_list,
            'blocking_construction_list': blocking_construction_list,
            'start_date': start_date,
            'end_date': end_date,
            'fire_flag': fire_flag,
            'high_place_flag': high_place_flag,
            'entering_the_tank_flag': entering_the_tank_flag,
            'heavy_equipment_flag': heavy_equipment_flag,
            'brake_off_flag': brake_off_flag,
            'contamination_flag': contamination_flag,
            'blockage_flag': blockage_flag,
        }

        return render(request, 'fms/parts/daily_construction/construction_display.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


@require_POST
def construction_data_get(request):
    try:
        target_id = int(request.POST['id'])

        construction_data = DailyConstruction.objects.get(id=target_id)

        if construction_data.scheduled_construction_id is None:
            construction_id = construction_data.small_construction_id
            construction_class = '小口工事'
        else:
            construction_id = construction_data.scheduled_construction_id
            construction_class = '項目工事'
        construction_name = construction_data.construction_name
        construction_date = construction_data.construction_date
        fire_flag = construction_data.fires_flag
        drilling_flag = construction_data.drilling_flag
        blockage_flag = construction_data.blockage_flag
        notification_flag = construction_data.notification_flag
        high_place_flag = construction_data.high_place_flag
        entering_the_tank_flag = construction_data.entering_the_tank_flag
        heavy_equipment_flag = construction_data.heavy_equipment_flag
        brake_off_flag = construction_data.brake_off_flag
        contamination_flag = construction_data.contamination_flag
        rem = construction_data.rem

        data = {
            'construction_id': construction_id,
            'construction_name': construction_name,
            'construction_class': construction_class,
            'construction_date': construction_date,
            'fire_flag': fire_flag,
            'drilling_flag': drilling_flag,
            'blockage_flag': blockage_flag,
            'notification_flag': notification_flag,
            'high_place_flag': high_place_flag,
            'entering_the_tank_flag': entering_the_tank_flag,
            'heavy_equipment_flag': heavy_equipment_flag,
            'brake_off_flag': brake_off_flag,
            'contamination_flag': contamination_flag,
            'rem': rem,
        }

        return render(request, 'fms/parts/daily_construction/construction_detail.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


#@require_POST
def daily_construction_entry_page(request):
    try:
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']

        #fire_flag = request.POST['sel_fire_flag']
        #high_place_flag = request.POST['sel_high_place_flag']
        #entering_the_tank_flag = request.POST['sel_entering_the_tank_flag']
        #heavy_equipment_flag = request.POST['sel_heavy_equipment_flag']
        #brake_off_flag = request.POST['sel_brake_off_flag']
        #contamination_flag = request.POST['sel_contamination_flag']
        fire_flag = 0
        high_place_flag = 0
        entering_the_tank_flag = 0
        heavy_equipment_flag = 0
        brake_off_flag = 0
        contamination_flag = 0
        blockage_flag = 0

        d_today = datetime.date.today()

        year = d_today.year
        month = d_today.month
        last_day = calendar.monthrange(year, month)[1]

        if start_date == '':
            start_date = str(year) + '-' + str(month) + '-01'
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')

        if end_date == '':
            end_date = str(year) + '-' + str(month) + '-' + str(last_day)
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')

        construction_list = get_construction_list(start_date, end_date, fire_flag, high_place_flag, entering_the_tank_flag, heavy_equipment_flag, brake_off_flag, contamination_flag, blockage_flag)

        blocking_construction_list = get_blocking_construction_list(start_date, end_date, fire_flag, high_place_flag, entering_the_tank_flag, heavy_equipment_flag, brake_off_flag, contamination_flag, blockage_flag)

        data = {
            'construction_list': construction_list,
            'blocking_construction_list': blocking_construction_list,
            'start_date': start_date,
            'end_date': end_date,
        }

        return render(request, 'fms/parts/daily_construction/construction_position_entry.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


@require_POST
def construction_entry_data_get(request):
    try:
        target_id = int(request.POST['id'])

        construction_data = DailyConstruction.objects.get(id=target_id)

        if construction_data.scheduled_construction_id is None:
            construction_id = construction_data.small_construction_id
            construction_class = '小口工事'
        else:
            construction_id = construction_data.scheduled_construction_id
            construction_class = '項目工事'
        construction_name = construction_data.construction_name
        construction_date = construction_data.construction_date
        position_x = construction_data.position_x
        position_y = construction_data.position_y
        blockage_position_x = construction_data.blockage_position_x
        blockage_position_y = construction_data.blockage_position_y
        fire_flag = construction_data.fires_flag
        drilling_flag = construction_data.drilling_flag
        blockage_flag = construction_data.blockage_flag
        notification_flag = construction_data.notification_flag
        high_place_flag = construction_data.high_place_flag
        entering_the_tank_flag = construction_data.entering_the_tank_flag
        heavy_equipment_flag = construction_data.heavy_equipment_flag
        brake_off_flag = construction_data.brake_off_flag
        contamination_flag = construction_data.contamination_flag
        rem = construction_data.rem

        data = {
            'id': target_id,
            'construction_id': construction_id,
            'construction_name': construction_name,
            'construction_class': construction_class,
            'construction_date': construction_date,
            'position_x': position_x,
            'position_y': position_y,
            'blockage_position_x': blockage_position_x,
            'blockage_position_y': blockage_position_y,
            'fire_flag': fire_flag,
            'drilling_flag': drilling_flag,
            'blockage_flag': blockage_flag,
            'notification_flag': notification_flag,
            'high_place_flag': high_place_flag,
            'entering_the_tank_flag': entering_the_tank_flag,
            'heavy_equipment_flag': heavy_equipment_flag,
            'brake_off_flag': brake_off_flag,
            'contamination_flag': contamination_flag,
            'rem': rem,
        }

        return render(request, 'fms/parts/daily_construction/construction_entry.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


@require_POST
def position_data_entry(request):
    try:
        target_id = int(request.POST['target_id'])
        construction_date = request.POST['construction_date']
        position_x = request.POST['position_x']
        position_y = request.POST['position_y']
        blockage_position_x = request.POST['blockage_position_x']
        blockage_position_y = request.POST['blockage_position_y']
        blockage_flag = int(request.POST['blockage_flag'])

        construction_data = DailyConstruction.objects.get(id=target_id)
        if construction_date != 'None' and construction_date != '':
            construction_data.construction_date = construction_date
        if position_x != 'None' and position_x != '':
            construction_data.position_x = position_x
        if position_y != 'None' and position_y != '':
            construction_data.position_y = position_y
        if blockage_position_x != 'None' and blockage_position_x != '':
            construction_data.blockage_position_x = blockage_position_x
        if blockage_position_y != 'None' and blockage_position_y != '':
            construction_data.blockage_position_y = blockage_position_y
        construction_data.blockage_flag = blockage_flag
        construction_data.save()

        msg = '登録完了！！'

        ary = {
            'msg': msg,
        }

        return JsonResponse(ary)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


#@login_required
def file_import(request):
    try:
        # ベースディレクトリ
        if gethostname() == 'YWEBSERV1': #本番
            base_dir = 'D:\\python_tool_development\\isk-tools'
        elif gethostname() == 'I7161DD6': #テスト
            base_dir = 'C:\\python_tool_development\\isk-tools'
        else:
            # base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'
            base_dir = 'C:\\python_tool_development\\isk-tools'
        UPLOADE_DIR = base_dir + '\\static\\files\\fms\\daily_construction\\'

        if request.method != 'POST':

            msg = "アップロードできませんでした！！"

        else:
            file = request.FILES['file']
            path = os.path.join(UPLOADE_DIR, file.name)
            destination = open(path, 'wb')

            for chunk in file.chunks():
                destination.write(chunk)

            destination.close()

            wb = openpyxl.load_workbook(path)

            # sheets_num = len(wb.worksheets)

            j = 0
            k = 0

            for ws in wb.worksheets:
                if '日' in ws.title:
                    for i in range(8, 30, 2):
                        if ws.cell(row=i, column=5).value is not None and ws.cell(row=i, column=5).value != "":
                            new_continuation_item = ws.cell(row=i, column=1).value
                            construction_process = ws.cell(row=i, column=2).value
                            construction_name = ws.cell(row=i, column=5).value
                            construction_detail = ws.cell(row=i+1, column=5).value
                            construction_constructor = ws.cell(row=i, column=6).value
                            if ws.cell(row=i, column=7).fill == PatternFill(fill_type=None):
                                construction_fire_flag = 0
                            else:
                                construction_fire_flag = 1
                            if ws.cell(row=i, column=8).fill == PatternFill(fill_type=None):
                                construction_high_place_flag = 0
                            else:
                                construction_high_place_flag = 1
                            if ws.cell(row=i, column=9).fill == PatternFill(fill_type=None):
                                construction_entering_the_tank_flag = 0
                            else:
                                construction_entering_the_tank_flag = 1
                            if ws.cell(row=i, column=10).fill == PatternFill(fill_type=None):
                                construction_heavy_equipment_flag = 0
                            else:
                                construction_heavy_equipment_flag = 1
                            if ws.cell(row=i, column=11).fill == PatternFill(fill_type=None):
                                construction_brake_off_flag = 0
                            else:
                                construction_brake_off_flag = 1
                            if ws.cell(row=i, column=12).fill == PatternFill(fill_type=None):
                                construction_contamination_flag = 0
                            else:
                                construction_contamination_flag = 1
                            construction_charge_person = ws.cell(row=i, column=13).value
                            construction_data = ws.cell(row=4, column=6).value

                            if construction_name != "":
                                import_data, created = DailyConstruction.objects.get_or_create(construction_date=construction_data, construction_name=construction_name)
                                import_data.new_continuation_item = new_continuation_item
                                import_data.construction_class = 3
                                import_data.process = construction_process
                                import_data.detail = construction_detail
                                import_data.constructor = construction_constructor
                                import_data.fires_flag = construction_fire_flag
                                import_data.high_place_flag = construction_high_place_flag
                                import_data.entering_the_tank_flag = construction_entering_the_tank_flag
                                import_data.heavy_equipment_flag = construction_heavy_equipment_flag
                                import_data.brake_off_flag = construction_brake_off_flag
                                import_data.contamination_flag = construction_contamination_flag
                                import_data.charge_person = construction_charge_person
                                import_data.lost_flag = 0

                                import_data.save()

                                if created == 1:
                                    k += 1
                                else:
                                    j += 1
                        else:
                            break

            msg = str(j) + "件更新登録、" + str(k) + "件新規登録！！"

        ary = {
            'msg': msg
        }

        return JsonResponse(ary)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise
