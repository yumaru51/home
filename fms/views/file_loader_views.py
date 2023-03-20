import openpyxl
import mimetypes
import traceback
from django.utils.encoding import smart_str
from django.views.decorators.http import require_GET, require_POST
from django.http import FileResponse
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.http.response import JsonResponse
import os
import datetime
from cpexcel.models import Tasks
from fms.models import AttachmentDocuments, DataEntryStepMaster
from fms.views.common_def_views import get_output_file_path
from fms.views.common_def_views import clear_input_file_folder, get_input_file_path, get_input_file_base_path
import datetime
from socket import gethostname
# 20210114 y-kawauchi 絞り込み機能対応
from django.db.models import Count, Q
from django.utils.timezone import make_aware
from fms.views.common_def_views import output_log_info, output_log_error, output_log_exception


@login_required
@require_GET
def file_download(request, data_id):
    try:
        attachment_documents = AttachmentDocuments.objects.get(id=data_id, lost_flag=0)

        # ベースディレクトリ
        if gethostname() == 'YWEBSERV1': #本番
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\Production'
        elif gethostname() == 'I7161DD6': #テスト
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'
        else:
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'

        # 故障対応と計画策定・実行versionでfile_path仕様分岐
        if attachment_documents.data is None:
            file_full_path = base_dir + attachment_documents.folder + '\\' + attachment_documents.file_name
            # file_full_path = base_dir + attachment_documents.folder + attachment_documents.file_name
        else:
            file_full_path = base_dir + attachment_documents.folder + attachment_documents.data + '\\' + attachment_documents.file_name

        # 20201212 y-kawauchi ファイル読込変更
        if os.path.isfile(file_full_path) is False:
            msg = '対象ファイルが存在しません！！'

        with open(file_full_path, 'rb') as fh:
            # response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            # response = HttpResponse(fh.read())
            response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(attachment_documents.file_name)[0] or 'application/octet-stream')
            response['Content-Disposition'] = 'inline; filename=' + attachment_documents.file_name

        return response
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

@login_required
@require_GET
def attached_file_download(request, data_id, file_name):
    try:

        file_full_path = '\\\\Ysqlserv4\\apps\\小口工事管理\\IEP\\機器管理添付ファイル\\' + str(data_id) + '\\' + file_name

        if os.path.isfile(file_full_path) is False:
            msg = '対象ファイルが存在しません！！'
        else:
            with open(file_full_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(file_name) or 'application/octet-stream')
                response['Content-Disposition'] = 'inline; filename=' + file_name

        return response
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

@login_required
@require_POST
def file_delete(request):
    try:
        attachment_documents = AttachmentDocuments.objects.get(id=request.POST['data_id'], lost_flag=0)

        file_folder = attachment_documents.folder
        div_id_name = attachment_documents.div_id_name
        file_class = attachment_documents.data
        delete_pb_disp_flag = 1

        # ベースディレクトリ
        if gethostname() == 'YWEBSERV1': #本番
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\Production'
        elif gethostname() == 'I7161DD6': #テスト
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'
        else:
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'
        # file_full_path = base_dir + attachment_documents.folder + attachment_documents.data + '\\' + attachment_documents.file_name
        # base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'

        # if attachment_documents.data is None:
        if attachment_documents.data == 'AbnormalityReport' or attachment_documents.data == 'CorrespondencePolicyDecision' or attachment_documents.data == 'LaboratoryDiagnosis' or attachment_documents.data == 'Evaluation' or attachment_documents.data == 'MaintenanceEstimate' or attachment_documents.data == 'InspectionAcceptance':
            file_full_path = base_dir + attachment_documents.folder + attachment_documents.data + '\\' + attachment_documents.file_name
            temporary_flag = 0
        else:
            file_full_path = base_dir + attachment_documents.folder + attachment_documents.data + '\\' + attachment_documents.file_name
            temporary_flag = 1

        # ファイル削除、レコード削除
        if os.path.isfile(file_full_path) is True:
            os.remove(file_full_path)
        attachment_documents.lost_flag = 1
        attachment_documents.save()
        # attachment_documents.delete()

        ary = {
            'msg': 'データを削除しました',
            'file_folder': file_folder,
            'div_id_name': div_id_name,
            'delete_pb_disp_flag': delete_pb_disp_flag,
            'file_class': file_class,
            # 削除機能後のリスト更新一時対応　処理が分かれているためFLAGで処理を分ける
            'temporary_flag': temporary_flag,
        }
        return JsonResponse(ary)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

# 共通処理のため削除予定
@login_required
@require_POST
def file_delete_other_type(request):
    try:
        attachment_documents = AttachmentDocuments.objects.get(id=request.POST['data_id'], lost_flag=0)

        file_folder = attachment_documents.folder
        div_id_name = attachment_documents.div_id_name
        delete_pb_disp_flag = 1

        # ベースディレクトリ
        if gethostname() == 'YWEBSERV1': #本番
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\Production'
        elif gethostname() == 'I7161DD6': #テスト
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'
        else:
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'

        if attachment_documents.data is None:
            file_full_path = base_dir + attachment_documents.folder + '\\' + attachment_documents.file_name
        else:
            file_full_path = base_dir + attachment_documents.folder + attachment_documents.data + '\\' + attachment_documents.file_name

        # ファイル削除、レコード削除
        if os.path.isfile(file_full_path) is True:
            os.remove(file_full_path)
        attachment_documents.lost_flag = 1
        attachment_documents.save()
        # attachment_documents.delete()

        ary = {
            'msg': 'データを削除しました',
            'file_folder': file_folder,
            'div_id_name': div_id_name,
            'delete_pb_disp_flag': delete_pb_disp_flag,
        }
        return JsonResponse(ary)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

@login_required
@require_POST
def file_upload(request):
    try:
        # 日時フォーマット対応　'YYYYMMDD'
        # dt_now = datetime.datetime.now()
        dt_now_naive = datetime.datetime.now()
        dt_now = make_aware(dt_now_naive)
        str_date = datetime.datetime.now().strftime('%Y%m%d')
        t_username = request.user.username
        file_target = request.POST['file_target']
        file_budget_id = request.POST['file_budget_id']
        file_work_id = request.POST['file_work_id']
        div_id_name = request.POST['div_id_name']

        # ※idの入力チェックはJS側で行うこと

        # ベースディレクトリ
        if gethostname() == 'YWEBSERV1': #本番
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\Production'
        elif gethostname() == 'I7161DD6': #テスト
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'
        else:
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'

        # ファイルが1件もない、または新規画面の時　アップロード不可
        if request.FILES.__len__() == 0:
            msg = "ファイルが選択されていません！"
        else:
            # ファイルが複数ある場合1ファイルずつ処理を行う。
            for file_list in request.FILES.dict():
                file = request.FILES[file_list]
                file_name = str_date + "_" + file.name

            # アップロードファイルパス
            upload_file_path = base_dir + '\\' + file_target + '\\' + str(file_budget_id) + '\\' + str(file_work_id) + '\\' + file_list + '\\'
            # 存在チェック　なければフォルダ作成
            # if os.path.exists(upload_file_path) != True:
            if os.path.exists(upload_file_path) is not True:
                os.makedirs(upload_file_path, exist_ok=True)
            path = os.path.join(upload_file_path, file_name)

            with open(path, 'wb') as f:
                for chunk in file.chunks():
                    f.write(chunk)

            attachment_documents_data, created = AttachmentDocuments.objects.get_or_create(file_name=file_name, folder='\\' + file_target + '\\' + str(file_budget_id) + '\\' + str(file_work_id) + '\\')
            attachment_documents_data.div_id_name = div_id_name
            attachment_documents_data.data = file_list
            attachment_documents_data.entry_time = dt_now
            attachment_documents_data.entry_user = t_username
            attachment_documents_data.lost_flag = 0
            attachment_documents_data.save()

            msg = "アップロード完了！！"

        ary = {
            'msg': msg,
            'div_id_name': div_id_name,
        }
        return JsonResponse(ary)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

@login_required
@require_POST
def file_upload_other_type(request):
    try:
        dt_now = datetime.datetime.now()
        dt_now_naive = datetime.datetime.now()
        dt_now = make_aware(dt_now_naive)
        # str_year = str(dt_now.year)
        # str_month = str(dt_now.month)
        # str_day = str(dt_now.day)
        # str_date = str_year + str_month + str_day
        str_date = datetime.datetime.now().strftime('%Y%m%d')
        t_username = request.user.username
        file_folder = request.POST['file_folder']
        div_id_name = request.POST['div_id_name']
        delete_pb_disp_flag = int(request.POST['delete_pb_disp_flag'])
        check_item_value = request.POST['check_item']
        file_target = request.POST['file_target']
        file_first_layer_id = request.POST['file_first_layer_id']
        file_second_layer_id = request.POST['file_second_layer_id']
        file_class = request.POST['file_class']

        # ベースディレクトリ
        # base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'
        if gethostname() == 'YWEBSERV1': #本番
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\Production'
        elif gethostname() == 'I7161DD6': #テスト
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'
        else:
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'

        # ファイルが1件もない時　アップロード不可
        if request.FILES.__len__() == 0:
            msg = "ファイルをアップロードできませんでした！"
            target_file_folder = file_folder
        elif check_item_value == "" or request.POST['file_first_layer_id'] == '0' or request.POST['file_first_layer_id'] == '':
            msg = "案件を先に保存してください！"
            target_file_folder = file_folder
        else:
            file = request.FILES['file']

            file_name = str_date + "_" + file.name
            # アップロードファイルパス
            # upload_file_path = base_dir + file_folder + '\\'
            if file_second_layer_id is not "":
                upload_file_path = base_dir + '\\' + file_target + '\\' + str(file_first_layer_id) + '\\' + str(file_second_layer_id) + '\\' + file_class + '\\'
            else:
                upload_file_path = base_dir + '\\' + file_target + '\\' + str(file_first_layer_id) + '\\' + file_class + '\\'


            # 存在チェック　なければフォルダ作成
            # if os.path.exists(upload_file_path) != True:
            #     os.makedirs(upload_file_path)
            if os.path.exists(upload_file_path) is not True:
                os.makedirs(upload_file_path, exist_ok=True)
            path = os.path.join(upload_file_path, file_name)

            with open(path, 'wb') as f:
                for chunk in file.chunks():
                    f.write(chunk)

            if file_second_layer_id is not "":
                target_file_folder = '\\' + file_target + '\\' + str(file_first_layer_id) + '\\' + str(file_second_layer_id) + '\\' + file_class + '\\'
            else:
                target_file_folder = '\\' + file_target + '\\' + str(file_first_layer_id) + '\\'

            attachment_documents_data, created = AttachmentDocuments.objects.get_or_create(folder=target_file_folder, data=file_class, file_name=file_name)
            attachment_documents_data.div_id_name = div_id_name
            attachment_documents_data.data = file_class
            attachment_documents_data.entry_time = dt_now
            attachment_documents_data.entry_user = t_username
            attachment_documents_data.lost_flag = 0
            attachment_documents_data.save()

            msg = "アップロード完了！！"

        ary = {
            'msg': msg,
            'file_folder': target_file_folder,
            'div_id_name': div_id_name,
            'delete_pb_disp_flag': delete_pb_disp_flag,
            'file_class': file_class,
        }
        return JsonResponse(ary)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

@login_required
@require_POST
def file_upload_file_list(request):
    try:
        # ここでfolderをrequestできるようにする。

        file_target = request.POST['file_target']
        file_budget_id = request.POST['file_budget_id']
        file_work_id = request.POST['file_work_id']
        div_id_name = request.POST['div_id_name']
        current_tab = request.POST['current_tab']

        if div_id_name == 'budget':
            file_work_id = 0
            file_target = 'budget'

        # 格納先path作成　idを汎用的にするため「level4_id」「level5_id」としたい。
        file_folder = '\\' + file_target + '\\' + str(file_budget_id) + '\\' + str(file_work_id) + '\\'

        # 「upload_file_list.html」絞り込み機能表示
        # uploadfile_list_num = AttachmentDocuments.objects.filter(folder=file_folder, div_id_name=div_id_name, lost_flag=0).count()
        uploadfile_list_num = AttachmentDocuments.objects.filter(folder=file_folder,
                                                                 div_id_name=div_id_name,
                                                                 lost_flag=0
                                                                ).values('data').annotate(total=Count('data')).count()
        if current_tab != 'execution_specification':
            uploadfile_list = AttachmentDocuments.objects.filter(folder=file_folder, div_id_name=div_id_name, lost_flag=0)
        else:
            uploadfile_list_num = AttachmentDocuments.objects.filter(folder=file_folder,
                                                                     div_id_name=div_id_name,
                                                                     lost_flag=0
                                                                     ).values('data').annotate(total=Count('data')).count()

            uploadfile_list = AttachmentDocuments.objects.filter(Q(folder=file_folder),
                                                                 Q(div_id_name=div_id_name),
                                                                 Q(lost_flag=0)
                                                                )
            if uploadfile_list_num > 0:
                test = 1

        data_list = AttachmentDocuments.objects.filter(folder=file_folder,
                                                       div_id_name=div_id_name,
                                                       lost_flag=0
                                                       ).values('data').order_by('data').distinct()

        # 20210114 y-kawauchi 削除機能FLAG追加
        # 画面が編集可能状態のときのみ削除も可とする。「dataentrystepmaster」にデータがなければ編集不可なので削除も不可、あれば編集可なので削除も可
        if DataEntryStepMaster.objects.filter(step_id=request.POST['this_step'],
                                              target_table=request.POST['div_id_name'],
                                              lost_flag=0).count() == 0:
            delete_pb_disp_flag = 0
        else:
            delete_pb_disp_flag = 1

        data = {
            'uploadfile_list_num': uploadfile_list_num,
            'uploadfile_list': uploadfile_list,
            'div_id_name': div_id_name,
            'delete_pb_disp_flag': delete_pb_disp_flag,
            'data_list': data_list,
        }

        return render(request, 'fms/parts/common/upload_file_list.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

@login_required
@require_POST
def file_upload_file_list_other_type(request):
    try:
        file_folder = request.POST['file_folder']
        div_id_name = request.POST['div_id_name']
        delete_pb_disp_flag = int(request.POST['delete_pb_disp_flag'])
        file_class = request.POST['file_class']

        # ベースディレクトリ
        if gethostname() == 'YWEBSERV1': #本番
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\Production'
        elif gethostname() == 'I7161DD6': #テスト
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'
        else:
            base_dir = '\\\\Ydomnserv\\common\\部門間フォルダ\\FacilityData\\test'

        upload_file_path = base_dir + '\\' + file_folder + '\\' + file_class + '\\'

        uploadfile_list_num = AttachmentDocuments.objects.filter(folder=file_folder, data=file_class, lost_flag=0).count()

        uploadfile_list = AttachmentDocuments.objects.filter(folder=file_folder, data=file_class, lost_flag=0)

        # ネットワークドライブからファイル一覧を取得する

        # file_folderからphenomenon_idを取得する
        phenomenon_id = file_folder.replace('\\FailureCorrespondence\\', '')
        phenomenon_id = phenomenon_id.replace('\\', '')
        attached_file_dir = '\\\\Ysqlserv4\\apps\\小口工事管理\\IEP\\機器管理添付ファイル\\' + phenomenon_id

        attached_file_list = ""
        if os.path.exists(attached_file_dir):
            files = os.listdir(attached_file_dir)
            attached_file_list = [f for f in files if os.path.isfile(os.path.join(attached_file_dir, f))]

        attached_file_list_num = len(attached_file_list)

        data = {
            'uploadfile_list_num': uploadfile_list_num,
            'uploadfile_list': uploadfile_list,
            'delete_pb_disp_flag': delete_pb_disp_flag,
            'phenomenon_id': phenomenon_id,
            'attached_file_list_num': attached_file_list_num,
            'attached_file_list': attached_file_list,
        }

        return render(request, 'fms/parts/common/upload_file_list.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

@login_required
def file_import(request):
    try:
        DIFF_JST_FROM_UTC = 18
        # JST = timezone(timedelta(hours=+9), 'JST')

        # now = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)
        now_naive = datetime.datetime.now()
        now = make_aware(now_naive)

        t_username = request.user.username

        # BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        BASE_DIR = 'C:\\PythonProjects\\isk_tools\\'
        # UPLOADE_DIR = os.path.dirname(os.path.abspath(__file__)) + '/static/files/' + t_username + '/'
        UPLOADE_DIR = os.path.join(BASE_DIR, 'static\\files\\cpexcel\\' + t_username + '\\')
        if request.method != 'POST':
            # return render(request, 'upload_form/form.html')
            msg = "アップロードできませんでした！！"

        else:
            file = request.FILES['file']
            path = os.path.join(UPLOADE_DIR, file.name)
            destination = open(path, 'wb')

            for chunk in file.chunks():
                destination.write(chunk)

            destination.close()

            wb = openpyxl.load_workbook(path)

            data_sheet = wb['DATA']

            row_num = data_sheet.max_row

            msg = ""

            if data_sheet.cell(row=1, column=2).value != "":
                program_id = data_sheet.cell(row=1, column=2).value

                j = 0
                k = 0

                for i in range(5, row_num+1):

                    if data_sheet.cell(row=i, column=1).value == "record":

                        source_sheet_name = data_sheet.cell(row=i, column=3).value
                        source_cell_name = data_sheet.cell(row=i, column=4).value
                        post_to_sheet_name = data_sheet.cell(row=i, column=5).value
                        post_to_cell_name = data_sheet.cell(row=i, column=6).value
                        lost_flag = data_sheet.cell(row=i, column=7).value

                        if data_sheet.cell(row=i, column=2).value is None:

                            Tasks(program_id=program_id, source_sheet_name=source_sheet_name, source_cell_name=source_cell_name, post_to_sheet_name=post_to_sheet_name, post_to_cell_name=post_to_cell_name, entry_date=now, entry_operator=t_username, lost_flag=0).save()

                            k += 1

                        else:

                            task_id = data_sheet.cell(row=i, column=2).value
                            task_data = Tasks.objects.get(id=task_id)
                            task_data.source_sheet_name = source_sheet_name
                            task_data.source_cell_name = source_cell_name
                            task_data.post_to_sheet_name = post_to_sheet_name
                            task_data.post_to_cell_name = post_to_cell_name
                            task_data.lost_flag = lost_flag
                            task_data.update_date = now
                            task_data.update_operator = t_username
                            task_data.save()

                            j += 1

                msg = str(j) + "件更新登録、" + str(k) + "件新規登録！！"

            else:
                msg = "プログラムidが指定されていないため、登録不可！！"

        ary = {
            'msg': msg
        }

        return JsonResponse(ary)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

# コード参照のみ不使用のため削除予定
# 2021/01/01 ueda 追加
@login_required
@require_POST
def get_upload_file_list(request):
    try:
        file_folder = request.POST['file_folder']
        delete_pb_disp_flag = int(request.POST['delete_pb_disp_flag'])

        upload_file_list_num = AttachmentDocuments.objects.filter(folder=file_folder, lost_flag=0).count()

        upload_file_list = AttachmentDocuments.objects.filter(folder=file_folder, lost_flag=0)

        data = {
            'upload_file_list_num': upload_file_list_num,
            'upload_file_list': upload_file_list,
            'delete_pb_disp_flag': delete_pb_disp_flag,
        }

        return render(request, 'fms/parts/common/upload_file_list.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

# 出力ファイル取得
@login_required
@require_GET
def output_file_download(request, target_folder, file_name):
    try:
        file_path = get_output_file_path(target_folder, file_name)
        if os.path.isfile(file_path) is False:
            msg = '対象ファイルが存在しません！！'
        else:
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(file_name) or 'application/octet-stream')
                response['Content-Disposition'] = 'inline; filename=' + file_name

        return response
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

# サーバーファイル取得
@login_required
@require_GET
def server_file_download(request):
    target_path = request.GET['target_path']
    file_name = request.GET['file_name']
    file_path = target_path + r"/" + file_name

    if os.path.isfile(file_path) is False:
        print('file_name', file_name)
        print('target_path', target_path)
        print('対象ファイルが存在しません！！')
        # ステータスコードを指定
        response = HttpResponse(status=404)
    else:
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(file_name) or 'application/octet-stream')
            response['Content-Disposition'] = 'inline; filename=' + file_name

    return response

@login_required
@require_POST
def input_file_upload(request):
    try:
        if request.FILES.__len__() == 0:
            msg = "ファイルが選択されていません！"
            ary = {
                'msg': msg,
            }
            return JsonResponse(ary)

        file = request.FILES['file']
        target_folder = request.POST['target_folder']
        file_folder_path = get_input_file_base_path(target_folder)
        file_full_path = get_input_file_path(target_folder, file.name)
        # 存在チェック　なければフォルダ作成
        if os.path.exists(file_folder_path) is not True:
            os.makedirs(file_folder_path, exist_ok=True)

        # フォルダ内のファイル削除
        clear_input_file_folder(target_folder)

        # ファイル保存
        with open(file_full_path, 'wb') as f:
            for chunk in file.chunks():
                f.write(chunk)

        msg = "アップロード完了！！"
        ary = {
            'msg': msg,
        }
        return JsonResponse(ary)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

