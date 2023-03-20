import datetime
import time
import traceback
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_POST, require_GET
from django.db.models import Q
from common.common_def import date_to_many_type
from fms.views.common_def_views import get_template_file_path, get_output_file_path
from fms.views.common_def_views import clear_output_file_folder
from fms.views.common_def_views import output_log_info, output_log_error, output_log_exception
from fms.views.common_def_views import clear_input_file_folder, get_input_file_path, get_input_file_base_path
from fms.views.common_def_views import get_operator_permission
from fms.views.budget_views import budget_no_registration, budget_no_registration_list
from fms.views.common_def_views import output_log_exception
import openpyxl
from config.settings.settings_common import BASE_DIR
import mimetypes
from fms.models import MaintenanceEquipment
from fms.models.temporary_models import TemporaryMakePlantiaImport
from plantia.models import MasterLocation, FcltyLdgr, ServiceUser
from fms.models import Budget, BudgetLaw
from fms.models import ProSpecificationUnit
from fms.models import DepartmentMaster
from fms.models import UserAttribute, User
from fms.models import EquipmentHistoryReport, Phenomenon
from fms.models import DivisionMaster
from fms.models import PurposeClassMaster, ProcessMaster, BudgetClassMaster, PeriodClassMaster
from plantia.models import MasterMgtCls

from openpyxl.styles.borders import Border, Side
from django.utils.timezone import make_aware


def temporary_response_menu(request):
    try:
        # サイドメニュー用継承データ取得
        t_username = request.user.username
        user_division_cd = request.POST['user_division_cd']
        user_department_cd = request.POST['user_department_cd']
        user_authority = int(request.POST['user_authority'])
        confirm_user = request.POST['confirm_user']
        permit_user = request.POST['permit_user']
        plantia_import_permission_list = TemporaryMakePlantiaImport.objects.filter(user_id=t_username, lost_flag=0)

        data = {
            'user_division_cd': user_division_cd,
            'user_department_cd': user_department_cd,
            'user_authority': user_authority,
            'confirm_user': confirm_user,
            'permit_user': permit_user,
            'plantia_import_permission_list': plantia_import_permission_list,
            'operator_permission_list': get_operator_permission(t_username),
        }

        return render(request, 'fms/parts/temporary_response/temporary_response_menu.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


def make_ordered_construction_excel_page(request):
    try:
        user_division_cd = request.POST['user_division_cd']
        user_department_cd = request.POST['user_department_cd']
        user_authority = int(request.POST['user_authority'])
        confirm_user = request.POST['confirm_user']
        permit_user = request.POST['permit_user']

        departments_list = DepartmentMaster.objects.filter(lost_flag=0).all().order_by('display_order')
        data = {
            'user_division_cd': user_division_cd,
            'user_department_cd': user_department_cd,
            'user_authority': user_authority,
            'confirm_user': confirm_user,
            'permit_user': permit_user,
            'departments_list': departments_list,
        }

        return render(request, 'fms/parts/temporary_response/make_ordered_construction_excel.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

# 台帳問い合わせexcel作成
@login_required
@require_GET
def make_ordered_construction_excel(request, department, business_year, start_date, end_date):
    try:
        DIFF_JST_FROM_UTC = 9
        now_naive = datetime.datetime.now()
        now = make_aware(now_naive)

        date_str = date_to_many_type(now)
        today_str = date_str.str_type_date_jp

        start_datetime = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)

        operator = request.user.username

        # templates_file_full_path = BASE_DIR + '\\static\\files\\fms\\temporary_response\\ordered_construction_excel.xlsx'
        templates_file_full_path = '\\\\ydomnserv\\common\\部門間フォルダ\\FacilityData\\template_files\\ordered_construction\\ordered_construction_excel.xlsx'

        wb_new_file = openpyxl.load_workbook(templates_file_full_path)

        ws_new_sheet = wb_new_file['Sheet1']

        sql = """SELECT fms_prospecificationunit.*, fms_budget.budget_no as bdg_no, fms_departmentmaster.department_name, fms_suppliermaster.supplier_name"""
        # sql = sql + """, fms_proestimates.confirmed_last_estimated_amount, fms_proestimates.confirmed_estimated_amount_af_nego"""
        sql = sql + """, fms_mcframe.total_price, fms_mcframe.discount_price"""
        sql = sql + """, fms_mcframe.purchase_order_no"""

        sql = sql + """, fms_mcframe.numbering"""
        sql = sql + """, fms_mcframe.year"""
        sql = sql + """, fms_mcframe.detail_no"""

        sql = sql + """, CASE fms_proestimates.confirmed_vendor"""
        sql = sql + """ WHEN fms_proestimates.candidate_vendor1 THEN fms_proestimates.delivery_date1_to"""
        sql = sql + """ WHEN fms_proestimates.candidate_vendor2 THEN fms_proestimates.delivery_date2_to"""
        sql = sql + """ WHEN fms_proestimates.candidate_vendor3 THEN fms_proestimates.delivery_date3_to"""
        sql = sql + """ WHEN fms_proestimates.candidate_vendor4 THEN fms_proestimates.delivery_date4_to"""
        sql = sql + """ WHEN fms_proestimates.candidate_vendor5 THEN fms_proestimates.delivery_date5_to"""
        sql = sql + """ END as delivery_date_to"""
        sql = sql + """ FROM ((((fms_prospecificationunit"""
        sql = sql + """ LEFT JOIN fms_budget ON fms_prospecificationunit.budget_id=fms_budget.budget_id)"""
        sql = sql + """ LEFT JOIN fms_proindividualcontractdoc ON fms_prospecificationunit.construction_id=fms_proindividualcontractdoc.construction_id)"""
        sql = sql + """ LEFT JOIN fms_departmentmaster ON fms_prospecificationunit.department=fms_departmentmaster.department_cd)"""
        sql = sql + """ LEFT JOIN fms_suppliermaster ON fms_proindividualcontractdoc.vendor_code=fms_suppliermaster.supplier_cd)"""
        sql = sql + """ LEFT JOIN fms_proestimates ON fms_prospecificationunit.construction_id=fms_proestimates.construction_id"""
        sql = sql + """ LEFT JOIN fms_mcframe ON fms_prospecificationunit.construction_id = fms_mcframe.order_id"""
        # sql = sql + """ WHERE fms_mcframe.purchase_order_no Is Not Null AND fms_mcframe.purchase_order_no <> '' AND fms_proindividualcontractdoc.lost_flag =0 AND fms_proestimates.lost_flag =0 AND fms_prospecificationunit.lost_flag =0"""
        sql = sql + """ WHERE (fms_mcframe.status Is Not Null AND fms_mcframe.status <> 4) AND fms_proindividualcontractdoc.lost_flag =0 AND fms_proestimates.lost_flag =0 AND fms_prospecificationunit.lost_flag =0"""
        sql = sql + """ AND fms_mcframe.work_class <> 'SW' """

        if department != "NON":
            sql = sql + """ AND fms_prospecificationunit.department='""" + department + """'"""
        if business_year != "NON":
            sql = sql + """ AND fms_budget.business_year_id=""" + business_year
        if start_date != "NON":
            # date_str = date_to_many_type(start_date)
            # start_date = date_str.str_type_date_erp
            start_date = start_date.replace('-', '')
            sql = sql + """ AND CONVERT(int, fms_mcframe.order_date)>=""" + str(start_date)
            # sql = sql + """ AND fms_mcframe.order_date>=""" + str(start_date)
        if end_date != "NON":
            # date_str = date_to_many_type(start_date)
            # end_date = date_str.str_type_date_erp
            end_date = end_date.replace('-', '')
            sql = sql + """ AND CONVERT(int, fms_mcframe.order_date)<=""" + str(end_date)
            # sql = sql + """ AND fms_mcframe.order_date<=""" + str(end_date)

        construction_data_list = ProSpecificationUnit.objects.raw(sql)

        i = 2

        for construction_data_list in construction_data_list:
            ws_new_sheet.cell(row=i, column=1).value = construction_data_list.budget_id  # 予算ID
            ws_new_sheet.cell(row=i, column=2).value = construction_data_list.numbering  # 発行NO
            ws_new_sheet.cell(row=i, column=3).value = construction_data_list.year  # 年度
            ws_new_sheet.cell(row=i, column=4).value = construction_data_list.bdg_no  # 予算NO
            ws_new_sheet.cell(row=i, column=5).value = construction_data_list.department_name  # 部署名
            ws_new_sheet.cell(row=i, column=6).value = construction_data_list.supplier_name  # 発注先名
            ws_new_sheet.cell(row=i, column=7).value = construction_data_list.total_price  # 見積額
            ws_new_sheet.cell(row=i, column=8).value = 0  # 出精値引
            # ws_new_sheet.cell(row=i, column=9).value = construction_data_list.  # 再見積額
            ws_new_sheet.cell(row=i, column=10).value = construction_data_list.discount_price  # 一次ネゴ
            ws_new_sheet.cell(row=i, column=13).value = construction_data_list.discount_price  # ネゴ計
            ws_new_sheet.cell(row=i, column=11).value = 0  # 二次ネゴ
            ws_new_sheet.cell(row=i, column=12).value = 0  # 三次ネゴ
            if construction_data_list.total_price is not None and construction_data_list.discount_price is not None:
                ws_new_sheet.cell(row=i, column=14).value = construction_data_list.total_price - construction_data_list.discount_price  # 発注額
            ws_new_sheet.cell(row=i, column=15).value = construction_data_list.construction_id  # 工事ID
            # ws_new_sheet.cell(row=i, column=16).value = today_str  # 工事No
            ws_new_sheet.cell(row=i, column=17).value = construction_data_list.work_name  # 工事項目名
            ws_new_sheet.cell(row=i, column=18).value = construction_data_list.delivery_date_to  # 工期
            ws_new_sheet.cell(row=i, column=19).value = construction_data_list.purchase_order_no  # 注文No
            ws_new_sheet.cell(row=i, column=20).value = construction_data_list.detail_no  # 明細番号

            i += 1

        side = Side(style='thin', color='000000')

        # set border (black thin line)
        border = Border(top=side, bottom=side, left=side, right=side)

        # write in sheet
        for row in ws_new_sheet:
            for cell in row:
                ws_new_sheet[cell.coordinate].border = border

        new_file_name = 'result_ordered_construction_excel.xlsx'

        # new_file_name_fullpath = BASE_DIR + '\\static\\files\\fms\\temporary_response\\' + new_file_name
        new_file_name_fullpath = '\\\\ydomnserv\\common\\部門間フォルダ\\FacilityData\\template_files\\ordered_construction\\' + new_file_name

        wb_new_file.save(new_file_name_fullpath)

        with open(new_file_name_fullpath, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(new_file_name)[0] or 'application/octet-stream')
            response['Content-Disposition'] = 'inline; filename=' + new_file_name

        return response
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


@login_required
@require_POST
def make_plantia_import_excel_page(request):
    try:
        user_division_cd = request.POST['user_division_cd']
        user_department_cd = request.POST['user_department_cd']
        user_authority = int(request.POST['user_authority'])
        confirm_user = request.POST['confirm_user']
        permit_user = request.POST['permit_user']
        # アクセス可能ユーザーをデータで持ち、ログインユーザーがアクセス権限あればページを開く、なければアラート表示後ページ遷移
        test = TemporaryMakePlantiaImport.objects.filter(lost_flag=0)
        for test in test:
            if test.user.username == request.user.username:
                data = {
                    'user_division_cd': user_division_cd,
                    'user_department_cd': user_department_cd,
                    'user_authority': user_authority,
                    'confirm_user': confirm_user,
                    'permit_user': permit_user,
                }
                return render(request, 'fms/parts/temporary_response/make_plantia_import_excel_page.html', data)
        return HttpResponse("<script>alert('アクセス権がありません!');window.history.back(-1);</script>")
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# Plantiaインポート用excel１レコード出力
def make_plantia_import_excel_row(ws_sheet, row, export_report_item, data):
    ws_sheet.cell(row=row, column=1).value = "Record"
    ws_sheet.cell(row=row, column=3).value = "$NEXTVAL$"
    ws_sheet.cell(row=row, column=4).value = export_report_item.phenomenon_id2
    ws_sheet.cell(row=row, column=5).value = export_report_item.m_site_skey
    ws_sheet.cell(row=row, column=6).value = export_report_item.m_mgt_cls_skey
    ws_sheet.cell(row=row, column=7).value = export_report_item.maintenance_personnel

    ws_sheet.cell(row=row, column=8).value = data['fclty_ldgr_skey']

    ws_sheet.cell(row=row, column=9).value = export_report_item.m_location_skey
    ws_sheet.cell(row=row, column=10).value = export_report_item.maintenance_name
    ws_sheet.cell(row=row, column=11).value = export_report_item.report_detail
    ws_sheet.cell(row=row, column=12).value = export_report_item.serious_breakdown_case
    ws_sheet.cell(row=row, column=13).value = export_report_item.cycle_reference_date
    ws_sheet.cell(row=row, column=14).value = export_report_item.start_date
    ws_sheet.cell(row=row, column=15).value = export_report_item.completion_date
    ws_sheet.cell(row=row, column=16).value = export_report_item.time_of_occurrence

    ws_sheet.cell(row=row, column=17).value = data['s_user_skey']

    ws_sheet.cell(row=row, column=18).value = export_report_item.person_in_charge_of_the_original_section
    ws_sheet.cell(row=row, column=19).value = export_report_item.phenomenon_details
    ws_sheet.cell(row=row, column=20).value = export_report_item.cause_detail
    ws_sheet.cell(row=row, column=21).value = export_report_item.countermeasure
    ws_sheet.cell(row=row, column=22).value = export_report_item.special_note_construction_work
    ws_sheet.cell(row=row, column=23).value = export_report_item.message
    ws_sheet.cell(row=row, column=24).value = export_report_item.stop_time
    ws_sheet.cell(row=row, column=25).value = export_report_item.m_phenomenon_cd_skey
    ws_sheet.cell(row=row, column=26).value = export_report_item.m_position_cd_skey_phenomenon
    ws_sheet.cell(row=row, column=27).value = export_report_item.m_condition_cd_skey
    ws_sheet.cell(row=row, column=28).value = export_report_item.m_position_cd_skey_condition
    ws_sheet.cell(row=row, column=29).value = export_report_item.m_cause_cd_skey
    ws_sheet.cell(row=row, column=30).value = export_report_item.m_position_cd_skey_cause
    ws_sheet.cell(row=row, column=31).value = export_report_item.m_result_cd_skey
    ws_sheet.cell(row=row, column=32).value = export_report_item.m_treatment_cd_skey
    ws_sheet.cell(row=row, column=33).value = export_report_item.m_position_cd_skey_treatment
    ws_sheet.cell(row=row, column=34).value = export_report_item.m_exe_cls_skey
    ws_sheet.cell(row=row, column=40).value = export_report_item.repair_time
    ws_sheet.cell(row=row, column=45).value = "小口工事"
    ws_sheet.cell(row=row, column=46).value = export_report_item.s_specman_list_value_skey
    ws_sheet.cell(row=row, column=47).value = export_report_item.special_note_production
    ws_sheet.cell(row=row, column=48).value = export_report_item.items_to_be_sent_production
    ws_sheet.cell(row=row, column=49).value = export_report_item.attachment
    ws_sheet.cell(row=row, column=50).value = export_report_item.is_need_input_plantia

    ws_sheet.cell(row=row, column=51).value = data['location_nm_1']
    ws_sheet.cell(row=row, column=52).value = data['eqpt_id']
    ws_sheet.cell(row=row, column=53).value = data['user_id']


# Plantiaインポートデータ外部データ取得（兼エラーチェック）
def get_plantia_import_data(export_report_item):
    msg = ''
    data = {}
    # 工事担当者のs_keyをPLANTIA側から取得(未設定の場合は空欄)
    s_user_skey = ''
    if export_report_item.construction_representative is not None:
        representative_user_list = User.objects.filter(
            username=str.lower(export_report_item.construction_representative))

        if representative_user_list.count() < 1:
            # 該当の工事担当者が登録されていない
            msg = '　phenomenon_id:' + str(export_report_item.phenomenon_id2) + ',工事担当者(User未登録)\n'
            return msg, data

        representative_user = representative_user_list[0]
        if representative_user.last_name is None or representative_user.first_name is None:
            # 工事担当者の氏名が登録されていない
            msg = '　phenomenon_id:' + str(export_report_item.phenomenon_id2) + ',工事担当者(User氏名未設定)\n'
            return msg, data

        representative_user_name = representative_user.last_name + ' ' + representative_user.first_name
        construction_representative_user_list = ServiceUser.objects.filter(
            user_nm=representative_user_name, deleted_flg=0)

        if construction_representative_user_list.count() < 1:
            # PLANTIA側のユーザー情報と不一致の場合はエラー表示
            msg = '　phenomenon_id:' + str(export_report_item.phenomenon_id2) + ',工事担当者(PLANTIA User未登録)\n'
            return msg, data
        s_user_skey = construction_representative_user_list[0].s_user_skey

    # ロケーション番号をPLANTIA側から取得
    if export_report_item.m_location_skey is not None:
        location_list = MasterLocation.objects.filter(
            m_location_skey=export_report_item.m_location_skey, deleted_flg=0)
        if location_list.count() > 0:
            location_nm_1 = location_list[0].location_nm_1
        else:
            msg = '　phenomenon_id:' + str(export_report_item.phenomenon_id2) + ',ロケーション_skey(PLANTIA未登録)\n'
            return msg, data
    else:
        # 未設定の場合はエラー表示
        msg = '　phenomenon_id:' + str(export_report_item.phenomenon_id2) + ',ロケーション番号(未設定)\n'
        return msg, data

    # 保全担当者のuser_idをPLANTIA側から取得
    if export_report_item.maintenance_personnel is not None:
        maintenance_user_list = ServiceUser.objects.filter(
            s_user_skey=export_report_item.maintenance_personnel, deleted_flg=0)
        if maintenance_user_list.count() > 0:
            user_id = maintenance_user_list[0].user_id
        else:
            # PLANTIA側のユーザー情報と不一致の場合はエラー表示
            msg = '　phenomenon_id:' + str(export_report_item.phenomenon_id2) + ',保全担当者(PLANTIA User未登録)\n'
            return msg, data
    else:
        # 未設定の場合はエラー表示
        msg = '　phenomenon_id:' + str(export_report_item.phenomenon_id2) + ',保全担当者(未設定)\n'
        return msg, data

    data = {
        's_user_skey': s_user_skey,
        'location_nm_1': location_nm_1,
        'user_id': user_id,
    }
    return msg, data


# Plantiaインポート用excel作成
@login_required
@require_POST
def make_plantia_import_excel(request):
    try:
        target_folder = 'PLANTIA'
        templates_file_name = 'PLANTIAインポートシートテンプレート'
        create_file_name = 'PLANTIAインポートFile'
        templates_file_ext = '.xlsx'
        new_sheet_name = 'DATA-MAP'

        output_unique_id_list = request.POST.getlist("output_unique_id_list[]")

        if len(output_unique_id_list) < 1:
            msg = '出力対象がありません'
            data = {
                'msg': msg,
                'target_folder': '',
                'file_name': '',
            }
            return JsonResponse(data)

        # 前回ファイルを削除
        clear_output_file_folder(target_folder)

        # 共有フォルダパス取得
        templates_file_full_name = templates_file_name + templates_file_ext
        templates_file_path = get_template_file_path(target_folder, templates_file_full_name)

        # テンプレートファイル読込
        wb_new_file = openpyxl.load_workbook(templates_file_path)
        ws_new_sheet = wb_new_file[new_sheet_name]

        # 出力対象レコード取得
        export_report_list = EquipmentHistoryReport.objects.filter(id__in=output_unique_id_list, lost_flag=0)

        # 出力先頭行指定（テンプレートファイルに依存)
        i = 5
        # 異常データ数カウント
        error_data_count = 0
        msg = "PLANTIAインポート用データに異常があります！！\n" + "以下のデータを確認してください\n" + "\n"

        # シートにレポート情報を出力(外部データ取得と、１行分の出力処理を共通化)
        for export_report_list_item in export_report_list:
            # 外部データ取得
            ret_msg, ret_data = get_plantia_import_data(export_report_list_item)
            if len(ret_msg) > 0:
                error_data_count += 1
                msg += ret_msg
                continue

            equipment_list = MaintenanceEquipment.objects.filter(
                phenomenon_id=export_report_list_item.phenomenon_id2, lost_flag=0)

            # 機番が無い場合
            if equipment_list.count() == 0:
                # 指定行に出力（共通関数化)
                export_data = {
                    's_user_skey': ret_data['s_user_skey'],
                    'location_nm_1': ret_data['location_nm_1'],
                    'user_id': ret_data['user_id'],
                    'eqpt_id': '',
                    'fclty_ldgr_skey': '',
                }
                make_plantia_import_excel_row(ws_new_sheet, i, export_report_list_item, export_data)
                i += 1

            # 機番がある場合
            else:
                for equipment_item in equipment_list:
                    # 機器番号をPLANTIA側から取得
                    if equipment_item.t_fclty_ldgr_skey is not None:
                        fclty_ldgr_list = FcltyLdgr.objects.filter(
                            t_fclty_ldgr_skey=equipment_item.t_fclty_ldgr_skey)
                        if fclty_ldgr_list.count() > 0:
                            eqpt_id = fclty_ldgr_list[0].eqpt_id
                        else:
                            # PLANTIA側のユーザー情報と不一致の場合はエラー表示
                            error_data_count += 1
                            msg += '　phenomenon_id:' + str(
                                equipment_item.phenomenon_id) + ',機器番号(PLANTIA未登録)\n'
                            continue
                    else:
                        # 未設定の場合はエラー表示
                        error_data_count += 1
                        msg += '　phenomenon_id:' + str(equipment_item.phenomenon_id) + ',機器番号(未設定)\n'
                        continue

                    # 指定行に出力（共通関数化)
                    export_data = {
                        's_user_skey': ret_data['s_user_skey'],
                        'location_nm_1': ret_data['location_nm_1'],
                        'user_id': ret_data['user_id'],
                        'fclty_ldgr_skey': equipment_item.t_fclty_ldgr_skey,
                        'eqpt_id': eqpt_id,
                    }
                    make_plantia_import_excel_row(ws_new_sheet, i, export_report_list_item, export_data)
                    i += 1

        # 異常データが無ければ
        if error_data_count < 1:
            # 罫線描画？
            side = Side(style='thin', color='000000')
            # set border (black thin line)
            border = Border(top=side, bottom=side, left=side, right=side)
            # write in sheet
            for row in ws_new_sheet:
                for cell in row:
                    ws_new_sheet[cell.coordinate].border = border

            # ダウンロード用ファイルに保存
            # new_file_name = 'PLANTIAインポートFile.xlsx'
            # 現在日時を取得。ファイル名に加える --- (*1)
            today = datetime.datetime.now().strftime('_%Y%m%d%H%M%S')
            new_file_name = create_file_name + today + templates_file_ext
            output_file_path = get_output_file_path(target_folder, new_file_name)
            wb_new_file.save(output_file_path)

            # 出力対象フラグ更新
            for export_report_list_item in export_report_list:
                # 出力対象フラグ更新
                export_report_list_item.export_complete_flag = 2
                export_report_list_item.save()

            msg = '出力完了'
        else:
            new_file_name = ""

        data = {
            'msg': msg,
            'target_folder': target_folder,
            'file_name': new_file_name,
        }

        return JsonResponse(data)

    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# 故障対応工程完了済み案件一覧の絞込のパーツ表示
@require_POST
def end_maintenance_order_filter(request):
    from fms.views.common_def_views import get_department_person_list
    try:
        # ・案件ID
        # ・案件名
        # ・工場
        facility_lists = MasterLocation.objects.filter(parent_m_location_skey__isnull=False, deleted_flg=0).order_by('display_order')
        # ・工事完了日
        # ・機器番号
        fcltyldgr = FcltyLdgr.objects.filter(deleted_flg=0)

        # ・保全G担当者 (ユーザ全体からの抽出用)
        mng_charge_person_list = get_department_person_list('MNG')

        # 管理区分(機器番号フィルタ用)
        mgt_cls_lists = MasterMgtCls.objects.filter(deleted_flg=0).all().order_by('display_order')

        data = {
            'facility_lists': facility_lists,
            'mng_charge_person_list': mng_charge_person_list,
            'fcltyldgr': fcltyldgr,
            'mgt_cls_lists': mgt_cls_lists,
        }

        return render(request, 'fms/parts/temporary_response/end_maintenance_order_filter.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# 故障対応工程完了済み案件一覧
@require_POST
def end_maintenance_order_list(request):
    try:
        sel_phenomenon_id = request.POST['sel_phenomenon_id']
        sel_project_name = request.POST['sel_project_name']

        # mgt_cls = request.POST['mgt_cls']
        facility = request.POST['facility']
        eqpt_no = request.POST['eqpt_no']

        sel_completion_date = request.POST['sel_completion_date']
        sel_mng_charge_person = request.POST['sel_mng_charge_person']

        filter_export_complete_flag_list = request.POST.getlist('filter_export_complete_flag_list[]')

        where_str = ""
        where_parm1 = []

        # 検索条件
        # 案件ID
        if sel_phenomenon_id != "":
            where_phenomenon_id = Q(phenomenon_id2=sel_phenomenon_id)
        else:
            where_phenomenon_id = Q()
        # 案件名
        if sel_project_name != "":
            where_project_name = Q(maintenance_name__icontains=sel_project_name)
        else:
            where_project_name = Q()
        # 工場
        if facility != "":
            where_facility = Q(m_location_skey=facility)
        else:
            where_facility = Q()
        # 機器番号
        if eqpt_no != "":
            # 機器番号はEquipmentHistoryReportテーブルに登録されないため、
            # MaintenanceEquipmentテーブルからphenomenon_idを取得してフィルタする
            maintenance_equipment_phenomenon_id_list = []
            maintenance_equipment = MaintenanceEquipment.objects.filter(t_fclty_ldgr_skey=int(eqpt_no), lost_flag=0)
            for maintenance_equipment_item in maintenance_equipment:
                maintenance_equipment_phenomenon_id_list.append(maintenance_equipment_item.phenomenon_id)
            # phenomenon_idは複数件Hitする可能性があるため、IN句検索
            where_eqpt_no = Q(phenomenon_id2__in=maintenance_equipment_phenomenon_id_list)
        else:
            where_eqpt_no = Q()
        # 工事完了日
        if sel_completion_date != "":
            date_str = date_to_many_type(sel_completion_date)
            completion_date = date_str.str_type_date_hyphen
            where_completion_date = Q(completion_date=completion_date)
        else:
            where_completion_date = Q()

        # 保全G担当者
        if sel_mng_charge_person != "":
            charge_person = User.objects.get(username=sel_mng_charge_person)
            charge_person_name = charge_person.last_name + ' ' + charge_person.first_name
            charge_person_list = ServiceUser.objects.filter(user_nm=charge_person_name, deleted_flg=0)
            if charge_person_list.count() > 0:
                maintenance_personnel = charge_person_list[0].s_user_skey
                where_mng_charge_person = Q(maintenance_personnel=maintenance_personnel)
            else:
                where_mng_charge_person = Q()
        else:
            where_mng_charge_person = Q()
        # 出力状態
        where_filter_export_complete_flag = Q()
        if len(filter_export_complete_flag_list):
            for filter_export_complete_flag_list_item in filter_export_complete_flag_list:
                where_filter_export_complete_flag.add(Q(export_complete_flag=filter_export_complete_flag_list_item), Q.OR)

        end_maintenance_order_lists = EquipmentHistoryReport.objects.filter(Q(lost_flag=0)
                                                                            & where_phenomenon_id
                                                                            & where_project_name
                                                                            & where_facility
                                                                            & where_eqpt_no
                                                                            & where_completion_date
                                                                            & where_mng_charge_person
                                                                            & where_filter_export_complete_flag
                                                                            ).exclude(export_complete_flag=0).exclude(export_complete_flag=None).order_by('phenomenon_id2')
        end_maintenance_order_list_num = len(list(end_maintenance_order_lists))

        phenomenon_id_list = []
        for end_maintenance_order_lists_item in end_maintenance_order_lists:
            # phenomenonテーブルデータ取得用
            phenomenon_id_list.append(end_maintenance_order_lists_item.phenomenon_id2)

            # 工場のサロゲートキーを日本語名に書き換え
            if end_maintenance_order_lists_item.m_location_skey is not None:
                m_location_skey = str(end_maintenance_order_lists_item.m_location_skey)
                master_location_count = MasterLocation.objects.filter(m_location_skey=m_location_skey, deleted_flg=0).count()
                if master_location_count > 0:
                    end_maintenance_order_lists_item.m_location_skey = MasterLocation.objects.get(m_location_skey=m_location_skey,
                                                                                                  deleted_flg=0).location_nm_1
            # phenomenon_idにて機器番号のサロゲートキー取得し、機器番号に書き換え
            maintenance_equipment_data = MaintenanceEquipment.objects.filter(phenomenon_id=end_maintenance_order_lists_item.phenomenon_id2,
                                                                             lost_flag=0)
            if len(maintenance_equipment_data) > 0:
                loop = 0
                equipment_list = ''
                for maintenance_equipment_data_item in maintenance_equipment_data:
                    if loop != 0:
                        equipment_list += '/'
                    t_fclty_ldgr_skey = str(maintenance_equipment_data_item.t_fclty_ldgr_skey)
                    fclty_ldgr_count = FcltyLdgr.objects.filter(t_fclty_ldgr_skey=t_fclty_ldgr_skey).count()
                    if fclty_ldgr_count > 0:
                        equipment_list += FcltyLdgr.objects.get(t_fclty_ldgr_skey=t_fclty_ldgr_skey).eqpt_id
                    loop += 1
                end_maintenance_order_lists_item.t_fclty_ldgr_skey = equipment_list
            else:
                end_maintenance_order_lists_item.t_fclty_ldgr_skey = ''
            # 保全G担当者のサロゲートキーを日本語名に書き換え
            if end_maintenance_order_lists_item.maintenance_personnel is not None:
                maintenance_personnel = str(end_maintenance_order_lists_item.maintenance_personnel)
                service_user_count = ServiceUser.objects.filter(s_user_skey=maintenance_personnel, deleted_flg=0).count()
                if service_user_count > 0:
                    end_maintenance_order_lists_item.maintenance_personnel = ServiceUser.objects.get(s_user_skey=maintenance_personnel,
                                                                                                     deleted_flg=0).user_nm

        phenomenon_lists = Phenomenon.objects.filter(phenomenon_id__in=phenomenon_id_list, lost_flag=0)
        phenomenon_lists_num = len(list(phenomenon_lists))

        msg = "故障対応工程完了済み案件一覧取得完了"

        data = {
            'end_maintenance_order_lists': end_maintenance_order_lists,
            'end_maintenance_order_list_num': end_maintenance_order_list_num,
            'phenomenon_lists': phenomenon_lists,
            'phenomenon_lists_num': phenomenon_lists_num,
            'msg': msg,
        }

        return render(request, 'fms/parts/temporary_response/end_maintenance_order_list.html', data)

    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# ヒアリング資料出力画面表示
def make_hearing_document_page(request):
    try:
        t_username = request.user.username
        t_user_last_name = request.user.last_name
        t_user_first_name = request.user.first_name
        t_user_is_superuser = request.user.is_superuser
        user_name = t_user_last_name + t_user_first_name

        user_division_cd = request.POST['user_division_cd']
        user_department_cd = request.POST['user_department_cd']
        user_authority = int(request.POST['user_authority'])
        confirm_user = request.POST['confirm_user']
        permit_user = request.POST['permit_user']

        # ユーザーの部署名を取得
        department_data = DepartmentMaster.objects.get(department_cd=user_department_cd)
        user_department = department_data.department_name

        # ユーザーの部門名を取得
        division_data = DivisionMaster.objects.get(division_cd=user_division_cd)
        user_division = division_data.division_name

        # level5_step_top_page を利用するが、ステップリスト空で表示
        step_name = 'ヒアリング資料出力'
        step_num_list = []
        level5_step_id = 133000000

        data = {
            't_user_name': t_username,
            'user_first_name': t_user_first_name,
            'user_last_name': t_user_last_name,
            't_user_is_superuser': t_user_is_superuser,
            'user_name': user_name,
            'user_department': user_department,
            'user_division': user_division,
            'user_division_cd': user_division_cd,
            'user_department_cd': user_department_cd,
            'user_authority': user_authority,
            'confirm_user': confirm_user,
            'permit_user': permit_user,
            'step_name': step_name,
            'step_num_list': step_num_list,
            'level5_step_id': level5_step_id,
        }

        return render(request, 'fms/parts/temporary_response/make_hearing_document_page.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise

# ヒアリング資料excel作成
@login_required
@require_POST
def make_hearing_document(request):
    try:
        # POST情報取得
        budget_id_list = request.POST.getlist("budget_id_list[]")

        target_folder = 'hearing_document'
        templates_file_name = 'ヒアリング資料'
        templates_file_ext = '.xlsx'
        budget_sheet_name = '原紙'
        index_sheet_name = '目次'
        total_sheet_name = '集計'

        output_budget_list, department_id_all, total_info_list = get_hearing_document_output_budget_list(budget_id_list)

        if len(output_budget_list) < 1:
            msg = '出力対象予算がありません'
            data = {
                'msg': msg,
                'target_folder': '',
                'file_name': '',
            }
            return JsonResponse(data)

        # 前回ファイルを削除
        clear_output_file_folder(target_folder)

        # 共有フォルダパス取得
        templates_file_full_name = templates_file_name + templates_file_ext
        templates_file_path = get_template_file_path(target_folder, templates_file_full_name)

        # テンプレートファイル読込
        wb_new_file = openpyxl.load_workbook(templates_file_path)
        index_sheet = wb_new_file[index_sheet_name]
        total_sheet = wb_new_file[total_sheet_name]

        # 予算分のシートコピー(先頭予算以外)
        for budget_item in output_budget_list:
            if output_budget_list[0].budget_id != budget_item.budget_id:
                ws_new_sheet = wb_new_file.copy_worksheet(wb_new_file[budget_sheet_name])
                ws_new_sheet.title = str(budget_item.budget_id)

        # 先頭予算のシート名設定
        ws_new_sheet = wb_new_file[budget_sheet_name]
        ws_new_sheet.title = str(output_budget_list[0].budget_id)

        # 各シートに出力 start
        index = 1
        for budget_item in output_budget_list:

            # 目次シート出力
            make_hearing_document_index_sheet(index_sheet, budget_item, index)

            # 予算シート出力
            budget_sheet = wb_new_file[str(budget_item.budget_id)]
            make_hearing_document_budget_sheet(budget_sheet, budget_item, index)

            index = index + 1
        # 各シートに出力 end

        # 集計シート出力
        make_hearing_document_total_sheet(total_sheet, total_info_list, department_id_all)

        # 目次シートの罫線描画
        set_border_index_sheet(index_sheet, index - 1)

        # ダウンロード用ファイルに保存
        today = datetime.datetime.now().strftime('_%Y%m%d_%H%M')
        new_file_name = templates_file_name + today + templates_file_ext
        output_file_path = get_output_file_path(target_folder, new_file_name)
        wb_new_file.save(output_file_path)

        msg = '出力完了'
        data = {
            'msg': msg,
            'target_folder': target_folder,
            'file_name': new_file_name,
        }
        return JsonResponse(data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# ヒアリング資料出力対象予算IDリスト取得 兼集計処理
def get_hearing_document_output_budget_list(budget_id_list):
    total_info_list = [0, 0, 0, 0, 0, 0, 0, 0]
    output_budget_list = []
    department_id = ''
    department_id_all = ''

    # 工事区分を判定し、出力対象リストを作成 + 集計
    for budget_id in budget_id_list:
        budget_list = Budget.objects.filter(budget_id=budget_id, lost_flag=0)
        if len(budget_list) != 1:
            continue
        budget_item = budget_list[0]

        if budget_item.budget_class_id == 1:
            # 特別建設
            total_index = 0
        elif budget_item.budget_class_id == 2:
            # 別枠建設
            total_index = 2
        elif budget_item.budget_class_id == 3:
            # 一般建設
            total_index = 4
        elif budget_item.budget_class_id == 4:
            # 改修項目
            total_index = 6
        else:
            # その他は集計対象外とする
            continue

        # 件数、合計額集計
        total_info_list[total_index] = total_info_list[total_index] + 1
        total_info_list[total_index + 1] = total_info_list[total_index + 1] + budget_item.budget_price

        # 部署名統一チェック
        if department_id == '':
            department_id = budget_item.budget_main_department_id
            department_id_all = budget_item.budget_main_department_id
        elif department_id != budget_item.budget_main_department_id:
            department_id_all = ''

        # 出力対象リストに追加
        output_budget_list.append(budget_item)

    return output_budget_list, department_id_all, total_info_list


# 指定シートの罫線描画
def set_border_index_sheet(ws_new_sheet, count):
    # 罫線描画
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    for row_num in range(3, count + 4):
        for col_num in range(2, 17):
            ws_new_sheet.cell(row=row_num, column=col_num).border = border

    return


# 集計シートデータ出力
def make_hearing_document_total_sheet(ws_new_sheet, total_info, department_id):
    # 特別建設
    ws_new_sheet.cell(row=9, column=6).value = total_info[0]
    ws_new_sheet.cell(row=9, column=8).value = total_info[1]
    # 別枠建設
    ws_new_sheet.cell(row=10, column=6).value = total_info[2]
    ws_new_sheet.cell(row=10, column=8).value = total_info[3]
    # 一般建設
    ws_new_sheet.cell(row=11, column=6).value = total_info[4]
    ws_new_sheet.cell(row=11, column=8).value = total_info[5]
    # 改修項目
    ws_new_sheet.cell(row=12, column=6).value = total_info[6]
    ws_new_sheet.cell(row=12, column=8).value = total_info[7]

    # 部署名
    if department_id != '':
        department_item = DepartmentMaster.objects.get(
            department_cd=department_id, lost_flag=0)
        ws_new_sheet.cell(row=5, column=5).value = department_item.department_name

    return


# 目次シートデータ出力
def make_hearing_document_index_sheet(ws_new_sheet, budget_item, index):
    row = index + 3
    ws_new_sheet.cell(row=row, column=2).value = index
    ws_new_sheet.cell(row=row, column=3).value = budget_item.budget_id
    ws_new_sheet.cell(row=row, column=4).value = budget_item.business_year_id
    ws_new_sheet.cell(row=row, column=8).value = budget_item.budget_name
    ws_new_sheet.cell(row=row, column=9).value = budget_item.start_date
    ws_new_sheet.cell(row=row, column=10).value = budget_item.end_date
    ws_new_sheet.cell(row=row, column=11).value = budget_item.budget_price
    # 備考と除去資産は結合して出力
    budget_rem_str = ''
    if budget_item.budget_rem is not None:
        budget_rem_str += budget_item.budget_rem
        budget_rem_str += '\n'

    if budget_item.remove_assets is not None:
        budget_rem_str += budget_item.remove_assets

    ws_new_sheet.cell(row=row, column=16).value = budget_rem_str

    # 別テーブルのデータを出力
    department_item = DepartmentMaster.objects.get(
        department_cd=budget_item.budget_main_department_id, lost_flag=0)
    ws_new_sheet.cell(row=row, column=5).value = department_item.department_name

    budget_class_item = BudgetClassMaster.objects.get(
        budget_class_cd=budget_item.budget_class_id, lost_flag=0)
    ws_new_sheet.cell(row=row, column=6).value = budget_class_item.budget_class_name

    period_class_item = PeriodClassMaster.objects.get(
        period_class_cd=budget_item.period_class_id, lost_flag=0)
    ws_new_sheet.cell(row=row, column=7).value = period_class_item.period_class_name

    process_item = ProcessMaster.objects.get(
        process_cd2=budget_item.facility_process_id, lost_flag=0)
    ws_new_sheet.cell(row=row, column=12).value = process_item.process_cd
    ws_new_sheet.cell(row=row, column=13).value = process_item.process_name

    purpose_class_item = PurposeClassMaster.objects.get(
        purpose_class_cd=budget_item.purpose_class_id, lost_flag=0)
    ws_new_sheet.cell(row=row, column=14).value = budget_item.purpose_class_id
    ws_new_sheet.cell(row=row, column=15).value = purpose_class_item.purpose_class_name

    return


# 工事区分に応じて、予算シートのタイトルを取得
def get_hearing_document_budget_class_title(budget_class_id):
    title_str = ''
    if budget_class_id == 1:
        # 特別建設
        title_str = '年度　特別建設工事(購入)予算項目申請書'
    elif budget_class_id == 2:
        # 別枠建設
        title_str = '年度　別枠建設工事(購入)予算項目申請書'
    elif budget_class_id == 3:
        # 一般建設
        title_str = '年度　経常一般建設工事(購入)予算項目申請書'
    elif budget_class_id == 4:
        # 改修項目
        title_str = '年度　改修項目工事(購入)予算項目申請書'

    return title_str


# 予算シートデータ出力
def make_hearing_document_budget_sheet(ws_new_sheet, budget_item, index):
    ws_new_sheet.cell(row=3, column=4).value = budget_item.business_year_id
    ws_new_sheet.cell(row=7, column=2).value = str(index)
    ws_new_sheet.cell(row=8, column=2).value = budget_item.budget_id
    ws_new_sheet.cell(row=6, column=4).value = budget_item.budget_name
    ws_new_sheet.cell(row=6, column=9).value = budget_item.budget_price
    ws_new_sheet.cell(row=7, column=4).value = budget_item.start_date
    ws_new_sheet.cell(row=7, column=5).value = budget_item.order_date
    ws_new_sheet.cell(row=8, column=4).value = budget_item.end_date
    ws_new_sheet.cell(row=8, column=5).value = budget_item.delivery_date
    ws_new_sheet.cell(row=7, column=7).value = budget_item.pre_order_flag
    ws_new_sheet.cell(row=8, column=7).value = budget_item.asdm_flag

    ws_new_sheet.cell(row=10, column=1).value = budget_item.purpose
    ws_new_sheet.cell(row=10, column=6).value = budget_item.detail
    ws_new_sheet.cell(row=8, column=10).value = budget_item.effect
    ws_new_sheet.cell(row=20, column=10).value = budget_item.influence_for_operation
    ws_new_sheet.cell(row=24, column=10).value = budget_item.influence_for_quality

    # 備考と除去資産は結合して出力
    budget_rem_str = ''
    if budget_item.budget_rem is not None:
        budget_rem_str += budget_item.budget_rem
        budget_rem_str += '\n'

    if budget_item.remove_assets is not None:
        budget_rem_str += budget_item.remove_assets

    ws_new_sheet.cell(row=24, column=6).value = budget_rem_str

    # 別テーブルのデータを出力
    department_item = DepartmentMaster.objects.get(
        department_cd=budget_item.budget_main_department_id, lost_flag=0)
    ws_new_sheet.cell(row=5, column=4).value = department_item.department_name

    purpose_class_item = PurposeClassMaster.objects.get(
        purpose_class_cd=budget_item.purpose_class_id, lost_flag=0)
    ws_new_sheet.cell(row=6, column=10).value = budget_item.purpose_class_id
    ws_new_sheet.cell(row=6, column=11).value = purpose_class_item.purpose_class_name

    process_item = ProcessMaster.objects.get(
        process_cd2=budget_item.facility_process_id, lost_flag=0)
    ws_new_sheet.cell(row=5, column=9).value = \
        process_item.process_cd + '(' + process_item.process_name + ')'

    budget_class_item = BudgetClassMaster.objects.get(
        budget_class_cd=budget_item.budget_class_id, lost_flag=0)
    ws_new_sheet.cell(row=5, column=2).value = budget_class_item.budget_class_name

    # 工事区分に応じてタイトル変更
    title_str = get_hearing_document_budget_class_title(budget_item.budget_class_id)
    ws_new_sheet.cell(row=3, column=5).value = title_str

    period_class_item = PeriodClassMaster.objects.get(
        period_class_cd=budget_item.period_class_id, lost_flag=0)
    ws_new_sheet.cell(row=6, column=2).value = period_class_item.period_class_name

    law_str = ''
    law_list = BudgetLaw.objects.filter(
        budget_id=budget_item.budget_id)
    for law_item in law_list:
        if law_str != '':
            law_str += '、'
        law_str += law_item.law_name
    ws_new_sheet.cell(row=7, column=9).value = law_str

    return


# 予算No登録ページ表示
@login_required
@require_POST
def make_budget_registration_list_page(request):
    try:
        t_username = request.user.username
        user_name = request.user.last_name + request.user.first_name

        user_division_cd = request.POST['user_division_cd']
        user_department_cd = request.POST['user_department_cd']
        user_authority = int(request.POST['user_authority'])
        confirm_user = request.POST['confirm_user']
        permit_user = request.POST['permit_user']

        budget_type = request.POST['budget_type']

        # ユーザーの部署名を取得
        department_data = DepartmentMaster.objects.get(department_cd=user_department_cd)
        user_department_name = department_data.department_name

        # ユーザーの部門名を取得
        division_data = DivisionMaster.objects.get(division_cd=user_division_cd)
        user_division_name = division_data.division_name

        # 追加予算登録は 工事企画G(CPG)のメンバーのみ可能
        if user_department_cd != 'CPG':
            return HttpResponse("<script>alert('アクセス権がありません!');window.history.back(-1);</script>")

        step_num_list = []
        if budget_type == 'normal':
            step_name = '通常申請：予算No登録'
            level5_step_id = 133000000
        else:
            step_name = '追加予算：予算No登録'
            level5_step_id = 136000000

        data = {
            't_user_name': t_username,
            'user_name': user_name,
            'user_department': user_department_name,
            'user_division': user_division_name,
            'user_division_cd': user_division_cd,
            'user_department_cd': user_department_cd,
            'user_authority': user_authority,
            'confirm_user': confirm_user,
            'permit_user': permit_user,
            'step_name': step_name,
            'step_num_list': step_num_list,
            'budget_type': budget_type,
            'level5_step_id': level5_step_id,
        }
        return render(request, 'fms/parts/temporary_response/make_budget_registration_list_page.html', data)

    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# 追加予算登録リスト出力
@login_required
@require_POST
def make_budget_registration_list(request):
    try:
        # POST情報取得
        budget_id_list = request.POST.getlist("budget_id_list[]")

        target_folder = 'budget_registration_list'
        templates_file_name = 'ERPインポート用ファイル'
        templates_file_ext = '.xlsx'
        export_sheet_name = 'エクスポート'

        output_budget_list = get_budget_registration_list(budget_id_list)

        if len(output_budget_list) < 1:
            msg = '出力対象予算がありません'
            data = {
                'msg': msg,
                'target_folder': '',
                'file_name': '',
            }
            return JsonResponse(data)

        # 前回ファイルを削除
        clear_output_file_folder(target_folder)

        # 共有フォルダパス取得
        templates_file_full_name = templates_file_name + templates_file_ext
        templates_file_path = get_template_file_path(target_folder, templates_file_full_name)

        # テンプレートファイル読込
        wb_new_file = openpyxl.load_workbook(templates_file_path)
        export_sheet = wb_new_file[export_sheet_name]

        # シートに出力 start
        index = 1
        for budget_item in output_budget_list:
            make_budget_registration_data(export_sheet, budget_item, index)
            index = index + 1
        # シートに出力 end

        # 罫線出力
        set_border_budget_registration_sheet(export_sheet, index)

        # ダウンロード用ファイルに保存
        today = datetime.datetime.now().strftime('_%Y%m%d_%H%M')
        new_file_name = templates_file_name + today + templates_file_ext
        output_file_path = get_output_file_path(target_folder, new_file_name)
        wb_new_file.save(output_file_path)

        msg = '出力完了'
        data = {
            'msg': msg,
            'target_folder': target_folder,
            'file_name': new_file_name,
        }
        return JsonResponse(data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# 追加予算登録出力対象予算取得
def get_budget_registration_list(budget_id_list):
    output_budget_list = Budget.objects.filter(budget_id__in=budget_id_list, lost_flag=0)
    return output_budget_list


# 追加予算データ出力
def make_budget_registration_data(ws_new_sheet, budget_item, index):
    row = index + 1
    ws_new_sheet.cell(row=row, column=3).value = budget_item.budget_class.budget_class_name
    ws_new_sheet.cell(row=row, column=4).value = budget_item.period_class.period_class_name

    # 四日市予算額追加は四日市追加、本社予算額追加は本社追加と表示
    if budget_item.application_class.application_class_cd == 3:
        ws_new_sheet.cell(row=row, column=5).value = '四日市追加'
    elif budget_item.application_class.application_class_cd == 5:
        ws_new_sheet.cell(row=row, column=5).value = '本社追加'
    else:
        ws_new_sheet.cell(row=row, column=5).value = budget_item.application_class.application_class_name

    ws_new_sheet.cell(row=row, column=7).value = budget_item.budget_id
    ws_new_sheet.cell(row=row, column=8).value = budget_item.budget_main_department.department_name

    ws_new_sheet.cell(row=row, column=12).value = budget_item.facility_process.process_name

    process_item = ProcessMaster.objects.get(
        process_cd2=budget_item.facility_process_id, lost_flag=0)
    ws_new_sheet.cell(row=row, column=13).value = process_item.process_cd

    # 予算No付与済の予算を選択された時に気づかせるために、予算Noを出力する
    if budget_item.budget_no is not None:
        ws_new_sheet.cell(row=row, column=14).value = budget_item.budget_no

    ws_new_sheet.cell(row=row, column=15).value = budget_item.budget_name

    ws_new_sheet.cell(row=row, column=17).value = budget_item.purpose_class_id
    ws_new_sheet.cell(row=row, column=19).value = budget_item.budget_price

    # pre_order_flagが要ならばY 不要ならばN
    if budget_item.pre_order_flag is None or budget_item.pre_order_flag == '不要':
        ws_new_sheet.cell(row=row, column=24).value = 'N'
    else:
        ws_new_sheet.cell(row=row, column=24).value = 'Y'

    # asdm_flagが定修ならばY 定修外ならN
    if budget_item.asdm_flag is None or budget_item.asdm_flag == '定修外':
        ws_new_sheet.cell(row=row, column=25).value = 'N'
    else:
        ws_new_sheet.cell(row=row, column=25).value = 'Y'

    ws_new_sheet.cell(row=row, column=27).value = budget_item.start_date
    ws_new_sheet.cell(row=row, column=28).value = budget_item.end_date

    return


# 指定シートの罫線描画
def set_border_budget_registration_sheet(ws_new_sheet, count):
    # 罫線描画
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    for row_num in range(2, count + 1):
        for col_num in range(1, 34):
            ws_new_sheet.cell(row=row_num, column=col_num).border = border
    return


# 追加予算登録リスト入力
@login_required
@require_POST
def import_budget_registration_list(request):
    try:
        # POST情報取得
        if request.FILES.__len__() == 0:
            msg = "ファイルが選択されていません！"
            ary = {
                'msg': msg,
            }
            return JsonResponse(ary)

        file = request.FILES['file']
        target_folder = request.POST['target_folder']
        budget_type = request.POST['budget_type']

        file_full_path = get_input_file_path(target_folder, file.name)

        # Excelファイル読込
        wb_file = openpyxl.load_workbook(file_full_path)

        # 先頭のシート取得
        ws = wb_file.worksheets[0]

        # 予算情報の読込
        budget_registration_data_list, result_msg = read_budget_registration_data(ws)
        if result_msg != '':
            ary = {'msg': result_msg, }
            return JsonResponse(ary)

        # 各予算に予算Noを付与する(通常申請：追加予算共通処理)
        msg = budget_no_registration_list(request, budget_registration_data_list, budget_type)
        data = {
            'msg': msg,
        }
        return JsonResponse(data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# 追加予算登録データ取得
def read_budget_registration_data(ws):
    row = 2
    budget_registration_data = list()
    msg = ''
    while True:
        budget_id = ws.cell(row=row, column=7).value
        budget_no = ws.cell(row=row, column=14).value

        # 正常判定
        if budget_id is not None and budget_id != '' and budget_no is not None and budget_no != '':
            budget_registration_data.append({'budget_id': budget_id, 'budget_no': budget_no})
            row = row + 1
            continue

        # 終了判定(予算ID、予算NOともに空)
        if (budget_id is None or budget_id == '') and (budget_no is None or budget_no == ''):
            break
        else:
            msg = 'インポートデータの' + str(row) + '行目に異常データがあります！！！'
            break

    if len(budget_registration_data) < 1 and msg == '':
        msg = "インポートできる予算がありません！！！"

    return budget_registration_data, msg



# 繰越申請書出力用ページ表示
def make_carry_forward_document_page(request):
    try:
        t_username = request.user.username
        t_user_last_name = request.user.last_name
        t_user_first_name = request.user.first_name
        t_user_is_superuser = request.user.is_superuser
        user_name = t_user_last_name + t_user_first_name

        user_division_cd = request.POST['user_division_cd']
        user_department_cd = request.POST['user_department_cd']
        user_authority = int(request.POST['user_authority'])
        confirm_user = request.POST['confirm_user']
        permit_user = request.POST['permit_user']

        # ユーザーの部署名を取得
        department_data = DepartmentMaster.objects.get(department_cd=user_department_cd)
        user_department = department_data.department_name

        # ユーザーの部門名を取得
        division_data = DivisionMaster.objects.get(division_cd=user_division_cd)
        user_division = division_data.division_name

        # level5_step_top_page を利用するが、ステップリスト空で表示
        step_name = '予算繰越申請書出力'
        step_num_list = []
        level5_step_id = 213010000

        data = {
            't_user_name': t_username,
            'user_first_name': t_user_first_name,
            'user_last_name': t_user_last_name,
            't_user_is_superuser': t_user_is_superuser,
            'user_name': user_name,
            'user_department': user_department,
            'user_division': user_division,
            'user_division_cd': user_division_cd,
            'user_department_cd': user_department_cd,
            'user_authority': user_authority,
            'confirm_user': confirm_user,
            'permit_user': permit_user,
            'step_name': step_name,
            'step_num_list': step_num_list,
            'level5_step_id': level5_step_id,
        }

        return render(request, 'fms/parts/temporary_response/make_carry_forward_document_page.html', data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# 繰越申請書出力機能実行
def make_carry_forward_document(request):
    try:
        now_naive = datetime.datetime.now()
        now = make_aware(now_naive)

        # POST情報取得
        carry_forward_id_list = request.POST.getlist("carry_forward_id_list[]")

        target_folder = 'carry_forward_document'
        templates_file_name = '繰越申請書'
        templates_file_ext = '.xlsx'

        list_sheet_name = 'リスト'
        reason_sheet_name = '理由'

        # 指定された予算繰越情報と、予算情報の収集
        output_budget_list, output_carry_forward_list = \
            get_carry_forward_document_output_budget_list(carry_forward_id_list)
        if len(output_carry_forward_list) < 1:
            msg = '出力対象がありません'
            data = {
                'msg': msg,
                'target_folder': '',
                'file_name': '',
            }
            return JsonResponse(data)

        # 前回ファイルを削除
        clear_output_file_folder(target_folder)
        # 共有フォルダパス取得
        templates_file_full_name = templates_file_name + templates_file_ext
        templates_file_path = get_template_file_path(target_folder, templates_file_full_name)

        # テンプレートファイル読込(結合したセルが多いと異常に時間がかかるので注意)
        wb_new_file = openpyxl.load_workbook(templates_file_path)

        # 出力対象シート取得
        list_sheet = wb_new_file[list_sheet_name]
        reason_sheet = wb_new_file[reason_sheet_name]
        # 各シートに出力 start
        index_list = 0
        index_reason = 0
        for budget_item, carry_forward_item in zip(output_budget_list, output_carry_forward_list):
            # リストシート出力１行分
            make_carry_forward_document_list_sheet(list_sheet, budget_item, carry_forward_item, index_list)

            # 理由シート出力1行分
            make_carry_forward_document_reason_sheet(reason_sheet, budget_item, carry_forward_item, index_reason)

            index_list = index_list + 1
            index_reason = index_reason + 1
            main_budget_id = budget_item.budget_id

            # 関連予算の情報出力
            relation_budget_list = Budget.objects.filter(relation_budget_id=main_budget_id,
                                                         lost_flag=0).exclude(budget_id=main_budget_id)
            # 関連予算分ループ
            for relation_budget_item in relation_budget_list:
                make_carry_forward_document_list_sheet(list_sheet, relation_budget_item, '', index_list)
                index_list = index_list + 1

        # タイトル部分、日付部分の出力
        make_carry_forward_document_title(request, output_budget_list[0], now, list_sheet, reason_sheet)

        # リストシート印刷プロパティ指定
        print_area = 'A1:I'
        print_title_rows = '6:7'
        list_sheet.print_title_rows = print_title_rows
        list_sheet.print_area = print_area + str(index_list + 7)
        list_sheet.page_setup.orientation = list_sheet.ORIENTATION_LANDSCAPE
        list_sheet.page_setup.fitToWidth = 1
        list_sheet.page_setup.fitToHeight = 0
        list_sheet.sheet_properties.pageSetUpPr.fitToPage = True

        # 理由シート印刷プロパティ指定
        print_area = 'A1:U'
        print_title_rows = '5:5'
        reason_sheet.print_title_rows = print_title_rows
        reason_sheet.print_area = print_area + str(index_reason + 5)
        reason_sheet.page_setup.orientation = list_sheet.ORIENTATION_LANDSCAPE
        reason_sheet.page_setup.fitToWidth = 1
        reason_sheet.page_setup.fitToHeight = 0
        reason_sheet.sheet_properties.pageSetUpPr.fitToPage = True

        # ダウンロード用ファイルに保存
        today = now.strftime('_%Y%m%d_%H%M')
        new_file_name = templates_file_name + today + templates_file_ext
        output_file_path = get_output_file_path(target_folder, new_file_name)
        wb_new_file.save(output_file_path)

        msg = '出力完了'
        data = {
            'msg': msg,
            'target_folder': target_folder,
            'file_name': new_file_name,
        }
        return JsonResponse(data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


# 予算繰越申請書出力用 予算、繰越情報取得
def get_carry_forward_document_output_budget_list(carry_forward_id_list):
    from fms.models import BudgetCarryForward

    output_budget_list = []
    output_carry_forward_list = []

    # 出力対象リストを作成
    for carry_forward_id in carry_forward_id_list:
        carry_forward_data = BudgetCarryForward.objects.get(carry_forward_id=carry_forward_id, lost_flag=0)
        budget_data = Budget.objects.get(budget_id=carry_forward_data.budget_id, lost_flag=0)

        # 出力対象リストに追加
        output_budget_list.append(budget_data)
        output_carry_forward_list.append(carry_forward_data)

    return output_budget_list, output_carry_forward_list


# 予算繰越申請書 タイトル部分出力
def make_carry_forward_document_title(request, budget_item, now, list_sheet, reason_sheet):

    # ログインユーザー氏名取得
    t_user_last_name = request.user.last_name
    t_user_first_name = request.user.first_name
    user_name = t_user_last_name + ' ' + t_user_first_name

    # 年度取得
    business_year = budget_item.business_year_id

    # 出力日付
    today_title = now.strftime('%Y/%m/%d')

    # 年度出力
    list_sheet.cell(row=2, column=3).value = business_year
    reason_sheet.cell(row=2, column=6).value = business_year

    # 日付出力
    list_sheet.cell(row=1, column=9).value = today_title
    reason_sheet.cell(row=1, column=21).value = today_title

    # 作成者出力
    list_sheet.cell(row=2, column=9).value = user_name

    return


# 予算繰越申請書 リストシート出力
def make_carry_forward_document_list_sheet(ws_new_sheet, budget_item, carry_forward_item, index):
    from fms.views.execution_output_application_views import get_construction_data
    row_index = index + 8

    ws_new_sheet.cell(row=row_index, column=1).value = budget_item.budget_main_department.department_name
    ws_new_sheet.cell(row=row_index, column=2).value = budget_item.budget_no
    ws_new_sheet.cell(row=row_index, column=3).value = budget_item.budget_id
    ws_new_sheet.cell(row=row_index, column=4).value = budget_item.budget_name
    ws_new_sheet.cell(row=row_index, column=5).value = budget_item.budget_price

    # 追加予算側は繰越情報は出力しない
    if carry_forward_item != '':
        if carry_forward_item.order_complete_flag == 1:
            ws_new_sheet.cell(row=row_index, column=6).value = 'Y'
        else:
            ws_new_sheet.cell(row=row_index, column=6).value = 'N'

        # 実行額集計処理
        construction_count, construction_total_price = get_construction_data(budget_item.budget_id)
        ws_new_sheet.cell(row=row_index, column=7).value = construction_total_price

        ws_new_sheet.cell(row=row_index, column=8).value = carry_forward_item.carry_forward_price
        ws_new_sheet.cell(row=row_index, column=9).value = carry_forward_item.end_date

    return


# 予算繰越申請書 理由シート出力
def make_carry_forward_document_reason_sheet(ws_new_sheet, budget_item, carry_forward_item, index):
    row_index = index + 6

    ws_new_sheet.cell(row=row_index, column=1).value = budget_item.budget_no
    ws_new_sheet.cell(row=row_index, column=3).value = carry_forward_item.carry_forward_reason

    return


# 中期計画：単年度計画移行ページ表示
@login_required
@require_POST
def go_next_budget_plan_page(request):
    try:
        t_username = request.user.username
        user_name = request.user.last_name + request.user.first_name

        user_division_cd = request.POST['user_division_cd']
        user_department_cd = request.POST['user_department_cd']
        user_authority = int(request.POST['user_authority'])
        confirm_user = request.POST['confirm_user']
        permit_user = request.POST['permit_user']

        # ユーザーの部署名を取得
        department_data = DepartmentMaster.objects.get(department_cd=user_department_cd)
        user_department_name = department_data.department_name

        # ユーザーの部門名を取得
        division_data = DivisionMaster.objects.get(division_cd=user_division_cd)
        user_division_name = division_data.division_name

        # 中期計画：単年度計画移行は 工事企画G(CPG)のメンバーのみ可能
        if user_department_cd != 'CPG':
            return HttpResponse("<script>alert('アクセス権がありません!');window.history.back(-1);</script>")

        step_num_list = []
        step_name = '中期計画：単年度計画移行'
        level5_step_id = 132000000

        data = {
            't_user_name': t_username,
            'user_name': user_name,
            'user_department': user_department_name,
            'user_division': user_division_name,
            'user_division_cd': user_division_cd,
            'user_department_cd': user_department_cd,
            'user_authority': user_authority,
            'confirm_user': confirm_user,
            'permit_user': permit_user,
            'step_name': step_name,
            'step_num_list': step_num_list,
            'level5_step_id': level5_step_id,
        }
        return render(request, 'fms/parts/temporary_response/go_next_budget_plan_page.html', data)

    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise
