import datetime
import mimetypes
import glob
import os
# ログインユーザーを使用するmoduleをインポート
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
# postからの引数を使用できるmoduleをインポート
from django.views.decorators.http import require_POST
# excel操作
import openpyxl
# SQLSERVER接続
import pyodbc


# 酸化チタン_CL法_品質検討会資料
def quality_study_meeting_materials(request):
    # DB接続
    server = 'YSQLSERV4'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # SQL-トレース履歴銘柄一覧
    cursor.execute("SELECT [銘柄] FROM [T_TRACE_レポート引数_履歴] GROUP BY [銘柄]")
    tracereport1 = cursor.fetchall()

    # SQL-トレース履歴一覧
    cursor.execute("SELECT [銘柄], [ランNO] FROM [T_TRACE_レポート引数_履歴] GROUP BY [銘柄], [ランNO]")
    tracereport2 = cursor.fetchall()

    data = {
        'tracereport1': tracereport1,
        'tracereport2': tracereport2,
    }
    return render(request, 'report_output/CL法/品質検討会資料.html', data)


# 酸化チタン_CL法_品質検討会資料
def __select__brand(request):
    # 変数宣言
    value = request.POST['value']

    # DB接続
    server = 'YSQLSERV4'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # SQL-トレース履歴一覧
    cursor.execute("SELECT [銘柄], [ランNO] FROM [T_TRACE_REPORT_縦持ち_履歴] WHERE [銘柄] LIKE ? GROUP BY [銘柄], [ランNO] ORDER BY [ランNO] DESC", value)
    tracereport2 = cursor.fetchall()
    select = ''
    for tracereport2 in tracereport2:
        select += '<option value = ' + str(tracereport2.ランNO) + '>　' + str(tracereport2.ランNO) + '　</option>'

    data = {
        'select': select,
    }
    return JsonResponse(data)


# 酸化チタン_CL法_品質検討会資料
def trace_report(request, brand, run_no1, run_no2, run_no3):
    # 前回ファイル削除
    file_list = glob.glob("\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\品質検討会資料\\*品質検討会資料(CL法).xlsx")
    for file in file_list:
        # print("remove：{0}".format(file))
        os.remove(file)

    # DB接続
    server = 'YSQLSERV4'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # 操業データ_最大・最小・平均取得クエリ---------------------------------------------------------------------------------------------------------------------------
#     cursor.execute(" \
# SELECT \
#     REPLACE(項目, '_最大', '') AS 項目, \
#     MAX(CAST([MAX] AS NUMERIC(13,4))) AS [MAX], \
#     MIN(CAST([MIN] AS NUMERIC(13,4))) AS [MIN] \
# FROM ( \
#     SELECT \
#         日付, \
#         工程名, \
#         CASE p.seq \
#             WHEN 1 THEN 工程名 + '＿' + q.フィールド3 \
#             WHEN 5 THEN 工程名 + '＿' + q.フィールド7 \
#             WHEN 9 THEN 工程名 + '＿' + q.フィールド11 \
#             WHEN 17 THEN 工程名 + '＿' + q.フィールド19 \
#             WHEN 21 THEN 工程名 + '＿' + q.フィールド23 \
#             WHEN 25 THEN 工程名 + '＿' + q.フィールド27 \
#             WHEN 13 THEN 工程名 + '＿' + q.フィールド15 \
#             WHEN 29 THEN 工程名 + '＿' + q.フィールド31 \
#             WHEN 33 THEN 工程名 + '＿' + q.フィールド35 \
#             WHEN 37 THEN 工程名 + '＿' + q.フィールド39 \
#             WHEN 41 THEN 工程名 + '＿' + q.フィールド43 \
#             WHEN 45 THEN 工程名 + '＿' + q.フィールド47 \
#             WHEN 49 THEN 工程名 + '＿' + q.フィールド51 \
#             WHEN 53 THEN 工程名 + '＿' + q.フィールド55 \
#             WHEN 57 THEN 工程名 + '＿' + q.フィールド59 \
#             WHEN 61 THEN 工程名 + '＿' + q.フィールド63 \
#             WHEN 65 THEN 工程名 + '＿' + q.フィールド67 \
#             WHEN 69 THEN 工程名 + '＿' + q.フィールド71 \
#             WHEN 73 THEN 工程名 + '＿' + q.フィールド75 \
#             WHEN 77 THEN 工程名 + '＿' + q.フィールド79 \
#             WHEN 81 THEN 工程名 + '＿' + q.フィールド83 \
#             WHEN 85 THEN 工程名 + '＿' + q.フィールド87 \
#             WHEN 89 THEN 工程名 + '＿' + q.フィールド91 \
#             WHEN 93 THEN 工程名 + '＿' + q.フィールド95 \
#             WHEN 97 THEN 工程名 + '＿' + q.フィールド99 \
#             WHEN 101 THEN 工程名 + '＿' + q.フィールド103 \
#             WHEN 105 THEN 工程名 + '＿' + q.フィールド107 \
#             WHEN 109 THEN 工程名 + '＿' + q.フィールド111 \
#             WHEN 113 THEN 工程名 + '＿' + q.フィールド115 \
#             WHEN 117 THEN 工程名 + '＿' + q.フィールド119 \
#             WHEN 121 THEN 工程名 + '＿' + q.フィールド123 \
#             WHEN 125 THEN 工程名 + '＿' + q.フィールド127 \
#             WHEN 129 THEN 工程名 + '＿' + q.フィールド131 \
#             WHEN 133 THEN 工程名 + '＿' + q.フィールド135 \
#             WHEN 137 THEN 工程名 + '＿' + q.フィールド139 \
#             WHEN 141 THEN 工程名 + '＿' + q.フィールド143 \
#             WHEN 145 THEN 工程名 + '＿' + q.フィールド147 \
#         END AS 項目, \
#         CASE p.seq \
#             WHEN 1 THEN 3 \
#             WHEN 5 THEN 7 \
#             WHEN 9 THEN 11 \
#             WHEN 13 THEN 15 \
#             WHEN 17 THEN 19 \
#             WHEN 21 THEN 23 \
#             WHEN 25 THEN 27 \
#             WHEN 29 THEN 31 \
#             WHEN 33 THEN 35 \
#             WHEN 37 THEN 39 \
#             WHEN 41 THEN 43 \
#             WHEN 45 THEN 47 \
#             WHEN 49 THEN 51 \
#             WHEN 53 THEN 55 \
#             WHEN 57 THEN 59 \
#             WHEN 61 THEN 63 \
#             WHEN 65 THEN 67 \
#             WHEN 69 THEN 71 \
#             WHEN 73 THEN 75 \
#             WHEN 77 THEN 79 \
#             WHEN 81 THEN 83 \
#             WHEN 85 THEN 87 \
#             WHEN 89 THEN 91 \
#             WHEN 93 THEN 95 \
#             WHEN 97 THEN 99 \
#             WHEN 101 THEN 103 \
#             WHEN 105 THEN 107 \
#             WHEN 109 THEN 111 \
#             WHEN 113 THEN 115 \
#             WHEN 117 THEN 119 \
#             WHEN 121 THEN 123 \
#             WHEN 125 THEN 127 \
#             WHEN 129 THEN 131 \
#             WHEN 133 THEN 135 \
#             WHEN 137 THEN 139 \
#             WHEN 141 THEN 143 \
#             WHEN 145 THEN 147 \
#         END AS [フィールドindex(MAX)], \
#         CASE p.seq \
#             WHEN 1 THEN 4 \
#             WHEN 5 THEN 8 \
#             WHEN 9 THEN 12 \
#             WHEN 13 THEN 16 \
#             WHEN 17 THEN 20 \
#             WHEN 21 THEN 24 \
#             WHEN 25 THEN 28 \
#             WHEN 29 THEN 32 \
#             WHEN 33 THEN 36 \
#             WHEN 37 THEN 40 \
#             WHEN 41 THEN 44 \
#             WHEN 45 THEN 48 \
#             WHEN 49 THEN 52 \
#             WHEN 53 THEN 56 \
#             WHEN 57 THEN 60 \
#             WHEN 61 THEN 64 \
#             WHEN 65 THEN 68 \
#             WHEN 69 THEN 72 \
#             WHEN 73 THEN 76 \
#             WHEN 77 THEN 80 \
#             WHEN 81 THEN 84 \
#             WHEN 85 THEN 88 \
#             WHEN 89 THEN 92 \
#             WHEN 93 THEN 96 \
#             WHEN 97 THEN 100 \
#             WHEN 101 THEN 104 \
#             WHEN 105 THEN 108 \
#             WHEN 109 THEN 112 \
#             WHEN 113 THEN 116 \
#             WHEN 117 THEN 120 \
#             WHEN 121 THEN 124 \
#             WHEN 125 THEN 128 \
#             WHEN 129 THEN 132 \
#             WHEN 133 THEN 136 \
#             WHEN 137 THEN 140 \
#             WHEN 141 THEN 144 \
#             WHEN 145 THEN 148 \
#         END AS [フィールドindex(MIN)] \
#     FROM ( \
#         SELECT * \
#         FROM [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_履歴] \
#         WHERE (ランNO IN (?)) \
#             AND (item_value_flag = 1 OR item_value_flag = 3 OR item_value_flag = 6 OR item_value_flag = 9) \
#         ) AS q \
#     CROSS JOIN T_TRACE_REPORT_pivot AS p \
#     ) NAME \
# INNER JOIN ( \
#     SELECT * \
#     FROM ( \
#         SELECT \
#             日付, \
#             工程名, \
#             銘柄, \
#             トレースNO, \
#             CASE p.seq \
#                 WHEN 1 THEN q.フィールド3 \
#                 WHEN 5 THEN q.フィールド7 \
#                 WHEN 9 THEN q.フィールド11 \
#                 WHEN 13 THEN q.フィールド15 \
#                 WHEN 17 THEN q.フィールド19 \
#                 WHEN 21 THEN q.フィールド23 \
#                 WHEN 25 THEN q.フィールド27 \
#                 WHEN 29 THEN q.フィールド31 \
#                 WHEN 33 THEN q.フィールド35 \
#                 WHEN 37 THEN q.フィールド39 \
#                 WHEN 41 THEN q.フィールド43 \
#                 WHEN 45 THEN q.フィールド47 \
#                 WHEN 49 THEN q.フィールド51 \
#                 WHEN 53 THEN q.フィールド55 \
#                 WHEN 57 THEN q.フィールド59 \
#                 WHEN 61 THEN q.フィールド63 \
#                 WHEN 65 THEN q.フィールド67 \
#                 WHEN 69 THEN q.フィールド71 \
#                 WHEN 73 THEN q.フィールド75 \
#                 WHEN 77 THEN q.フィールド79 \
#                 WHEN 81 THEN q.フィールド83 \
#                 WHEN 85 THEN q.フィールド87 \
#                 WHEN 89 THEN q.フィールド91 \
#                 WHEN 93 THEN q.フィールド95 \
#                 WHEN 97 THEN q.フィールド99 \
#                 WHEN 101 THEN q.フィールド103 \
#                 WHEN 105 THEN q.フィールド107 \
#                 WHEN 109 THEN q.フィールド111 \
#                 WHEN 113 THEN q.フィールド115 \
#                 WHEN 117 THEN q.フィールド119 \
#                 WHEN 121 THEN q.フィールド123 \
#                 WHEN 125 THEN q.フィールド127 \
#                 WHEN 129 THEN q.フィールド131 \
#                 WHEN 133 THEN q.フィールド135 \
#                 WHEN 137 THEN q.フィールド139 \
#                 WHEN 141 THEN q.フィールド143 \
#                 WHEN 145 THEN q.フィールド147 \
#             END AS [MAX], \
#             CASE p.seq \
#                 WHEN 1 THEN q.フィールド4 \
#                 WHEN 5 THEN q.フィールド8 \
#                 WHEN 9 THEN q.フィールド12 \
#                 WHEN 13 THEN q.フィールド16 \
#                 WHEN 17 THEN q.フィールド20 \
#                 WHEN 21 THEN q.フィールド24 \
#                 WHEN 25 THEN q.フィールド28 \
#                 WHEN 29 THEN q.フィールド32 \
#                 WHEN 33 THEN q.フィールド36 \
#                 WHEN 37 THEN q.フィールド40 \
#                 WHEN 41 THEN q.フィールド44 \
#                 WHEN 45 THEN q.フィールド48 \
#                 WHEN 49 THEN q.フィールド52 \
#                 WHEN 53 THEN q.フィールド56 \
#                 WHEN 57 THEN q.フィールド60 \
#                 WHEN 61 THEN q.フィールド64 \
#                 WHEN 65 THEN q.フィールド68 \
#                 WHEN 69 THEN q.フィールド72 \
#                 WHEN 73 THEN q.フィールド76 \
#                 WHEN 77 THEN q.フィールド80 \
#                 WHEN 81 THEN q.フィールド84 \
#                 WHEN 85 THEN q.フィールド88 \
#                 WHEN 89 THEN q.フィールド92 \
#                 WHEN 93 THEN q.フィールド96 \
#                 WHEN 97 THEN q.フィールド100 \
#                 WHEN 101 THEN q.フィールド104 \
#                 WHEN 105 THEN q.フィールド108 \
#                 WHEN 109 THEN q.フィールド112 \
#                 WHEN 113 THEN q.フィールド116 \
#                 WHEN 117 THEN q.フィールド120 \
#                 WHEN 121 THEN q.フィールド124 \
#                 WHEN 125 THEN q.フィールド128 \
#                 WHEN 129 THEN q.フィールド132 \
#                 WHEN 133 THEN q.フィールド136 \
#                 WHEN 137 THEN q.フィールド140 \
#                 WHEN 141 THEN q.フィールド144 \
#                 WHEN 145 THEN q.フィールド148 \
#             END AS [MIN], \
#             CASE p.seq \
#                 WHEN 1 THEN 3 \
#                 WHEN 5 THEN 7 \
#                 WHEN 9 THEN 11 \
#                 WHEN 13 THEN 15 \
#                 WHEN 17 THEN 19 \
#                 WHEN 21 THEN 23 \
#                 WHEN 25 THEN 27 \
#                 WHEN 29 THEN 31 \
#                 WHEN 33 THEN 35 \
#                 WHEN 37 THEN 39 \
#                 WHEN 41 THEN 43 \
#                 WHEN 45 THEN 47 \
#                 WHEN 49 THEN 51 \
#                 WHEN 53 THEN 55 \
#                 WHEN 57 THEN 59 \
#                 WHEN 61 THEN 63 \
#                 WHEN 65 THEN 67 \
#                 WHEN 69 THEN 71 \
#                 WHEN 73 THEN 75 \
#                 WHEN 77 THEN 79 \
#                 WHEN 81 THEN 83 \
#                 WHEN 85 THEN 87 \
#                 WHEN 89 THEN 91 \
#                 WHEN 93 THEN 95 \
#                 WHEN 97 THEN 99 \
#                 WHEN 101 THEN 103 \
#                 WHEN 105 THEN 107 \
#                 WHEN 109 THEN 111 \
#                 WHEN 113 THEN 115 \
#                 WHEN 117 THEN 119 \
#                 WHEN 121 THEN 123 \
#                 WHEN 125 THEN 127 \
#                 WHEN 129 THEN 131 \
#                 WHEN 133 THEN 135 \
#                 WHEN 137 THEN 139 \
#                 WHEN 141 THEN 143 \
#                 WHEN 145 THEN 147 \
#             END AS [フィールドindex(MAX)], \
#             CASE p.seq \
#                 WHEN 1 THEN 4 \
#                 WHEN 5 THEN 8 \
#                 WHEN 9 THEN 12 \
#                 WHEN 13 THEN 16 \
#                 WHEN 17 THEN 20 \
#                 WHEN 21 THEN 24 \
#                 WHEN 25 THEN 28 \
#                 WHEN 29 THEN 32 \
#                 WHEN 33 THEN 36 \
#                 WHEN 37 THEN 40 \
#                 WHEN 41 THEN 44 \
#                 WHEN 45 THEN 48 \
#                 WHEN 49 THEN 52 \
#                 WHEN 53 THEN 56 \
#                 WHEN 57 THEN 60 \
#                 WHEN 61 THEN 64 \
#                 WHEN 65 THEN 68 \
#                 WHEN 69 THEN 72 \
#                 WHEN 73 THEN 76 \
#                 WHEN 77 THEN 80 \
#                 WHEN 81 THEN 84 \
#                 WHEN 85 THEN 88 \
#                 WHEN 89 THEN 92 \
#                 WHEN 93 THEN 96 \
#                 WHEN 97 THEN 100 \
#                 WHEN 101 THEN 104 \
#                 WHEN 105 THEN 108 \
#                 WHEN 109 THEN 112 \
#                 WHEN 113 THEN 116 \
#                 WHEN 117 THEN 120 \
#                 WHEN 121 THEN 124 \
#                 WHEN 125 THEN 128 \
#                 WHEN 129 THEN 132 \
#                 WHEN 133 THEN 136 \
#                 WHEN 137 THEN 140 \
#                 WHEN 141 THEN 144 \
#                 WHEN 145 THEN 148 \
#             END AS [フィールドindex(MIN)], \
#             [sort], \
#             [sort2] \
#         FROM ( \
#             SELECT * \
#             FROM [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_履歴] \
#             WHERE (ランNO IN (?)) \
#                 AND (item_value_flag = 2 OR item_value_flag = 4 OR item_value_flag = 7 OR item_value_flag = 10) \
#             ) AS q \
#         CROSS JOIN T_TRACE_REPORT_pivot AS p \
#         ) sub \
#     )DATA \
# ON NAME.日付 = DATA.日付 \
#     AND NAME.工程名 = DATA.工程名 \
#     AND NAME.[フィールドindex(MAX)] = DATA.[フィールドindex(MAX)] \
#     AND NAME.[フィールドindex(MIN)] = DATA.[フィールドindex(MIN)] \
# WHERE [項目] IS NOT NULL \
#     AND [銘柄] = ? \
# GROUP BY [項目], DATA.[sort], DATA.[sort2], DATA.[フィールドindex(MAX)] \
# ORDER BY DATA.[sort], DATA.[sort2], DATA.[フィールドindex(MAX)]", run_no1, run_no1, brand, )
    cursor.execute(" \
    SELECT \
                    CONCAT(IIF([M_QC-One_品検項目].[表記名] = '', '', CONCAT([M_QC-One_品検項目].[表記名], '_') ), [M_QC-One_品検項目].[出力項目] ) AS 出力項目 \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド2] AS 最小 \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド1] AS 最大 \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] AS 平均 \
    FROM            [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_縦持ち_履歴] \
    INNER JOIN      [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON              [T_TRACE_REPORT_縦持ち_履歴].[ID]		= [M_QC-One_品検項目].[SEQ] \
    AND			    [M_QC-One_品検項目].[工場]			= '0010' \
    AND			    [T_TRACE_REPORT_縦持ち_履歴].[項目名]	= [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	    [OP_ENTRY_INORG].[dbo].[T_TRACE_レポート引数_履歴] \
    ON			    [T_TRACE_レポート引数_履歴].[銘柄]		  = [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
    AND			    [T_TRACE_レポート引数_履歴].[ランNO]	  = [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
    AND 			[T_TRACE_レポート引数_履歴] .[銘柄]       = ? \
    AND			    [T_TRACE_レポート引数_履歴] .[ランNO]     = ? \
    GROUP BY	    CONCAT( IIF ( [M_QC-One_品検項目].[表記名] = '', '',CONCAT ( [M_QC-One_品検項目 ].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド1] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド2] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[ID] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日自] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日至] \
                    , [M_QC-One_品検項目].[出力順序] \
    ORDER BY        [M_QC-One_品検項目].[出力順序]", brand, run_no1, )
    tracereport = cursor.fetchall()
    aggregate_list = tracereport

    # 操業データ_前回平均取得クエリ-----------------------------------------------------------------------------------------------------------------------------
#     cursor.execute(" \
# SELECT \
#     REPLACE(項目, '_最大', '') AS 項目, \
#     AVG(CAST([AVG] AS NUMERIC(13,4))) AS [AVG] \
# FROM ( \
#     SELECT \
#         日付, \
#         工程名, \
#     CASE p.seq \
#         WHEN 1 THEN 工程名 + '＿' + q.フィールド3 \
#         WHEN 5 THEN 工程名 + '＿' + q.フィールド7 \
#         WHEN 9 THEN 工程名 + '＿' + q.フィールド11 \
#         WHEN 13 THEN 工程名 + '＿' + q.フィールド15 \
#         WHEN 17 THEN 工程名 + '＿' + q.フィールド19 \
#         WHEN 21 THEN 工程名 + '＿' + q.フィールド23 \
#         WHEN 25 THEN 工程名 + '＿' + q.フィールド27 \
#         WHEN 29 THEN 工程名 + '＿' + q.フィールド31 \
#         WHEN 33 THEN 工程名 + '＿' + q.フィールド35 \
#         WHEN 37 THEN 工程名 + '＿' + q.フィールド39 \
#         WHEN 41 THEN 工程名 + '＿' + q.フィールド43 \
#         WHEN 45 THEN 工程名 + '＿' + q.フィールド47 \
#         WHEN 49 THEN 工程名 + '＿' + q.フィールド51 \
#         WHEN 53 THEN 工程名 + '＿' + q.フィールド55 \
#         WHEN 57 THEN 工程名 + '＿' + q.フィールド59 \
#         WHEN 61 THEN 工程名 + '＿' + q.フィールド63 \
#         WHEN 65 THEN 工程名 + '＿' + q.フィールド67 \
#         WHEN 69 THEN 工程名 + '＿' + q.フィールド71 \
#         WHEN 73 THEN 工程名 + '＿' + q.フィールド75 \
#         WHEN 77 THEN 工程名 + '＿' + q.フィールド79 \
#         WHEN 81 THEN 工程名 + '＿' + q.フィールド83 \
#         WHEN 85 THEN 工程名 + '＿' + q.フィールド87 \
#         WHEN 89 THEN 工程名 + '＿' + q.フィールド91 \
#         WHEN 93 THEN 工程名 + '＿' + q.フィールド95 \
#         WHEN 97 THEN 工程名 + '＿' + q.フィールド99 \
#         WHEN 101 THEN 工程名 + '＿' + q.フィールド103 \
#         WHEN 105 THEN 工程名 + '＿' + q.フィールド107 \
#         WHEN 109 THEN 工程名 + '＿' + q.フィールド111 \
#         WHEN 113 THEN 工程名 + '＿' + q.フィールド115 \
#         WHEN 117 THEN 工程名 + '＿' + q.フィールド119 \
#         WHEN 121 THEN 工程名 + '＿' + q.フィールド123 \
#         WHEN 125 THEN 工程名 + '＿' + q.フィールド127 \
#         WHEN 129 THEN 工程名 + '＿' + q.フィールド131 \
#         WHEN 133 THEN 工程名 + '＿' + q.フィールド135 \
#         WHEN 137 THEN 工程名 + '＿' + q.フィールド139 \
#         WHEN 141 THEN 工程名 + '＿' + q.フィールド143 \
#         WHEN 145 THEN 工程名 + '＿' + q.フィールド147 \
#     END AS 項目, \
#     CASE p.seq \
#         WHEN 1 THEN 5 \
#         WHEN 5 THEN 9 \
#         WHEN 9 THEN 13 \
#         WHEN 13 THEN 17 \
#         WHEN 17 THEN 21 \
#         WHEN 21 THEN 25 \
#         WHEN 25 THEN 29 \
#         WHEN 29 THEN 33 \
#         WHEN 33 THEN 37 \
#         WHEN 37 THEN 41 \
#         WHEN 41 THEN 45 \
#         WHEN 45 THEN 49 \
#         WHEN 49 THEN 53 \
#         WHEN 53 THEN 57 \
#         WHEN 57 THEN 61 \
#         WHEN 61 THEN 65 \
#         WHEN 65 THEN 69 \
#         WHEN 69 THEN 73 \
#         WHEN 73 THEN 77 \
#         WHEN 77 THEN 81 \
#         WHEN 81 THEN 85 \
#         WHEN 85 THEN 89 \
#         WHEN 89 THEN 93 \
#         WHEN 93 THEN 97 \
#         WHEN 97 THEN 101 \
#         WHEN 101 THEN 105 \
#         WHEN 105 THEN 109 \
#         WHEN 109 THEN 113 \
#         WHEN 113 THEN 117 \
#         WHEN 117 THEN 121 \
#         WHEN 121 THEN 125 \
#         WHEN 125 THEN 129 \
#         WHEN 129 THEN 133 \
#         WHEN 133 THEN 137 \
#         WHEN 137 THEN 141 \
#         WHEN 141 THEN 145 \
#         WHEN 145 THEN 149 \
#         WHEN 149 THEN 153 \
#     END AS [フィールドindex(AVG)] \
#     FROM ( \
#         SELECT * \
#         FROM [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_履歴] \
#         WHERE (ランNO IN (?)) \
#             AND (item_value_flag = 1 OR item_value_flag = 3 OR item_value_flag = 6 OR item_value_flag = 9) \
#         ) AS q \
#     CROSS JOIN T_TRACE_REPORT_pivot AS p \
#     ) NAME \
# INNER JOIN ( \
#     SELECT * \
#     FROM ( \
#         SELECT \
#             日付, \
#             工程名, \
#             銘柄, \
#             トレースNO, \
#         CASE p.seq \
#             WHEN 1 THEN q.フィールド5 \
#             WHEN 5 THEN q.フィールド9 \
#             WHEN 9 THEN q.フィールド13 \
#             WHEN 13 THEN q.フィールド17 \
#             WHEN 17 THEN q.フィールド21 \
#             WHEN 21 THEN q.フィールド25 \
#             WHEN 25 THEN q.フィールド29 \
#             WHEN 29 THEN q.フィールド33 \
#             WHEN 33 THEN q.フィールド37 \
#             WHEN 37 THEN q.フィールド41 \
#             WHEN 41 THEN q.フィールド45 \
#             WHEN 45 THEN q.フィールド49 \
#             WHEN 49 THEN q.フィールド53 \
#             WHEN 53 THEN q.フィールド57 \
#             WHEN 57 THEN q.フィールド61 \
#             WHEN 61 THEN q.フィールド65 \
#             WHEN 65 THEN q.フィールド69 \
#             WHEN 69 THEN q.フィールド73 \
#             WHEN 73 THEN q.フィールド77 \
#             WHEN 77 THEN q.フィールド81 \
#             WHEN 81 THEN q.フィールド85 \
#             WHEN 85 THEN q.フィールド89 \
#             WHEN 89 THEN q.フィールド93 \
#             WHEN 93 THEN q.フィールド97 \
#             WHEN 97 THEN q.フィールド101 \
#             WHEN 101 THEN q.フィールド105 \
#             WHEN 105 THEN q.フィールド109 \
#             WHEN 109 THEN q.フィールド113 \
#             WHEN 113 THEN q.フィールド117 \
#             WHEN 117 THEN q.フィールド121 \
#             WHEN 121 THEN q.フィールド125 \
#             WHEN 125 THEN q.フィールド129 \
#             WHEN 129 THEN q.フィールド133 \
#             WHEN 133 THEN q.フィールド137 \
#             WHEN 137 THEN q.フィールド141 \
#             WHEN 141 THEN q.フィールド145 \
#             WHEN 145 THEN q.フィールド149 \
#         END AS [AVG], \
#         CASE p.seq \
#             WHEN 1 THEN 5 \
#             WHEN 5 THEN 9 \
#             WHEN 9 THEN 13 \
#             WHEN 13 THEN 17 \
#             WHEN 17 THEN 21 \
#             WHEN 21 THEN 25 \
#             WHEN 25 THEN 29 \
#             WHEN 29 THEN 33 \
#             WHEN 33 THEN 37 \
#             WHEN 37 THEN 41 \
#             WHEN 41 THEN 45 \
#             WHEN 45 THEN 49 \
#             WHEN 49 THEN 53 \
#             WHEN 53 THEN 57 \
#             WHEN 57 THEN 61 \
#             WHEN 61 THEN 65 \
#             WHEN 65 THEN 69 \
#             WHEN 69 THEN 73 \
#             WHEN 73 THEN 77 \
#             WHEN 77 THEN 81 \
#             WHEN 81 THEN 85 \
#             WHEN 85 THEN 89 \
#             WHEN 89 THEN 93 \
#             WHEN 93 THEN 97 \
#             WHEN 97 THEN 101 \
#             WHEN 101 THEN 105 \
#             WHEN 105 THEN 109 \
#             WHEN 109 THEN 113 \
#             WHEN 113 THEN 117 \
#             WHEN 117 THEN 121 \
#             WHEN 121 THEN 125 \
#             WHEN 125 THEN 129 \
#             WHEN 129 THEN 133 \
#             WHEN 133 THEN 137 \
#             WHEN 137 THEN 141 \
#             WHEN 141 THEN 145 \
#             WHEN 145 THEN 149 \
#             WHEN 149 THEN 153 \
#         END AS [フィールドindex(AVG)], \
#         [sort], \
#         [sort2] \
#         FROM ( \
#             SELECT * \
#             FROM [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_履歴] \
#             WHERE (ランNO IN (?)) \
#                 AND (item_value_flag = 2 OR item_value_flag = 4 OR item_value_flag = 7 OR item_value_flag = 10) \
#             ) AS q \
#         CROSS JOIN T_TRACE_REPORT_pivot AS p \
#         ) sub \
#     )DATA \
# ON NAME.日付 = DATA.日付 \
#     AND NAME.工程名 = DATA.工程名 \
#     AND NAME.[フィールドindex(AVG)] = DATA.[フィールドindex(AVG)] \
# WHERE [項目] IS NOT NULL \
#     AND [銘柄] = ? \
# GROUP BY [項目], DATA.[sort], DATA.[sort2], DATA.[フィールドindex(AVG)] \
# ORDER BY DATA.[sort], DATA.[sort2], DATA.[フィールドindex(AVG)]", run_no1, run_no1, brand, )
#     tracereport = cursor.fetchall()
#     avg1s_list = tracereport

    # 操業データ_前々回平均取得クエリ---------------------------------------------------------------------------------------------------------------------------
#     cursor.execute(" \
# SELECT \
#     REPLACE(項目, '_最大', '') AS 項目, \
#     AVG(CAST([AVG] AS NUMERIC(13,4))) AS [AVG] \
# FROM ( \
#     SELECT \
#         日付, \
#         工程名, \
#     CASE p.seq \
#         WHEN 1 THEN 工程名 + '＿' + q.フィールド3 \
#         WHEN 5 THEN 工程名 + '＿' + q.フィールド7 \
#         WHEN 9 THEN 工程名 + '＿' + q.フィールド11 \
#         WHEN 13 THEN 工程名 + '＿' + q.フィールド15 \
#         WHEN 17 THEN 工程名 + '＿' + q.フィールド19 \
#         WHEN 21 THEN 工程名 + '＿' + q.フィールド23 \
#         WHEN 25 THEN 工程名 + '＿' + q.フィールド27 \
#         WHEN 29 THEN 工程名 + '＿' + q.フィールド31 \
#         WHEN 33 THEN 工程名 + '＿' + q.フィールド35 \
#         WHEN 37 THEN 工程名 + '＿' + q.フィールド39 \
#         WHEN 41 THEN 工程名 + '＿' + q.フィールド43 \
#         WHEN 45 THEN 工程名 + '＿' + q.フィールド47 \
#         WHEN 49 THEN 工程名 + '＿' + q.フィールド51 \
#         WHEN 53 THEN 工程名 + '＿' + q.フィールド55 \
#         WHEN 57 THEN 工程名 + '＿' + q.フィールド59 \
#         WHEN 61 THEN 工程名 + '＿' + q.フィールド63 \
#         WHEN 65 THEN 工程名 + '＿' + q.フィールド67 \
#         WHEN 69 THEN 工程名 + '＿' + q.フィールド71 \
#         WHEN 73 THEN 工程名 + '＿' + q.フィールド75 \
#         WHEN 77 THEN 工程名 + '＿' + q.フィールド79 \
#         WHEN 81 THEN 工程名 + '＿' + q.フィールド83 \
#         WHEN 85 THEN 工程名 + '＿' + q.フィールド87 \
#         WHEN 89 THEN 工程名 + '＿' + q.フィールド91 \
#         WHEN 93 THEN 工程名 + '＿' + q.フィールド95 \
#         WHEN 97 THEN 工程名 + '＿' + q.フィールド99 \
#         WHEN 101 THEN 工程名 + '＿' + q.フィールド103 \
#         WHEN 105 THEN 工程名 + '＿' + q.フィールド107 \
#         WHEN 109 THEN 工程名 + '＿' + q.フィールド111 \
#         WHEN 113 THEN 工程名 + '＿' + q.フィールド115 \
#         WHEN 117 THEN 工程名 + '＿' + q.フィールド119 \
#         WHEN 121 THEN 工程名 + '＿' + q.フィールド123 \
#         WHEN 125 THEN 工程名 + '＿' + q.フィールド127 \
#         WHEN 129 THEN 工程名 + '＿' + q.フィールド131 \
#         WHEN 133 THEN 工程名 + '＿' + q.フィールド135 \
#         WHEN 137 THEN 工程名 + '＿' + q.フィールド139 \
#         WHEN 141 THEN 工程名 + '＿' + q.フィールド143 \
#         WHEN 145 THEN 工程名 + '＿' + q.フィールド147 \
#     END AS 項目, \
#     CASE p.seq \
#         WHEN 1 THEN 5 \
#         WHEN 5 THEN 9 \
#         WHEN 9 THEN 13 \
#         WHEN 13 THEN 17 \
#         WHEN 17 THEN 21 \
#         WHEN 21 THEN 25 \
#         WHEN 25 THEN 29 \
#         WHEN 29 THEN 33 \
#         WHEN 33 THEN 37 \
#         WHEN 37 THEN 41 \
#         WHEN 41 THEN 45 \
#         WHEN 45 THEN 49 \
#         WHEN 49 THEN 53 \
#         WHEN 53 THEN 57 \
#         WHEN 57 THEN 61 \
#         WHEN 61 THEN 65 \
#         WHEN 65 THEN 69 \
#         WHEN 69 THEN 73 \
#         WHEN 73 THEN 77 \
#         WHEN 77 THEN 81 \
#         WHEN 81 THEN 85 \
#         WHEN 85 THEN 89 \
#         WHEN 89 THEN 93 \
#         WHEN 93 THEN 97 \
#         WHEN 97 THEN 101 \
#         WHEN 101 THEN 105 \
#         WHEN 105 THEN 109 \
#         WHEN 109 THEN 113 \
#         WHEN 113 THEN 117 \
#         WHEN 117 THEN 121 \
#         WHEN 121 THEN 125 \
#         WHEN 125 THEN 129 \
#         WHEN 129 THEN 133 \
#         WHEN 133 THEN 137 \
#         WHEN 137 THEN 141 \
#         WHEN 141 THEN 145 \
#         WHEN 145 THEN 149 \
#         WHEN 149 THEN 153 \
#     END AS [フィールドindex(AVG)] \
#     FROM ( \
#         SELECT * \
#         FROM [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_履歴] \
#         WHERE (ランNO IN (?)) \
#             AND (item_value_flag = 1 OR item_value_flag = 3 OR item_value_flag = 6 OR item_value_flag = 9) \
#         ) AS q \
#     CROSS JOIN T_TRACE_REPORT_pivot as p \
#     ) NAME \
# INNER JOIN ( \
#     SELECT * \
#     FROM ( \
#         SELECT \
#             日付, \
#             工程名, \
#             銘柄, \
#             トレースNO, \
#         CASE p.seq \
#             WHEN 1 THEN q.フィールド5 \
#             WHEN 5 THEN q.フィールド9 \
#             WHEN 9 THEN q.フィールド13 \
#             WHEN 13 THEN q.フィールド17 \
#             WHEN 17 THEN q.フィールド21 \
#             WHEN 21 THEN q.フィールド25 \
#             WHEN 25 THEN q.フィールド29 \
#             WHEN 29 THEN q.フィールド33 \
#             WHEN 33 THEN q.フィールド37 \
#             WHEN 37 THEN q.フィールド41 \
#             WHEN 41 THEN q.フィールド45 \
#             WHEN 45 THEN q.フィールド49 \
#             WHEN 49 THEN q.フィールド53 \
#             WHEN 53 THEN q.フィールド57 \
#             WHEN 57 THEN q.フィールド61 \
#             WHEN 61 THEN q.フィールド65 \
#             WHEN 65 THEN q.フィールド69 \
#             WHEN 69 THEN q.フィールド73 \
#             WHEN 73 THEN q.フィールド77 \
#             WHEN 77 THEN q.フィールド81 \
#             WHEN 81 THEN q.フィールド85 \
#             WHEN 85 THEN q.フィールド89 \
#             WHEN 89 THEN q.フィールド93 \
#             WHEN 93 THEN q.フィールド97 \
#             WHEN 97 THEN q.フィールド101 \
#             WHEN 101 THEN q.フィールド105 \
#             WHEN 105 THEN q.フィールド109 \
#             WHEN 109 THEN q.フィールド113 \
#             WHEN 113 THEN q.フィールド117 \
#             WHEN 117 THEN q.フィールド121 \
#             WHEN 121 THEN q.フィールド125 \
#             WHEN 125 THEN q.フィールド129 \
#             WHEN 129 THEN q.フィールド133 \
#             WHEN 133 THEN q.フィールド137 \
#             WHEN 137 THEN q.フィールド141 \
#             WHEN 141 THEN q.フィールド145 \
#             WHEN 145 THEN q.フィールド149 \
#         END AS [AVG], \
#         CASE p.seq \
#             WHEN 1 THEN 5 \
#             WHEN 5 THEN 9 \
#             WHEN 9 THEN 13 \
#             WHEN 13 THEN 17 \
#             WHEN 17 THEN 21 \
#             WHEN 21 THEN 25 \
#             WHEN 25 THEN 29 \
#             WHEN 29 THEN 33 \
#             WHEN 33 THEN 37 \
#             WHEN 37 THEN 41 \
#             WHEN 41 THEN 45 \
#             WHEN 45 THEN 49 \
#             WHEN 49 THEN 53 \
#             WHEN 53 THEN 57 \
#             WHEN 57 THEN 61 \
#             WHEN 61 THEN 65 \
#             WHEN 65 THEN 69 \
#             WHEN 69 THEN 73 \
#             WHEN 73 THEN 77 \
#             WHEN 77 THEN 81 \
#             WHEN 81 THEN 85 \
#             WHEN 85 THEN 89 \
#             WHEN 89 THEN 93 \
#             WHEN 93 THEN 97 \
#             WHEN 97 THEN 101 \
#             WHEN 101 THEN 105 \
#             WHEN 105 THEN 109 \
#             WHEN 109 THEN 113 \
#             WHEN 113 THEN 117 \
#             WHEN 117 THEN 121 \
#             WHEN 121 THEN 125 \
#             WHEN 125 THEN 129 \
#             WHEN 129 THEN 133 \
#             WHEN 133 THEN 137 \
#             WHEN 137 THEN 141 \
#             WHEN 141 THEN 145 \
#             WHEN 145 THEN 149 \
#             WHEN 149 THEN 153 \
#         END AS [フィールドindex(AVG)], \
#         [sort], \
#         [sort2] \
#         FROM ( \
#             SELECT * \
#             FROM [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_履歴] \
#             WHERE (ランNO IN (?)) \
#                 AND (item_value_flag = 2 OR item_value_flag = 4 OR item_value_flag = 7 OR item_value_flag = 10) \
#             ) AS q \
#         CROSS JOIN T_TRACE_REPORT_pivot AS p \
#     ) sub \
#     )DATA \
# ON NAME.日付 = DATA.日付 \
#     AND NAME.工程名 = DATA.工程名 \
#     AND NAME.[フィールドindex(AVG)] = DATA.[フィールドindex(AVG)] \
# WHERE [項目] IS NOT NULL \
#     AND [銘柄] = ? \
# GROUP BY [項目], DATA.[sort], DATA.[sort2], DATA.[フィールドindex(AVG)] \
# ORDER BY DATA.[sort], DATA.[sort2], DATA.[フィールドindex(AVG)]", run_no2, brand, )
    cursor.execute(" \
    SELECT \
                    CONCAT(IIF([M_QC-One_品検項目].[表記名] = '', '', CONCAT([M_QC-One_品検項目].[表記名], '_') ), [M_QC-One_品検項目].[出力項目] ) AS 出力項目 \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] AS 平均 \
    FROM		    [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_縦持ち_履歴] \
    INNER JOIN	    [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON			    [T_TRACE_REPORT_縦持ち_履歴].[ID] = [M_QC-One_品検項目].[SEQ] \
    AND			    [M_QC-One_品検項目].[工場] = '0010' \
    AND			    [T_TRACE_REPORT_縦持ち_履歴].[項目名 ] = [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	    [OP_ENTRY_INORG].[dbo].[T_TRACE_レポート引数_履歴] \
    ON			    [T_TRACE_レポート引数_履歴].[銘柄]    = [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
    AND			    [T_TRACE_レポート引数_履歴].[ランNO]  = [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
    AND			    [T_TRACE_レポート引数_履歴] .[銘柄]   = ? \
    AND			    [T_TRACE_レポート引数_履歴] .[ランNO] = ? \
    GROUP BY        CONCAT( IIF ( [M_QC-One_品検項目].[表記名] = '', '',CONCAT ( [M_QC-One_品検項目 ].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド1] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド2] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[ID] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日自] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日至] \
                  , [M_QC-One_品検項目].[出力順序] \
    ORDER BY	    [M_QC-One_品検項目].[出力順序]", brand, run_no2, )
    tracereport = cursor.fetchall()
    avg2s_list = tracereport

    # 操業データ_前々々回平均取得クエリ-------------------------------------------------------------------------------------------------------------------------
#     cursor.execute(" \
# SELECT \
#     REPLACE(項目, '_最大', '') AS 項目, \
#     AVG(CAST([AVG] AS NUMERIC(13,4))) AS [AVG] \
#     FROM ( \
#         SELECT \
#             日付, \
#             工程名, \
#         CASE p.seq \
#             WHEN 1 THEN 工程名 + '＿' + q.フィールド3 \
#             WHEN 5 THEN 工程名 + '＿' + q.フィールド7 \
#             WHEN 9 THEN 工程名 + '＿' + q.フィールド11 \
#             WHEN 13 THEN 工程名 + '＿' + q.フィールド15 \
#             WHEN 17 THEN 工程名 + '＿' + q.フィールド19 \
#             WHEN 21 THEN 工程名 + '＿' + q.フィールド23 \
#             WHEN 25 THEN 工程名 + '＿' + q.フィールド27 \
#             WHEN 29 THEN 工程名 + '＿' + q.フィールド31 \
#             WHEN 33 THEN 工程名 + '＿' + q.フィールド35 \
#             WHEN 37 THEN 工程名 + '＿' + q.フィールド39 \
#             WHEN 41 THEN 工程名 + '＿' + q.フィールド43 \
#             WHEN 45 THEN 工程名 + '＿' + q.フィールド47 \
#             WHEN 49 THEN 工程名 + '＿' + q.フィールド51 \
#             WHEN 53 THEN 工程名 + '＿' + q.フィールド55 \
#             WHEN 57 THEN 工程名 + '＿' + q.フィールド59 \
#             WHEN 61 THEN 工程名 + '＿' + q.フィールド63 \
#             WHEN 65 THEN 工程名 + '＿' + q.フィールド67 \
#             WHEN 69 THEN 工程名 + '＿' + q.フィールド71 \
#             WHEN 73 THEN 工程名 + '＿' + q.フィールド75 \
#             WHEN 77 THEN 工程名 + '＿' + q.フィールド79 \
#             WHEN 81 THEN 工程名 + '＿' + q.フィールド83 \
#             WHEN 85 THEN 工程名 + '＿' + q.フィールド87 \
#             WHEN 89 THEN 工程名 + '＿' + q.フィールド91 \
#             WHEN 93 THEN 工程名 + '＿' + q.フィールド95 \
#             WHEN 97 THEN 工程名 + '＿' + q.フィールド99 \
#             WHEN 101 THEN 工程名 + '＿' + q.フィールド103 \
#             WHEN 105 THEN 工程名 + '＿' + q.フィールド107 \
#             WHEN 109 THEN 工程名 + '＿' + q.フィールド111 \
#             WHEN 113 THEN 工程名 + '＿' + q.フィールド115 \
#             WHEN 117 THEN 工程名 + '＿' + q.フィールド119 \
#             WHEN 121 THEN 工程名 + '＿' + q.フィールド123 \
#             WHEN 125 THEN 工程名 + '＿' + q.フィールド127 \
#             WHEN 129 THEN 工程名 + '＿' + q.フィールド131 \
#             WHEN 133 THEN 工程名 + '＿' + q.フィールド135 \
#             WHEN 137 THEN 工程名 + '＿' + q.フィールド139 \
#             WHEN 141 THEN 工程名 + '＿' + q.フィールド143 \
#             WHEN 145 THEN 工程名 + '＿' + q.フィールド147 \
#         END AS 項目, \
#         CASE p.seq \
#             WHEN 1 THEN 5 \
#             WHEN 5 THEN 9 \
#             WHEN 9 THEN 13 \
#             WHEN 13 THEN 17 \
#             WHEN 17 THEN 21 \
#             WHEN 21 THEN 25 \
#             WHEN 25 THEN 29 \
#             WHEN 29 THEN 33 \
#             WHEN 33 THEN 37 \
#             WHEN 37 THEN 41 \
#             WHEN 41 THEN 45 \
#             WHEN 45 THEN 49 \
#             WHEN 49 THEN 53 \
#             WHEN 53 THEN 57 \
#             WHEN 57 THEN 61 \
#             WHEN 61 THEN 65 \
#             WHEN 65 THEN 69 \
#             WHEN 69 THEN 73 \
#             WHEN 73 THEN 77 \
#             WHEN 77 THEN 81 \
#             WHEN 81 THEN 85 \
#             WHEN 85 THEN 89 \
#             WHEN 89 THEN 93 \
#             WHEN 93 THEN 97 \
#             WHEN 97 THEN 101 \
#             WHEN 101 THEN 105 \
#             WHEN 105 THEN 109 \
#             WHEN 109 THEN 113 \
#             WHEN 113 THEN 117 \
#             WHEN 117 THEN 121 \
#             WHEN 121 THEN 125 \
#             WHEN 125 THEN 129 \
#             WHEN 129 THEN 133 \
#             WHEN 133 THEN 137 \
#             WHEN 137 THEN 141 \
#             WHEN 141 THEN 145 \
#             WHEN 145 THEN 149 \
#             WHEN 149 THEN 153 \
#         END AS [フィールドindex(AVG)] \
#         FROM ( \
#             SELECT * \
#             FROM [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_履歴] \
#             WHERE (ランNO IN (?)) \
#                 AND (item_value_flag = 1 OR item_value_flag = 3 OR item_value_flag = 6 OR item_value_flag = 9) \
#             ) AS q \
#         CROSS JOIN T_TRACE_REPORT_pivot as p \
#     ) NAME \
# INNER JOIN ( \
#     SELECT * \
#     FROM ( \
#         SELECT \
#             日付, \
#             工程名, \
#             銘柄, \
#             トレースNO, \
#         CASE p.seq \
#             WHEN 1 THEN q.フィールド5 \
#             WHEN 5 THEN q.フィールド9 \
#             WHEN 9 THEN q.フィールド13 \
#             WHEN 13 THEN q.フィールド17 \
#             WHEN 17 THEN q.フィールド21 \
#             WHEN 21 THEN q.フィールド25 \
#             WHEN 25 THEN q.フィールド29 \
#             WHEN 29 THEN q.フィールド33 \
#             WHEN 33 THEN q.フィールド37 \
#             WHEN 37 THEN q.フィールド41 \
#             WHEN 41 THEN q.フィールド45 \
#             WHEN 45 THEN q.フィールド49 \
#             WHEN 49 THEN q.フィールド53 \
#             WHEN 53 THEN q.フィールド57 \
#             WHEN 57 THEN q.フィールド61 \
#             WHEN 61 THEN q.フィールド65 \
#             WHEN 65 THEN q.フィールド69 \
#             WHEN 69 THEN q.フィールド73 \
#             WHEN 73 THEN q.フィールド77 \
#             WHEN 77 THEN q.フィールド81 \
#             WHEN 81 THEN q.フィールド85 \
#             WHEN 85 THEN q.フィールド89 \
#             WHEN 89 THEN q.フィールド93 \
#             WHEN 93 THEN q.フィールド97 \
#             WHEN 97 THEN q.フィールド101 \
#             WHEN 101 THEN q.フィールド105 \
#             WHEN 105 THEN q.フィールド109 \
#             WHEN 109 THEN q.フィールド113 \
#             WHEN 113 THEN q.フィールド117 \
#             WHEN 117 THEN q.フィールド121 \
#             WHEN 121 THEN q.フィールド125 \
#             WHEN 125 THEN q.フィールド129 \
#             WHEN 129 THEN q.フィールド133 \
#             WHEN 133 THEN q.フィールド137 \
#             WHEN 137 THEN q.フィールド141 \
#             WHEN 141 THEN q.フィールド145 \
#             WHEN 145 THEN q.フィールド149 \
#         END AS [AVG], \
#         CASE p.seq \
#             WHEN 1 THEN 5 \
#             WHEN 5 THEN 9 \
#             WHEN 9 THEN 13 \
#             WHEN 13 THEN 17 \
#             WHEN 17 THEN 21 \
#             WHEN 21 THEN 25 \
#             WHEN 25 THEN 29 \
#             WHEN 29 THEN 33 \
#             WHEN 33 THEN 37 \
#             WHEN 37 THEN 41 \
#             WHEN 41 THEN 45 \
#             WHEN 45 THEN 49 \
#             WHEN 49 THEN 53 \
#             WHEN 53 THEN 57 \
#             WHEN 57 THEN 61 \
#             WHEN 61 THEN 65 \
#             WHEN 65 THEN 69 \
#             WHEN 69 THEN 73 \
#             WHEN 73 THEN 77 \
#             WHEN 77 THEN 81 \
#             WHEN 81 THEN 85 \
#             WHEN 85 THEN 89 \
#             WHEN 89 THEN 93 \
#             WHEN 93 THEN 97 \
#             WHEN 97 THEN 101 \
#             WHEN 101 THEN 105 \
#             WHEN 105 THEN 109 \
#             WHEN 109 THEN 113 \
#             WHEN 113 THEN 117 \
#             WHEN 117 THEN 121 \
#             WHEN 121 THEN 125 \
#             WHEN 125 THEN 129 \
#             WHEN 129 THEN 133 \
#             WHEN 133 THEN 137 \
#             WHEN 137 THEN 141 \
#             WHEN 141 THEN 145 \
#             WHEN 145 THEN 149 \
#             WHEN 149 THEN 153 \
#         END AS [フィールドindex(AVG)], \
#         [sort], \
#         [sort2] \
#         FROM ( \
#             SELECT * \
#             FROM [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_履歴] \
#             WHERE (ランNO IN (?)) \
#                 AND (item_value_flag = 2 OR item_value_flag = 4 OR item_value_flag = 7 OR item_value_flag = 10) \
#             ) AS q \
#         CROSS JOIN T_TRACE_REPORT_pivot AS p \
#         ) sub \
#     )DATA \
# ON NAME.日付 = DATA.日付 \
#     AND NAME.工程名 = DATA.工程名 \
#     AND NAME.[フィールドindex(AVG)] = DATA.[フィールドindex(AVG)] \
# WHERE [項目] IS NOT NULL \
#     AND [銘柄] = ? \
# GROUP BY [項目], DATA.[sort], DATA.[sort2], DATA.[フィールドindex(AVG)] \
# ORDER BY DATA.[sort], DATA.[sort2], DATA.[フィールドindex(AVG)]", run_no3, brand, )
    cursor.execute(" \
    SELECT \
                    CONCAT(IIF([M_QC-One_品検項目].[表記名] = '', '', CONCAT([M_QC-One_品検項目].[表記名], '_') ), [M_QC-One_品検項目].[出力項目] ) AS 出力項目 \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] AS 平均 \
    FROM		    [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_縦持ち_履歴] \
    INNER JOIN	    [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON			    [T_TRACE_REPORT_縦持ち_履歴].[ID] = [M_QC-One_品検項目].[SEQ] \
    AND			    [M_QC-One_品検項目].[工場] = '0010' \
    AND			    [T_TRACE_REPORT_縦持ち_履歴].[項目名 ] = [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	    [OP_ENTRY_INORG].[dbo].[T_TRACE_レポート引数_履歴] \
    ON			    [T_TRACE_レポート引数_履歴].[銘柄] = [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
    AND			    [T_TRACE_レポート引数_履歴].[ランNO] = [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
    AND			    [T_TRACE_レポート引数_履歴] .[銘柄]		= ? \
    AND			    [T_TRACE_レポート引数_履歴] .[ランNO]		= ? \
    GROUP BY        CONCAT( IIF ( [M_QC-One_品検項目].[表記名] = '', '',CONCAT ( [M_QC-One_品検項目 ].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド1] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド2] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[ID] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日自] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日至] \
                  , [M_QC-One_品検項目].[出力順序] \
    ORDER BY	    [M_QC-One_品検項目].[出力順序]", brand, run_no3, )
    tracereport = cursor.fetchall()
    avg3s_list = tracereport

    cursor.execute(" \
    SELECT \
        TOP 1 \
        操業年月日自, \
        操業年月日至 \
    FROM \
        [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_縦持ち_履歴] \
    WHERE \
        銘柄 = ? \
        AND ランNO = ? \
    ", brand, run_no1, )
    tracereport = cursor.fetchall()
    for tracereport in tracereport:
        operation_date_start = tracereport.操業年月日自.strftime('%Y%m%d%H%M')
        operation_date_start = operation_date_start[0:4] + "/" + operation_date_start[4:6] + "/" + operation_date_start[6:8]
        operation_date_end = tracereport.操業年月日至.strftime('%Y%m%d%H%M')
        operation_date_end = operation_date_end[0:4] + "/" + operation_date_end[4:6] + "/" + operation_date_end[6:8]

    print('データ取得完了')

    # 現在日時を取得。ファイル名に加える --- (*1)
    today = datetime.datetime.now().strftime('%Y%m%d%H%M')

    # EXCEL読込 --- (*2)
    template = '\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\テンプレート\\【帳票テンプレート】品質検討会資料(CL法).xlsx'
    print(template)
    wb = openpyxl.load_workbook(template)

    # アクティブなワークシートを得る --- (*3)
    ws = wb["品質検討会資料"]
    ws2 = wb["form"]
    ws3 = wb["pos"]

    # データ書き込み --- (*4)
    ws["F1"] = "XXXX年    X月度    " + brand + "    XX工程"
    ws["A2"] = "（ " + str(run_no1) + " ）前    回    ラ    ン    概    況 [ " + operation_date_start + " ～ " + operation_date_end + " ]"

    for i in range(5, aggregate_list.__len__()+5, 1):
        ws['A' + str(i)] = aggregate_list[i-5][0]
        ws['F' + str(i)] = aggregate_list[i-5][1]
        ws['H' + str(i)] = aggregate_list[i-5][2]
        ws['J' + str(i)] = aggregate_list[i-5][3]
        # for j in range(0, avg1s_list.__len__(), 1):
        #     if aggregate_list[i-5][0] == avg1s_list[j][0]:
        #         ws['J' + str(i)] = avg1s_list[j][1]
        for j in range(0, avg2s_list.__len__(), 1):
            if aggregate_list[i-5][0] == avg2s_list[j][0]:
                ws['L' + str(i)] = avg2s_list[j][1]
        for j in range(0, avg3s_list.__len__(), 1):
            if aggregate_list[i-5][0] == avg3s_list[j][0]:
                ws['N' + str(i)] = avg3s_list[j][1]
        # 自動集計部最大項目数以上は取得不可 テンプレートに収まりきらないので切り捨てる。
        if i == 64:
            break

    print('データ抽出完了')

    # パターンの取得 ["pos"]から銘柄比較してパターン取得
    pattern = ''
    for row in ws3.iter_rows(min_row=51, max_row=451, min_col=2, max_col=2):
        for data in row:
            if data.value == brand:
                pattern = ws3.cell(row=data.row, column=data.column + 1).value

    # パターン毎に処理分岐
    if pattern == 'A':
        for i in range(1, 8):
            for j in range(1, 18):
                copy = ws2.cell(row=i+23, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'B':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+32, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'C':
        for i in range(1, 12):
            for j in range(1, 18):
                copy = ws2.cell(row=i+41, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'D':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+53, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'E':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+62, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'F':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+66, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'G':
        for i in range(1, 11):
            for j in range(1, 18):
                copy = ws2.cell(row=i+70, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'H':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+81, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'I':
        for i in range(1, 14):
            for j in range(1, 18):
                copy = ws2.cell(row=i+85, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'J':
        for i in range(1, 7):
            for j in range(1, 18):
                copy = ws2.cell(row=i+99, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'K':
        for i in range(1, 8):
            for j in range(1, 18):
                copy = ws2.cell(row=i+106, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'L':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+114, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'M':
        for i in range(1, 5):
            for j in range(1, 18):
                copy = ws2.cell(row=i+123, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'N':
        for i in range(1, 5):
            for j in range(1, 18):
                copy = ws2.cell(row=i+128, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)

    # header
    ws.oddHeader.right.text = today[0:4] + "年" + today[4:6] + "月" + today[6:8] + "日\n"

    # 署名
    img = openpyxl.drawing.image.Image("\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\テンプレート\\署名.png")
    ws.add_image(img, 'F92')

    # ファイルを保存 --- (*5)
    file_path = '\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\品質検討会資料\\' + today + '_品質検討会資料(CL法).xlsx'
    file_name = today + '_品質検討会資料(CL法).xlsx'
    wb.save(file_path)

    with open(file_path, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(file_name)[0] or 'application/octet-stream')
    return response


# 酸化チタン_S法_品質検討会資料
def s_quality_study_meeting_materials(request):
    # DB接続
    server = 'YSQLSERV4'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # SQL-トレース履歴銘柄一覧
    cursor.execute("SELECT [銘柄] FROM [T_TRACE_S法_レポート引数_履歴] GROUP BY [銘柄]")
    tracereport1 = cursor.fetchall()

    # SQL-トレース履歴一覧
    cursor.execute("SELECT [銘柄], [ランNO] FROM [T_TRACE_S法_レポート引数_履歴] GROUP BY [銘柄], [ランNO]")
    tracereport2 = cursor.fetchall()

    data = {
        'tracereport1': tracereport1,
        'tracereport2': tracereport2,
    }
    return render(request, 'report_output/S法/品質検討会資料.html', data)


# 酸化チタン_S法_品質検討会資料
def s___select__brand(request):
    # 変数宣言
    value = request.POST['value']

    # DB接続
    server = 'YSQLSERV4'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # SQL-トレース履歴一覧
    cursor.execute("SELECT [銘柄], [ランNO] FROM [T_TRACE_S法_REPORT_縦持ち_履歴] WHERE [銘柄] LIKE ? GROUP BY [銘柄], [ランNO] ORDER BY [ランNO] DESC", value)
    tracereport2 = cursor.fetchall()
    select = ''
    for tracereport2 in tracereport2:
        select += '<option value = ' + str(tracereport2.ランNO) + '>　' + str(tracereport2.ランNO) + '　</option>'

    data = {
        'select': select,
    }
    return JsonResponse(data)


# 酸化チタン_S法_品質検討会資料
def s_trace_report(request, brand, run_no1, run_no2, run_no3):

    # 前回ファイル削除
    file_list = glob.glob("\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\品質検討会資料\\*品質検討会資料(S法).xlsx")
    for file in file_list:
        # print("remove：{0}".format(file))
        os.remove(file)

    # DB接続
    server = 'YSQLSERV4'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    cursor.execute(" \
    SELECT \
                  CONCAT( IIF( [M_QC-One_品検項目].[表記名] = '', '', CONCAT( [M_QC-One_品検項目].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] )	AS 出力項目 \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド2]			AS 最小 \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド1]			AS 最大	\
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3]			AS 平均 \
    FROM		  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_REPORT_縦持ち_履歴] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON			  [T_TRACE_S法_REPORT_縦持ち_履歴].[ID]		= [M_QC-One_品検項目].[SEQ] \
    AND			  [M_QC-One_品検項目].[工場]			= '0020' \
    AND			  [T_TRACE_S法_REPORT_縦持ち_履歴].[項目名]	= [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_レポート引数_履歴] \
    ON			  [T_TRACE_S法_レポート引数_履歴].[銘柄]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
    AND			  [T_TRACE_S法_レポート引数_履歴].[ランNO]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
    AND			  [T_TRACE_S法_レポート引数_履歴].[銘柄]		= ? \
    AND			  [T_TRACE_S法_レポート引数_履歴].[ランNO]		= ? \
    GROUP BY	  CONCAT( IIF( [M_QC-One_品検項目].[表記名] = '', '', CONCAT( [M_QC-One_品検項目].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド1] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド2] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ID] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日自] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日至] \
                , [M_QC-One_品検項目].[出力順序] \
    ORDER BY	  [M_QC-One_品検項目].[出力順序]", brand, run_no1, )
    tracereport = cursor.fetchall()
    aggregate_list = tracereport

    cursor.execute(" \
    SELECT \
                  CONCAT(IIF([M_QC-One_品検項目].[表記名] = '', '', CONCAT([M_QC-One_品検項目].[表記名], '_') ), [M_QC-One_品検項目].[出力項目] ) AS 出力項目 \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3] AS 平均 \
    FROM		  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_REPORT_縦持ち_履歴] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON			  [T_TRACE_S法_REPORT_縦持ち_履歴].[ID]		= [M_QC-One_品検項目].[SEQ] \
    AND			  [M_QC-One_品検項目].[工場]			= '0020' \
    AND			  [T_TRACE_S法_REPORT_縦持ち_履歴].[項目名 ]	= [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_レポート引数_履歴] \
    ON			  [T_TRACE_S法_レポート引数_履歴].[銘柄 ]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
    AND			  [T_TRACE_S法_レポート引数_履歴].[ランNO]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
    AND			  [T_TRACE_S法_レポート引数_履歴] .[銘柄]		= ? \
    AND			  [T_TRACE_S法_レポート引数_履歴] .[ランNO]		= ? \
    GROUP BY	  CONCAT( IIF ( [M_QC-One_品検項目].[表記名] = '', '',CONCAT ( [M_QC-One_品検項目].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド1] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド2] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ID] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日自] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日至] \
                , [M_QC-One_品検項目].[出力順序] \
    ORDER BY	  [M_QC-One_品検項目].[出力順序]", brand, run_no2, )
    tracereport = cursor.fetchall()
    avg2s_list = tracereport

    cursor.execute(" \
    SELECT \
                  CONCAT(IIF([M_QC-One_品検項目].[表記名] = '', '', CONCAT([M_QC-One_品検項目].[表記名], '_') ), [M_QC-One_品検項目].[出力項目] ) AS 出力項目 \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3] AS 平均 \
    FROM		  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_REPORT_縦持ち_履歴] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON			  [T_TRACE_S法_REPORT_縦持ち_履歴].[ID]		= [M_QC-One_品検項目].[SEQ] \
    AND			  [M_QC-One_品検項目].[工場]			= '0020' \
    AND			  [T_TRACE_S法_REPORT_縦持ち_履歴].[項目名 ]	= [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_レポート引数_履歴] \
    ON			  [T_TRACE_S法_レポート引数_履歴].[銘柄 ]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
    AND			  [T_TRACE_S法_レポート引数_履歴].[ランNO]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
    AND			  [T_TRACE_S法_レポート引数_履歴] .[銘柄]		= ? \
    AND			  [T_TRACE_S法_レポート引数_履歴] .[ランNO]		= ? \
    GROUP BY	  CONCAT( IIF ( [M_QC-One_品検項目].[表記名] = '', '',CONCAT ( [M_QC-One_品検項目].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド1] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド2] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ID] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日自] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日至] \
                , [M_QC-One_品検項目].[出力順序] \
    ORDER BY	  [M_QC-One_品検項目].[出力順序]", brand, run_no3, )
    tracereport = cursor.fetchall()
    avg3s_list = tracereport

    cursor.execute(" \
    SELECT \
        TOP 1 \
        操業年月日自, \
        操業年月日至 \
    FROM \
        [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_REPORT_縦持ち_履歴] \
    WHERE \
        銘柄 = ? \
        AND ランNO = ? \
    ", brand, run_no1, )
    tracereport = cursor.fetchall()
    for tracereport in tracereport:
        operation_date_start = tracereport.操業年月日自.strftime('%Y%m%d%H%M')
        operation_date_start = operation_date_start[0:4] + "/" + operation_date_start[4:6] + "/" + operation_date_start[6:8]
        operation_date_end = tracereport.操業年月日至.strftime('%Y%m%d%H%M')
        operation_date_end = operation_date_end[0:4] + "/" + operation_date_end[4:6] + "/" + operation_date_end[6:8]

    print('データ取得完了')

    # 現在日時を取得。ファイル名に加える --- (*1)
    today = datetime.datetime.now().strftime('%Y%m%d%H%M')

    # EXCEL読込 --- (*2)
    template = '\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\テンプレート\\【帳票テンプレート】品質検討会資料(S法).xlsx'
    print(template)
    wb = openpyxl.load_workbook(template)

    # アクティブなワークシートを得る --- (*3)
    ws = wb["品質検討会資料"]
    ws2 = wb["form"]
    ws3 = wb["pos"]

    # データ書き込み --- (*4)
    ws["F1"] = "XXXX年    X月度    " + brand + "    XX工程"
    ws["A2"] = "（ " + str(run_no1) + " ）前    回    ラ    ン    概    況 [ " + operation_date_start + " ～ " + operation_date_end + " ]"

    for i in range(5, aggregate_list.__len__()+5, 1):
        ws['A' + str(i)] = aggregate_list[i-5][0]
        ws['F' + str(i)] = aggregate_list[i-5][1]
        ws['H' + str(i)] = aggregate_list[i-5][2]
        ws['J' + str(i)] = aggregate_list[i-5][3]
        # for j in range(0, avg1s_list.__len__(), 1):
        #     if aggregate_list[i-5][0] == avg1s_list[j][0]:
        #         ws['J' + str(i)] = avg1s_list[j][1]
        for j in range(0, avg2s_list.__len__(), 1):
            if aggregate_list[i-5][0] == avg2s_list[j][0]:
                ws['L' + str(i)] = avg2s_list[j][1]
        for j in range(0, avg3s_list.__len__(), 1):
            if aggregate_list[i-5][0] == avg3s_list[j][0]:
                ws['N' + str(i)] = avg3s_list[j][1]
        # 自動集計部最大項目数以上は取得不可 テンプレートに収まりきらないので切り捨てる
        if i == 64:
            break

    print('データ抽出完了')

    # パターンの取得 ["pos"]から銘柄比較してパターン取得
    pattern = ''
    for row in ws3.iter_rows(min_row=51, max_row=451, min_col=2, max_col=2):
        for data in row:
            if data.value == brand:
                pattern = ws3.cell(row=data.row, column=data.column + 1).value

    # パターン毎に処理分岐
    if pattern == 'A':
        for i in range(1, 8):
            for j in range(1, 18):
                copy = ws2.cell(row=i+23, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'B':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+32, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'C':
        for i in range(1, 12):
            for j in range(1, 18):
                copy = ws2.cell(row=i+41, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'D':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+53, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'E':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+62, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'F':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+66, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'G':
        for i in range(1, 11):
            for j in range(1, 18):
                copy = ws2.cell(row=i+70, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'H':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+81, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'I':
        for i in range(1, 14):
            for j in range(1, 18):
                copy = ws2.cell(row=i+85, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'J':
        for i in range(1, 7):
            for j in range(1, 18):
                copy = ws2.cell(row=i+99, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'K':
        for i in range(1, 8):
            for j in range(1, 18):
                copy = ws2.cell(row=i+106, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'L':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+114, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'M':
        for i in range(1, 5):
            for j in range(1, 18):
                copy = ws2.cell(row=i+123, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'N':
        for i in range(1, 5):
            for j in range(1, 18):
                copy = ws2.cell(row=i+128, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)

    # header
    ws.oddHeader.right.text = today[0:4] + "年" + today[4:6] + "月" + today[6:8] + "日\n"

    # 署名
    img = openpyxl.drawing.image.Image("\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\テンプレート\\署名.png")
    ws.add_image(img, 'F92')

    # ファイルを保存 --- (*5)
    file_path = '\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\品質検討会資料\\' + today + '_品質検討会資料(S法).xlsx'
    file_name = today + '_品質検討会資料(S法).xlsx'
    wb.save(file_path)

    with open(file_path, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(file_name)[0] or 'application/octet-stream')
    return response

