import datetime
import openpyxl
import subprocess
import pathlib
import os
import datetime
import traceback

from django.db import connections
from django.contrib.auth.decorators import login_required
# django関係のreturn関係のmoduleをインポート
from django.http import JsonResponse
from django.http import HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.utils.timezone import make_aware
from django.db.models import Q

from common.common_def import date_to_many_type
from config.settings.settings_common import BASE_DIR
from openpyxl.styles.borders import Border, Side
import mimetypes

from fms.views.common_def_views import get_template_file_path, get_attachment_folder_name

from fms.models import Budget, BudgetClassMaster, Log
from fms.models import User
from fms.models import AttachmentDocuments
from fms.models import ProBudgetUnit, ProSpecificationUnit
from fms.models import DepartmentMaster
from fms.models import SupplierMaster
from fms.models import ErpConstruction
from fms.models import StopWorkCause
from fms.models import ProcessMaster
from socket import gethostname
from fms.views.common_def_views import output_log_info, output_log_error, output_log_exception


# 台帳問い合わせexcel作成
@login_required
@require_POST
def execution_output_application(request):
    try:
        DIFF_JST_FROM_UTC = 9
        # JST = timezone(timedelta(hours=+9), 'JST')

        # now = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)
        now_naive = datetime.datetime.now()
        now = make_aware(now_naive)

        date_str = date_to_many_type(now)
        today_str = date_str.str_type_date_jp

        start_datetime = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)

        operator = request.user.username

        budget_id = int(request.POST['budget_id'])
        construction_id = int(request.POST['construction_id'])
        target = request.POST["target"]
        this_step = int(request.POST["this_step"])

        fold = 'proindividualcontractdoc'

        if target == "" or target is None:
            print('targetが存在しないため保存先がありません')
            data = {
                'msg': 'targetが存在しないため保存先がありません',
                'div_id_name': fold,
            }
            return JsonResponse(data)

        # テンプレートファイルのパス設定
        execution_output_application_templates_file_full_path = BASE_DIR + '\\static\\files\\fms\\execution_output_application\\【テンプレート】実行申請書.xlsx'

        # テンプレートファイル読み込み
        execution_output_application_file = openpyxl.load_workbook(execution_output_application_templates_file_full_path)

        # テンプレートファイルの各シートを読み込み
        # 【注意】テンプレートファイルのシートの名前や順番を変更すると正しく動作しなくなります
        sheet_name_list = execution_output_application_file.sheetnames
        # 表紙
        front_cover_sheet = execution_output_application_file[sheet_name_list[0]]
        # 実行申請累積表
        execution_application_sheet = execution_output_application_file[sheet_name_list[1]]
        # 比較表
        comparison_table_sheet = execution_output_application_file[sheet_name_list[2]]

        # ベースディレクトリ
        if gethostname() == 'YWEBSERV1':  # 本番
            base_directory = r'\\Ydomnserv\common\部門間フォルダ\FacilityData\Production'

        else:
            base_directory = r'\\Ydomnserv\common\部門間フォルダ\FacilityData\test'

        target_folder = "execution_output_application"

        # 新規の実行申請書のファイル名とパスを指定
        # new_file_name = 'result_ordered_construction_excel.xlsx'
        # new_file_name_fullpath = BASE_DIR + '\\static\\files\\fms\\execution_output_application\\' + new_file_name
        if this_step == 212002001:
            new_file_name = '中止_実行申請書_'
        else:
            new_file_name = '実行申請書_'
        new_file_name += str(construction_id) + '.xlsx'
        folder_name = base_directory + "\\" + target + "\\" + str(budget_id) + "\\" + str(construction_id) + "\\" + target_folder + "\\"
        new_file_name_fullpath = folder_name + new_file_name

        # 新規の実行申請書を作成
        # execution_output_application_file.save(new_file_name_fullpath)

        with connections['fmsdb'].cursor() as cursor:
            sql = """ SELECT fms_budget.budget_id,  """
            sql += """ fms_budget.business_year_id,  """
            sql += """ fms_budget.budget_class_id,  """
            sql += """ fms_budget.facility_process_id,  """
            sql += """ fms_processmaster.process_name,  """
            sql += """ fms_budget.budget_price,  """
            sql += """ fms_budget.purpose,  """
            sql += """ fms_budget.detail,  """
            sql += """ fms_budget.delivery_date,  """
            sql += """ fms_budget.start_date,  """
            sql += """ fms_budget.end_date,  """
            sql += """ fms_budget.budget_no,  """
            sql += """ fms_probudgetunit.budget_id,  """
            sql += """ fms_probudgetunit.department,  """
            sql += """ fms_probudgetunit.budget_name,  """
            sql += """ fms_probudgetunit.original_sec_person_in_charge,  """
            sql += """ fms_proindividualcontractdoc.construction_classification,  """
            sql += """ fms_prospecificationunit.construction_id,  """
            sql += """ fms_prospecificationunit.work_name,  """
            sql += """ fms_prospecificationunit.budget_id,  """
            sql += """ fms_prospecificationunit.specification_person_in_charge,  """
            sql += """ fms_prospecificationunit.procurement_person_in_charge,  """
            sql += """ fms_erpconstruction.total_price,  """
            sql += """ fms_erpconstruction.discount_price,  """
            sql += """ fms_erpconstruction.order_date,  """
            sql += """ fms_erpconstruction.purchase_order_no,  """
            sql += """ fms_proestimates.candidate_vendor1,  """
            sql += """ fms_proestimates.last_estimated_amount1,  """
            sql += """ fms_proestimates.estimated_amount_af_nego1,  """
            sql += """ fms_proestimates.eva_delivery_date1,  """
            sql += """ fms_proestimates.eva_estimated_valuation_amount1,  """
            sql += """ fms_proestimates.eva_final_price1,  """
            sql += """ fms_proestimates.eva_other1,  """
            sql += """ fms_proestimates.candidate_vendor2,  """
            sql += """ fms_proestimates.last_estimated_amount2,  """
            sql += """ fms_proestimates.estimated_amount_af_nego2,  """
            sql += """ fms_proestimates.eva_delivery_date2,  """
            sql += """ fms_proestimates.eva_estimated_valuation_amount2,  """
            sql += """ fms_proestimates.eva_final_price2,  """
            sql += """ fms_proestimates.eva_other2,  """
            sql += """ fms_proestimates.candidate_vendor3,  """
            sql += """ fms_proestimates.last_estimated_amount3,  """
            sql += """ fms_proestimates.estimated_amount_af_nego3,  """
            sql += """ fms_proestimates.eva_delivery_date3,  """
            sql += """ fms_proestimates.eva_estimated_valuation_amount3,  """
            sql += """ fms_proestimates.eva_final_price3,  """
            sql += """ fms_proestimates.eva_other3,  """
            sql += """ fms_proestimates.candidate_vendor4,  """
            sql += """ fms_proestimates.last_estimated_amount4,  """
            sql += """ fms_proestimates.estimated_amount_af_nego4,  """
            sql += """ fms_proestimates.eva_delivery_date4,  """
            sql += """ fms_proestimates.eva_estimated_valuation_amount4,  """
            sql += """ fms_proestimates.eva_final_price4,  """
            sql += """ fms_proestimates.eva_other4,  """
            sql += """ fms_proestimates.candidate_vendor5,  """
            sql += """ fms_proestimates.last_estimated_amount5,  """
            sql += """ fms_proestimates.estimated_amount_af_nego5,  """
            sql += """ fms_proestimates.eva_delivery_date5,  """
            sql += """ fms_proestimates.eva_estimated_valuation_amount5,  """
            sql += """ fms_proestimates.eva_final_price5,  """
            sql += """ fms_proestimates.eva_other5,  """
            sql += """ fms_proestimates.fixed_delivery_date,  """    # 決定納期
            sql += """ fms_proestimates.fixed_delivery_date_from,  """    # 決定工期_FROM
            sql += """ fms_proestimates.fixed_delivery_date_to,  """    # 決定工期_TO
            sql += """ fms_proestimates.fixed_delivery_location,  """    # 決定納入場所
            sql += """ fms_proestimates.confirmed_vendor,  """    # 確定業者
            sql += """ fms_proestimates.confirmed_last_estimated_amount,  """    # 確定本見積額
            sql += """ fms_proestimates.confirmed_estimated_amount_af_nego,  """    # 確定交渉後見積額
            sql += """ fms_proestimates.estimated_deadline_date,  """    # 決定見積提出期日
            sql += """ fms_provendorevaluation.estimate_assessment_evaluation_sum,  """
            sql += """ fms_provendorevaluation.eva_final_price  """
            sql += """ from (((((((fms_budget  """
            sql += """ LEFT JOIN fms_processmaster ON fms_budget.facility_process_id=fms_processmaster.process_cd2)  """
            sql += """ LEFT JOIN fms_probudgetunit ON fms_budget.budget_id=fms_probudgetunit.budget_id)  """
            sql += """ LEFT JOIN fms_prospecificationunit ON fms_budget.budget_id=fms_prospecificationunit.budget_id)  """
            sql += """ LEFT JOIN fms_proindividualcontractdoc"""
            sql += """        ON fms_budget.budget_id=fms_proindividualcontractdoc.budget_id"""
            sql += """       AND fms_prospecificationunit.construction_id=fms_proindividualcontractdoc.construction_id)  """
            sql += """ LEFT JOIN fms_erpconstruction"""
            sql += """        ON fms_budget.budget_id=fms_erpconstruction.budget_id"""
            sql += """       AND fms_prospecificationunit.construction_id=fms_erpconstruction.order_id)  """
            sql += """ LEFT JOIN fms_proestimates"""
            sql += """        ON fms_budget.budget_id=fms_proestimates.budget_id"""
            sql += """       AND fms_prospecificationunit.construction_id=fms_proestimates.construction_id)  """
            sql += """ LEFT JOIN fms_provendorevaluation"""
            sql += """        ON fms_budget.budget_id=fms_provendorevaluation.budget_id"""
            sql += """       AND fms_prospecificationunit.construction_id=fms_provendorevaluation.construction_id)  """
            sql += """ where fms_budget.lost_flag=0   """
            sql += """   AND fms_processmaster.lost_flag=0   """
            sql += """   AND fms_probudgetunit.lost_flag=0   """
            sql += """   AND fms_proindividualcontractdoc.lost_flag=0  """
            sql += """   AND fms_prospecificationunit.lost_flag=0  """
            sql += """   AND fms_proestimates.lost_flag=0  """
            sql += """   AND fms_provendorevaluation.lost_flag=0  """

            if budget_id is not "NON":
                sql = sql + """ AND fms_budget.budget_id=""" + str(budget_id)
            if construction_id is not "NON":
                sql = sql + """ AND fms_prospecificationunit.construction_id=""" + str(construction_id)

            res = cursor.execute(sql)
            construction_data = res.fetchall()

        # 表紙シート 取得レコードは1件のはず インデックス付与のためのfor
        for construction_data in construction_data:
            # 作成日
            front_cover_sheet.cell(row=1, column=26).value = datetime.date.strftime(now, '%Y/%#m/%#d')
            # 年度
            front_cover_sheet.cell(row=3, column=6).value = construction_data.business_year_id
            # 工事区分
            budget_class_num = BudgetClassMaster.objects.filter(budget_class_cd=construction_data.budget_class_id,
                                                                lost_flag=0).count()
            if budget_class_num == 1:
                budget_class = BudgetClassMaster.objects.get(budget_class_cd=construction_data.budget_class_id,
                                                             lost_flag=0
                                                             ).budget_class_name
            elif budget_class_num > 1:
                budget_class = "エラー！工事区分cd重複"
            else:
                budget_class = "エラー！工事区分cd該当なし"
            front_cover_sheet.cell(row=3, column=13).value = budget_class
            # 予算名
            front_cover_sheet.cell(row=16, column=6).value = construction_data.budget_name
            # 予算NO
            front_cover_sheet.cell(row=15, column=6).value = construction_data.budget_no
            # 部署名
            if DepartmentMaster.objects.filter(department_cd=construction_data.department, lost_flag=0).count() > 0:
                department_name = DepartmentMaster.objects.filter(department_cd=construction_data.department,
                                                                  lost_flag=0).order_by('display_order')[0].department_name
                front_cover_sheet.cell(row=14, column=6).value = department_name
            else:
                front_cover_sheet.cell(row=14, column=6).value = ""
            # 設備工程
            front_cover_sheet.cell(row=14, column=17).value = construction_data.facility_process_id
            front_cover_sheet.cell(row=14, column=21).value = construction_data.process_name
            # 予算額
            front_cover_sheet.cell(row=15, column=17).value = construction_data.budget_price
            # 理由/目的
            front_cover_sheet.cell(row=18, column=2).value = construction_data.purpose
            # 内容
            front_cover_sheet.cell(row=18, column=14).value = construction_data.detail
            # 納期 date型から年/月/日(ゼロサプレスなし)の文字列に変換 (#削除でゼロサプレスあり)
            if construction_data.delivery_date is not None and construction_data.delivery_date != "":
                converted_delivery_date = datetime.date.strftime(construction_data.delivery_date, '%Y/%#m/%#d')
            else:
                converted_delivery_date = ""
            front_cover_sheet.cell(row=29, column=4).value = converted_delivery_date
            # 工期Start date型から年/月/日(ゼロサプレスなし)の文字列に変換
            if construction_data.start_date is not None and construction_data.start_date != "":
                converted_start_date = datetime.date.strftime(construction_data.start_date, '%Y/%#m/%#d')
            else:
                converted_start_date = ""
            front_cover_sheet.cell(row=29, column=16).value = converted_start_date
            # 工期End date型から年/月/日(ゼロサプレスなし)の文字列に変換
            if construction_data.end_date is not None and construction_data.end_date != "":
                converted_end_date = datetime.date.strftime(construction_data.end_date, '%Y/%#m/%#d')
            else:
                converted_end_date = ""
            front_cover_sheet.cell(row=29, column=21).value = converted_end_date
            # 購買発注No
            front_cover_sheet.cell(row=32, column=4).value = construction_data.purchase_order_no
            # 工事ID
            front_cover_sheet.cell(row=32, column=13).value = construction_data.construction_id
            # 実行額
            if construction_data.total_price is not None and construction_data.discount_price is not None:
                front_cover_sheet.cell(row=32, column=20).value = construction_data.total_price - construction_data.discount_price
            else:
                front_cover_sheet.cell(row=32, column=20).value = construction_data.confirmed_estimated_amount_af_nego
            # 工事名
            front_cover_sheet.cell(row=33, column=4).value = construction_data.work_name
            # 決定納期 date型から年/月/日(ゼロサプレスなし)の文字列に変換
            if construction_data.fixed_delivery_date is not None and construction_data.fixed_delivery_date != "":
                converted_fixed_delivery_date = datetime.date.strftime(construction_data.fixed_delivery_date, '%Y/%#m/%#d')
            else:
                converted_fixed_delivery_date = ""
            front_cover_sheet.cell(row=34, column=5).value = converted_fixed_delivery_date
            # 決定工期_FROM date型から年/月/日(ゼロサプレスなし)の文字列に変換
            if construction_data.fixed_delivery_date_from is not None and construction_data.fixed_delivery_date_from != "":
                converted_fixed_delivery_date_from = datetime.date.strftime(construction_data.fixed_delivery_date_from, '%Y/%#m/%#d')
            else:
                converted_fixed_delivery_date_from = ""
            front_cover_sheet.cell(row=34, column=11).value = converted_fixed_delivery_date_from
            # 決定工期_TO date型から年/月/日(ゼロサプレスなし)の文字列に変換
            if construction_data.fixed_delivery_date_to is not None and construction_data.fixed_delivery_date_to != "":
                converted_fixed_delivery_date_to = datetime.date.strftime(construction_data.fixed_delivery_date_to, '%Y/%#m/%#d')
            else:
                converted_fixed_delivery_date_to = ""
            front_cover_sheet.cell(row=34, column=19).value = converted_fixed_delivery_date_to

        # 表紙の内容を保存
        # execution_output_application_file.save(new_file_name_fullpath)
        # 自身の個別契約書作成日時を取得
        # if target == 'stop_work':
        #     my_log_data = Log.objects.filter(target='stop_work', action='permit', step=212001021,
        #                                      target_id=construction_id).order_by('-operation_datetime')[0]
        # else:
        #     my_log_data = Log.objects.filter(target='prospecificationunit', action='entry', step=241003001,
        #                                      target_id=construction_id).order_by('-operation_datetime')[0]
        # operation_datetime = my_log_data.operation_datetime

        erp_construction_data = ErpConstruction.objects.filter(order_id=construction_id,
                                                               budget_id=budget_id).order_by('-id')[0]
        if erp_construction_data.update_datetime is None or erp_construction_data.update_datetime == "":
            operation_datetime = erp_construction_data.entry_datetime
        else:
            operation_datetime = erp_construction_data.update_datetime
        # 実行申請累積表

        # if this_step == 212002001:
        #     cancel_data_get_log_target_conditions = "OR target='stop_work'"
        #     cancel_data_get_log_action_step_conditions = "OR(action='permit' AND step=212001021) "
        #     cancel_data_get_where_fmslog_conditions = "OR(fmslog.step=212001021) "
        # else:
        #     cancel_data_get_log_target_conditions = ""
        #     cancel_data_get_log_action_step_conditions = ""
        #     cancel_data_get_where_fmslog_conditions = ""

        with connections['fmsdb'].cursor() as cursor:
            sql = """ SELECT fms_budget.budget_id,  """
            sql += """ fms_budget.business_year_id,  """
            sql += """ fms_budget.budget_price,  """
            sql += """ fms_budget.budget_no,  """
            sql += """ fms_probudgetunit.budget_id,  """
            sql += """ fms_probudgetunit.budget_name,  """
            sql += """ fms_prospecificationunit.construction_id,  """
            sql += """ fms_prospecificationunit.work_name,  """
            sql += """ fms_prospecificationunit.budget_id,  """

            # sql += """ fms_proestimates.confirmed_last_estimated_amount AS total_price, """
            # sql += """ fms_proestimates.confirmed_last_estimated_amount - fms_proestimates.confirmed_estimated_amount_af_nego AS discount_price, """

            sql += """ fms_erpconstruction.total_price,  """
            sql += """ fms_erpconstruction.discount_price,  """
            sql += """ fms_erpconstruction.order_date,  """
            sql += """ fms_erpconstruction.entry_datetime,  """
            sql += """ fms_erpconstruction.entry_operator,  """
            sql += """ fms_erpconstruction.update_datetime,  """
            sql += """ fms_erpconstruction.update_operator, """
            sql += """ fms_erpconstruction.numbering, """
            sql += """ fms_erpconstruction.year, """
            sql += """ fms_erpconstruction.detail_no, """
            sql += """ fms_proestimates.fixed_delivery_date,  """    # 決定納期
            sql += """ fms_proestimates.fixed_delivery_date_from,  """    # 決定工期_FROM
            sql += """ fms_proestimates.fixed_delivery_date_to,  """    # 決定工期_TO
            sql += """ fms_proestimates.confirmed_vendor  """
            sql += """ from ((((fms_budget  """
            sql += """ LEFT JOIN fms_probudgetunit ON fms_budget.budget_id=fms_probudgetunit.budget_id)  """
            sql += """ LEFT JOIN fms_prospecificationunit ON fms_budget.budget_id=fms_prospecificationunit.budget_id)  """
            sql += """ LEFT JOIN fms_erpconstruction"""
            sql += """        ON fms_budget.budget_id=fms_erpconstruction.budget_id"""
            sql += """       AND fms_prospecificationunit.construction_id=fms_erpconstruction.order_id)  """
            sql += """ LEFT JOIN fms_proestimates"""
            sql += """        ON fms_budget.budget_id=fms_proestimates.budget_id"""
            sql += """       AND fms_prospecificationunit.construction_id=fms_proestimates.construction_id)  """
            sql += """ where fms_budget.lost_flag=0   """
            sql += """ AND fms_probudgetunit.lost_flag=0   """
            sql += """ AND fms_prospecificationunit.lost_flag=0  """
            sql += """ AND fms_proestimates.lost_flag=0  """
            sql += """ AND (fms_erpconstruction.status = 3 OR fms_erpconstruction.status = 10) """
            sql += """ AND case when fms_erpconstruction.update_datetime IS NULL 
                                then fms_erpconstruction.entry_datetime 
                                else fms_erpconstruction.update_datetime end <= '""" + str(operation_datetime) + """'"""

            if budget_id is not "NON":
                sql = sql + """ AND fms_budget.budget_id=""" + str(budget_id)

            sql += """ order by case when fms_erpconstruction.update_datetime IS NULL 
                                     then fms_erpconstruction.entry_datetime 
                                     else fms_erpconstruction.update_datetime end """

            res = cursor.execute(sql)
            construction_data_list = res.fetchall()

        # 実行申請累積表_総額
        with connections['fmsdb'].cursor() as cursor:
            # sql = """ SELECT SUM(confirmed_last_estimated_amount) AS sum_total_price_value, """
            # sql += """ SUM(confirmed_last_estimated_amount - confirmed_estimated_amount_af_nego) AS sum_discount_price_value """
            # sql += """ FROM fms_proestimates """
            # sql += """ WHERE budget_id = """ + str(budget_id)
            # sql += """ AND lost_flag = 0"""
            sql = """ SELECT SUM(total_price) AS sum_total_price_value, """
            sql += """ SUM(discount_price) AS sum_discount_price_value """
            sql += """ FROM fms_erpconstruction """
            sql += """ WHERE fms_erpconstruction.budget_id = """ + str(budget_id)
            sql += """ AND (status = 3 OR status = 10)"""
            sql += """ AND case when fms_erpconstruction.update_datetime IS NULL 
                                then fms_erpconstruction.entry_datetime 
                                else fms_erpconstruction.update_datetime end <= '""" + str(operation_datetime) + """'"""
            res = cursor.execute(sql)
            execution_application_total_value = res.fetchall()

        # 追加するシート数を算出 1シートにつき20件まで
        create_sheet_num = 1 + ((len(construction_data_list) - 1) // 20)

        # 最後尾に実行申請累積表シートを追加し、比較表の前に移動
        for loop in range(1, create_sheet_num):
            copy_sheet = execution_output_application_file.copy_worksheet(execution_application_sheet)
            copy_sheet.title = "実行申請累積表_" + str(loop + 1)
            execution_output_application_file.move_sheet(copy_sheet, offset=-1)

        execution_application_sheet.title = "実行申請累積表_1"

        # 2シート目以降の開始ループ番号
        loop_keeper = 0
        item_no = 1

        for loop in range(create_sheet_num):

            # 作成した実行申請累積表シートを書き込み先として指定
            sheet_name = "実行申請累積表_" + str(loop + 1)
            target_sheet = execution_output_application_file[sheet_name]

            # ページ数
            target_sheet.cell(row=1, column=25).value = str(loop + 1) + "/" + str(create_sheet_num) + "ページ"
            target_row = 9

            # 各ページに総額を記載
            for execution_application_total_value_item in execution_application_total_value:
                # 見積額_総額
                sum_total_price = execution_application_total_value_item.sum_total_price_value
                target_sheet.cell(row=30, column=13).value = sum_total_price
                # ネゴ額_総額
                sum_discount_price = execution_application_total_value_item.sum_discount_price_value
                target_sheet.cell(row=30, column=19).value = sum_discount_price
                # 実行額_総額
                target_sheet.cell(row=30, column=16).value = sum_total_price - sum_discount_price

            for construction_data_item in construction_data_list[loop_keeper:]:
                # 予算NO
                target_sheet.cell(row=4, column=6).value = construction_data_item.budget_no
                # 予算額
                target_sheet.cell(row=4, column=18).value = construction_data_item.budget_price
                # 予算名
                target_sheet.cell(row=5, column=6).value = construction_data_item.budget_name
                # 項番
                target_sheet.cell(row=target_row, column=2).value = construction_data_item.numbering
                item_no += 1
                # 工事名
                target_sheet.cell(row=target_row, column=3).value = construction_data_item.work_name
                # 工事ID
                target_sheet.cell(row=target_row, column=10).value = construction_data_item.construction_id
                # 見積額
                total_price = construction_data_item.total_price
                target_sheet.cell(row=target_row, column=13).value = total_price
                # ネゴ額
                discount_price = construction_data_item.discount_price
                target_sheet.cell(row=target_row, column=19).value = discount_price
                # 実行額
                target_sheet.cell(row=target_row, column=16).value = total_price - discount_price
                # 決定業者
                vendor_cd = construction_data_item.confirmed_vendor
                if vendor_cd is not None and vendor_cd != "":
                    vendor_data_count = SupplierMaster.objects.filter(supplier_cd=vendor_cd).count()
                    if vendor_data_count > 0:
                        vendor_name = SupplierMaster.objects.get(supplier_cd=vendor_cd).supplier_name
                    else:
                        vendor_name = vendor_cd
                else:
                    vendor_name = ""
                target_sheet.cell(row=target_row, column=22).value = vendor_name
                # 決定納期
                if total_price < 0:
                    input_data = "中止"
                else:
                    fixed_delivery_date = construction_data_item.fixed_delivery_date
                    if fixed_delivery_date is not None and fixed_delivery_date != "":
                        # 決定納期
                        input_data = fixed_delivery_date
                    elif construction_data_item.fixed_delivery_date_from is not None and construction_data_item.fixed_delivery_date_from != "":
                        # 決定納期
                        input_data = construction_data_item.fixed_delivery_date_from
                    else:
                        input_data = construction_data_item.fixed_delivery_date_to

                target_sheet.cell(row=target_row, column=26).value = input_data

                target_row += 1

                if target_row > 28:
                    loop_keeper = (loop + 1) * 20
                    break

            # 比較表シートの内容を保存
            # execution_output_application_file.save(new_file_name_fullpath)

        # 比較表
        with connections['fmsdb'].cursor() as cursor:
            sql = """ SELECT fms_budget.budget_id,  """
            sql += """ fms_budget.business_year_id,  """
            sql += """ fms_budget.budget_no,  """
            sql += """ fms_probudgetunit.budget_id,  """
            sql += """ fms_probudgetunit.department,  """
            sql += """ fms_probudgetunit.budget_name,  """
            sql += """ fms_probudgetunit.original_sec_person_in_charge,  """
            sql += """ fms_proindividualcontractdoc.construction_classification,  """
            sql += """ fms_prospecificationunit.construction_id,  """
            sql += """ fms_prospecificationunit.work_name,  """
            sql += """ fms_prospecificationunit.budget_id,  """
            sql += """ fms_prospecificationunit.specification_person_in_charge,  """
            sql += """ fms_prospecificationunit.procurement_person_in_charge,  """
            sql += """ fms_prospecificationunit.construction_id,  """
            sql += """ fms_prospecificationunit.cancel_flag,  """
            sql += """ fms_proestimates.fixed_delivery_date,  """    # 決定納期
            sql += """ fms_proestimates.fixed_delivery_date_from,  """    # 決定工期_FROM
            sql += """ fms_proestimates.fixed_delivery_date_to,  """    # 決定工期_TO
            sql += """ fms_proestimates.fixed_delivery_location,  """    # 決定納入場所
            sql += """ fms_proestimates.candidate_vendor1,  """
            sql += """ fms_proestimates.last_estimated_amount1,  """
            sql += """ fms_proestimates.estimated_amount_af_nego1,  """
            sql += """ fms_proestimates.eva_delivery_date1,  """
            sql += """ fms_proestimates.eva_estimated_valuation_amount1,  """
            sql += """ fms_proestimates.eva_final_price1,  """
            sql += """ fms_proestimates.eva_other1,  """
            sql += """ fms_proestimates.candidate_vendor2,  """
            sql += """ fms_proestimates.last_estimated_amount2,  """
            sql += """ fms_proestimates.estimated_amount_af_nego2,  """
            sql += """ fms_proestimates.eva_delivery_date2,  """
            sql += """ fms_proestimates.eva_estimated_valuation_amount2,  """
            sql += """ fms_proestimates.eva_final_price2,  """
            sql += """ fms_proestimates.eva_other2,  """
            sql += """ fms_proestimates.candidate_vendor3,  """
            sql += """ fms_proestimates.last_estimated_amount3,  """
            sql += """ fms_proestimates.estimated_amount_af_nego3,  """
            sql += """ fms_proestimates.eva_delivery_date3,  """
            sql += """ fms_proestimates.eva_estimated_valuation_amount3,  """
            sql += """ fms_proestimates.eva_final_price3,  """
            sql += """ fms_proestimates.eva_other3,  """
            sql += """ fms_proestimates.candidate_vendor4,  """
            sql += """ fms_proestimates.last_estimated_amount4,  """
            sql += """ fms_proestimates.estimated_amount_af_nego4,  """
            sql += """ fms_proestimates.eva_delivery_date4,  """
            sql += """ fms_proestimates.eva_estimated_valuation_amount4,  """
            sql += """ fms_proestimates.eva_final_price4,  """
            sql += """ fms_proestimates.eva_other4,  """
            sql += """ fms_proestimates.candidate_vendor5,  """
            sql += """ fms_proestimates.last_estimated_amount5,  """
            sql += """ fms_proestimates.estimated_amount_af_nego5,  """
            sql += """ fms_proestimates.eva_delivery_date5,  """
            sql += """ fms_proestimates.eva_estimated_valuation_amount5,  """
            sql += """ fms_proestimates.eva_final_price5,  """
            sql += """ fms_proestimates.eva_other5,  """
            sql += """ fms_proestimates.confirmed_vendor,  """    # 確定業者
            sql += """ fms_provendorevaluation.estimate_assessment_evaluation_sum, """
            sql += """ fms_provendorevaluation.eva_final_price  """
            sql += """ from (((((fms_budget  """
            sql += """ LEFT JOIN fms_probudgetunit ON fms_budget.budget_id=fms_probudgetunit.budget_id)  """
            sql += """ LEFT JOIN fms_prospecificationunit ON fms_budget.budget_id=fms_prospecificationunit.budget_id)  """
            sql += """ LEFT JOIN fms_proindividualcontractdoc"""
            sql += """        ON fms_budget.budget_id=fms_proindividualcontractdoc.budget_id"""
            sql += """       AND fms_prospecificationunit.construction_id=fms_proindividualcontractdoc.construction_id)  """
            sql += """ LEFT JOIN fms_proestimates"""
            sql += """        ON fms_budget.budget_id=fms_proestimates.budget_id"""
            sql += """       AND fms_prospecificationunit.construction_id=fms_proestimates.construction_id)  """
            sql += """ LEFT JOIN fms_provendorevaluation """
            sql += """        ON fms_budget.budget_id=fms_provendorevaluation.budget_id """
            sql += """       AND fms_prospecificationunit.construction_id=fms_provendorevaluation.construction_id) """
            sql += """ where fms_budget.lost_flag=0   """
            sql += """ AND fms_probudgetunit.lost_flag=0   """
            sql += """ AND fms_prospecificationunit.lost_flag=0  """
            sql += """ AND fms_proestimates.lost_flag=0  """
            sql += """ AND fms_proindividualcontractdoc.lost_flag=0  """
            sql += """ AND fms_provendorevaluation.lost_flag=0  """

            if budget_id is not "NON":
                sql = sql + """ AND fms_budget.budget_id=""" + str(budget_id)
            if construction_id is not "NON":
                sql = sql + """ AND fms_prospecificationunit.construction_id=""" + str(construction_id)

            res = cursor.execute(sql)
            proestimates_data = res.fetchall()

        # 比較表シート 取得レコードは1件のはず インデックス付与のためのfor
        for proestimates_data in proestimates_data:
            estimate_num = 0
            # 入力されている候補業者数を取得
            for loop in range(5):
                # 業者名
                confirmed_vendor = getattr(proestimates_data, "candidate_vendor" + str(loop + 1))
                if confirmed_vendor != "" and confirmed_vendor is not None:
                    estimate_num += 1

            # 候補業者数が4社以上なら必要分シートを作成する
            proestimates_sheet_num = 1 + ((estimate_num - 1) // 3)
            if proestimates_sheet_num > 1:
                for loop in range(1, proestimates_sheet_num):
                    copy_sheet = execution_output_application_file.copy_worksheet(comparison_table_sheet)
                    copy_sheet.title = "比較表_" + str(loop + 1)

            comparison_table_sheet.title = "比較表_1"

            # 確定業者記入欄に飛ぶ前のシート名退避変数
            return_sheet_name = "比較表_1"
            # 確定業者記入済みフラグ
            # fix_vendor_wrote_flag = 0
            # 2シート目以降の開始ループ番号
            loop_keeper = 0

            # 作成する比較表シート数分繰り返し
            for loop in range(proestimates_sheet_num):

                # 作成した比較表シートを書き込み先として指定
                sheet_name = "比較表_" + str(loop + 1)
                target_sheet = execution_output_application_file[sheet_name]

                # ページ数
                target_sheet.cell(row=2, column=22).value = str(loop + 1) + "/" + str(proestimates_sheet_num) + "ページ"
                # 工事分類
                target_sheet.cell(row=4, column=8).value = proestimates_data.construction_classification
                # 予算NO
                target_sheet.cell(row=4, column=22).value = proestimates_data.budget_no
                # 予算名
                target_sheet.cell(row=6, column=8).value = proestimates_data.budget_name
                # 予算ID
                target_sheet.cell(row=6, column=22).value = proestimates_data.budget_id
                # 原課
                if DepartmentMaster.objects.filter(department_cd=proestimates_data.department, lost_flag=0).count() > 0:
                    department_name = DepartmentMaster.objects.filter(department_cd=proestimates_data.department,
                                                                      lost_flag=0
                                                                      ).order_by('display_order')[0].department_name
                else:
                    department_name = ""
                target_sheet.cell(row=7, column=8).value = department_name
                # 原課担当者
                if User.objects.filter(username=proestimates_data.original_sec_person_in_charge, lost_flag=0).count() > 0:
                    user_data = User.objects.filter(username=proestimates_data.original_sec_person_in_charge,
                                                    lost_flag=0
                                                    ).order_by('display_order')[0]
                    user_name = user_data.last_name + user_data.first_name
                else:
                    user_name = ""
                target_sheet.cell(row=7, column=14).value = user_name
                # 工事ID
                target_sheet.cell(row=7, column=22).value = proestimates_data.construction_id
                # 工事名
                target_sheet.cell(row=8, column=8).value = proestimates_data.work_name
                # 仕様書担当者
                if User.objects.filter(username=proestimates_data.specification_person_in_charge, lost_flag=0).count() > 0:
                    user_data = User.objects.filter(username=proestimates_data.specification_person_in_charge,
                                                    lost_flag=0
                                                    ).order_by('display_order')[0]
                    user_name = user_data.last_name + user_data.first_name
                else:
                    user_name = ""
                target_sheet.cell(row=9, column=8).value = user_name
                # 調達G担当者
                if User.objects.filter(username=proestimates_data.procurement_person_in_charge, lost_flag=0).count() > 0:
                    user_data = User.objects.filter(username=proestimates_data.procurement_person_in_charge,
                                                    lost_flag=0
                                                    ).order_by('display_order')[0]
                    user_name = user_data.last_name + user_data.first_name
                else:
                    user_name = ""
                target_sheet.cell(row=9, column=14).value = user_name
                # 決定納期 date型から年/月/日(ゼロサプレスなし)の文字列に変換
                if proestimates_data.fixed_delivery_date is not None and proestimates_data.fixed_delivery_date != "":
                    fixed_delivery_date = datetime.date.strftime(proestimates_data.fixed_delivery_date, '%Y/%#m/%#d')
                else:
                    fixed_delivery_date = ""
                # 決定工期_FROM date型から年/月/日(ゼロサプレスなし)の文字列に変換
                if proestimates_data.fixed_delivery_date_from is not None and proestimates_data.fixed_delivery_date_from != "":
                    fixed_delivery_date_from = datetime.date.strftime(proestimates_data.fixed_delivery_date_from, '%Y/%#m/%#d')
                else:
                    fixed_delivery_date_from = ""
                # 決定工期_TO date型から年/月/日(ゼロサプレスなし)の文字列に変換
                if proestimates_data.fixed_delivery_date_to is not None and proestimates_data.fixed_delivery_date_to != "":
                    fixed_delivery_date_to = datetime.date.strftime(proestimates_data.fixed_delivery_date_to, '%Y/%#m/%#d')
                else:
                    fixed_delivery_date_to = ""

                if fixed_delivery_date is not None and fixed_delivery_date != "":

                    target_sheet.cell(row=9, column=22).value = fixed_delivery_date
                elif fixed_delivery_date_from is not None and fixed_delivery_date_from != "":
                    target_sheet.cell(row=9, column=22).value = fixed_delivery_date_from
                else:
                    target_sheet.cell(row=9, column=22).value = fixed_delivery_date_to

                # 見積査定評価まとめ
                target_sheet.cell(row=19, column=8).value = proestimates_data.estimate_assessment_evaluation_sum

                # 最終金額評価
                target_sheet.cell(row=23, column=8).value = proestimates_data.eva_final_price

            confirmed_vendor = ""
            confirmed_vendor_estimated_amount_af_nego = ""

            for column in range(loop_keeper, 5):  # 0~4
                # target_column = 8 + (6 * ((column + 1 - fix_vendor_wrote_flag) % 3))
                vendor_no = column + 1

                # 業者名
                target_vendor = getattr(proestimates_data, "candidate_vendor" + str(vendor_no))
                # 確定業者と同じ名前であれば1シート目1列目の確定業者記入欄へ
                if target_vendor == proestimates_data.confirmed_vendor:
                    # 現在のシート名を退避
                    # return_sheet_name = sheet_name
                    confirmed_vendor = proestimates_data.confirmed_vendor
                    sheet_name = "比較表_1"
                    target_sheet = execution_output_application_file[sheet_name]
                    target_column = 8
                    fix_vendor_wrote_flag = 1
                    confirmed_vendor_estimated_amount_af_nego = getattr(proestimates_data, "estimated_amount_af_nego" + str(vendor_no))

                    ans = comparison_table_sheet_each_vendor_info_make(target_sheet, target_column, proestimates_data, vendor_no)

                    # 業者名
                    vendor_cd = getattr(proestimates_data, "candidate_vendor" + str(vendor_no))
                    if vendor_cd is not None and vendor_cd != "":
                        vendor_data_count = SupplierMaster.objects.filter(supplier_cd=vendor_cd).count()
                        if vendor_data_count > 0:
                            vendor_name = SupplierMaster.objects.get(supplier_cd=vendor_cd).supplier_name
                        else:
                            vendor_name = vendor_cd
                    else:
                        vendor_name = ""
                    target_sheet.cell(row=28, column=14).value = vendor_name

                    # 最終金額評価
                    if proestimates_data.cancel_flag == 1:
                        target_sheet.cell(row=29, column=14).value = 0 - confirmed_vendor_estimated_amount_af_nego
                    else:
                        target_sheet.cell(row=29, column=14).value = confirmed_vendor_estimated_amount_af_nego

                    result = ans.get('result')

                    if result == 'Done':
                        break

            if this_step == 212002001:
                break

            i = 0
            fix_vendor_wrote_flag = 1
            for column in range(loop_keeper, 5):  # 0~4
                # target_column = 8 + (6 * ((column + 1 - fix_vendor_wrote_flag) % 3))
                vendor_no = column + 1

                    # print(target_sheet)

                # 業者名
                target_vendor = getattr(proestimates_data, "candidate_vendor" + str(vendor_no))
                if target_vendor is not None:
                    # 確定業者以外の処理
                    if target_vendor != proestimates_data.confirmed_vendor:
                        if i < 2:
                            sheet_name = "比較表_1"
                            target_sheet = execution_output_application_file[sheet_name]
                            target_column = 8 + 6 * (i + 1)
                        elif proestimates_sheet_num > 1:
                            sheet_name = "比較表_2"
                            target_sheet = execution_output_application_file[sheet_name]
                            target_column = 8 + 6 * (i - 2)

                        ans = comparison_table_sheet_each_vendor_info_make(target_sheet, target_column, proestimates_data, vendor_no)

                        result = ans.get('result')

                        if result == 'Done':
                            i += 1

            # 全シート分に記載
            for sheet_loop in range(1, proestimates_sheet_num):
                sheet_name = "比較表_" + str(sheet_loop + 1)
                target_sheet = execution_output_application_file[sheet_name]

                # 業者名
                target_sheet.cell(row=28, column=14).value = confirmed_vendor

                # 最終金額評価
                target_sheet.cell(row=29, column=14).value = confirmed_vendor_estimated_amount_af_nego

            # 確定業者記入欄に飛ぶ前のシートに戻す
            # sheet_name = return_sheet_name
            # target_sheet = execution_output_application_file[sheet_name]

        # フォルダが存在しない場合、フォルダを作成する
        if not pathlib.Path(folder_name).exists():
            os.makedirs(folder_name)

        # 指定ファイルが存在する場合、削除
        if pathlib.Path(new_file_name_fullpath).exists():
            subprocess.run('del ' + new_file_name_fullpath, shell=True)

        # 実行申請累積表シートの内容を保存
        execution_output_application_file.save(new_file_name_fullpath)

        # ファイルアップロード機能
        attach_folder_name = "\\" + target + "\\" + str(budget_id) + "\\" + str(construction_id) + '\\'
        attachment_documents_data, created = AttachmentDocuments.objects.get_or_create(file_name=new_file_name,
                                                                                       folder=attach_folder_name)
        attachment_documents_data.folder = '\\' + target + '\\' + str(budget_id) + '\\' + str(construction_id) + '\\'
        attachment_documents_data.div_id_name = fold
        attachment_documents_data.data = target_folder
        attachment_documents_data.entry_time = now
        attachment_documents_data.entry_user = request.user.username
        attachment_documents_data.lost_flag = 0
        attachment_documents_data.save()

        msg = '出力完了\nファイル一覧から取り出してください'

        data = {
            'budget_id': budget_id,
            'construction_id': construction_id,
            'msg': msg,
            'div_id_name': fold,
        }

        return JsonResponse(data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise


def comparison_table_sheet_each_vendor_info_make(target_sheet, target_column, proestimates_data, vendor_no):

    vendor_cd = getattr(proestimates_data, "candidate_vendor" + str(vendor_no))
    if vendor_cd is not None and vendor_cd != "":
        vendor_data_count = SupplierMaster.objects.filter(supplier_cd=vendor_cd).count()
        if vendor_data_count > 0:
            vendor_name = SupplierMaster.objects.get(supplier_cd=vendor_cd).supplier_name
        else:
            vendor_name = vendor_cd
    else:
        vendor_name = ""
    target_sheet.cell(row=10, column=target_column).value = vendor_name

    # 辞退フラグ(空欄 or 0円は辞退として扱う)
    refuse_flag = False
    # 見積額
    last_estimated_amount = getattr(proestimates_data, "last_estimated_amount" + str(vendor_no))
    if proestimates_data.cancel_flag == 1:
        target_sheet.cell(row=11, column=target_column).value = 0 - last_estimated_amount
    elif last_estimated_amount is not None and last_estimated_amount != '' and last_estimated_amount != 0:
        target_sheet.cell(row=11, column=target_column).value = last_estimated_amount
    else:
        target_sheet.cell(row=11, column=target_column).value = '辞退'
        refuse_flag = True

    # 発行NO
    # ws_new_sheet.cell(row=i, column=2).value = today_str

    # 発注額
    estimated_amount_af_nego = getattr(proestimates_data, "estimated_amount_af_nego" + str(vendor_no))
    if proestimates_data.cancel_flag == 1:
        target_sheet.cell(row=14, column=target_column).value = 0 - estimated_amount_af_nego
    elif estimated_amount_af_nego is not None and estimated_amount_af_nego != '' and estimated_amount_af_nego != 0:
        target_sheet.cell(row=14, column=target_column).value = estimated_amount_af_nego
    else:
        target_sheet.cell(row=14, column=target_column).value = '辞退'
        refuse_flag = True

    # ネゴ額、ネゴ率（辞退の場合は表示しない）
    if not refuse_flag and (last_estimated_amount is not None or last_estimated_amount != "") and \
            (estimated_amount_af_nego is not None or estimated_amount_af_nego != ""):
        if proestimates_data.cancel_flag == 1:
            # ネゴ額
            target_sheet.cell(row=12, column=target_column).value = 0
            # ネゴ率
            target_sheet.cell(row=13, column=target_column).value = 0
        else:
            # ネゴ額
            target_sheet.cell(row=12, column=target_column
                              ).value = last_estimated_amount - estimated_amount_af_nego
            # ネゴ率
            target_sheet.cell(row=13, column=target_column
                              ).value = (last_estimated_amount - estimated_amount_af_nego) / last_estimated_amount * 100

    # 工期/納期
    target_sheet.cell(row=15, column=target_column
                      ).value = getattr(proestimates_data, "eva_delivery_date" + str(vendor_no))

    # 見積査定
    target_sheet.cell(row=16, column=target_column
                      ).value = getattr(proestimates_data, "eva_estimated_valuation_amount" + str(vendor_no))

    # 最終金額評価
    target_sheet.cell(row=17, column=target_column
                      ).value = getattr(proestimates_data, "eva_final_price" + str(vendor_no))

    # その他
    target_sheet.cell(row=18, column=target_column
                      ).value = getattr(proestimates_data, "eva_other" + str(vendor_no))
    result = 'Done'

    ans = {
        'result': result,
    }

    return ans


# 実行額、件数取得
def get_construction_data(budget_id):
    construction_count = 0
    construction_total_price = 0

    prospec_list = ProSpecificationUnit.objects.filter(budget_id=budget_id, lost_flag=0)

    for prospec_item in prospec_list:
        construction_list = ErpConstruction.objects.filter(budget_id=budget_id, order_id=prospec_item.construction_id)
        construction_list = construction_list.filter(Q(status=3) | Q(status=10))
        for construction_item in construction_list:
            # 工事中止した場合、金額がマイナスのErpConstructionレコードが追加されるので、件数からは除外
            if (construction_item.total_price - construction_item.discount_price) > 0:
                construction_count = construction_count + 1
            # 実行額は合計する
            construction_total_price += construction_item.total_price - construction_item.discount_price

    return construction_count, construction_total_price

# 完了報告書excel作成
@login_required
@require_POST
def make_complete_report_file(request):
    try:
        # 実行日時取得
        now_naive = datetime.datetime.now()
        now = make_aware(now_naive)
        d_today = datetime.date.today()

        # POST情報取得
        target = request.POST["target"]
        this_step = int(request.POST["this_step"])
        budget_id = int(request.POST["budget_id"])
        if 213000000 < this_step < 213004000:
            current_tab = request.POST["current_tab"]
        else:
            current_tab = target

        # 固定値
        templates_file_ext = '.xlsx'
        templates_file_sheet = 'Sheet1'
        target_folder = 'complete_report'

        relation_budget_no_str = ''

        # 完了/中止 両対応用変数
        if this_step == 213002021:
            normal_flag = True
            templates_file_name = '予算完了報告書'
        else:
            normal_flag = False
            templates_file_name = '予算中止報告書'
        templates_file_full_name = templates_file_name + templates_file_ext
        new_file_name = templates_file_name + '_' + str(budget_id) + templates_file_ext

        # 共有フォルダパス取得
        folder_name, attach_folder_name = get_attachment_folder_name(target, budget_id, 0, target_folder)
        templates_file_path = get_template_file_path(target_folder, templates_file_full_name)

        # テンプレートファイル読込
        wb_new_file = openpyxl.load_workbook(templates_file_path)

        # 出力対象レコード取得(関連予算含む、予算Noが無いものは出力対象外)
        budget_list = Budget.objects.filter(relation_budget_id=budget_id,
                                            lost_flag=0).exclude(budget_no=None).order_by('budget_id')

        if len(budget_list) < 1:
            msg = '出力対象の予算がありません'
            data = {
                'msg': msg,
                'div_id_name': current_tab,
            }
            return JsonResponse(data)

        for budget_item in budget_list:
            # 予算NOが無い場合対処(予算IDを使用する)
            if budget_item.budget_no is not None:
                budget_no_str = budget_item.budget_no
            else:
                budget_no_str = str(budget_item.budget_id)

            # 関連予算分のシートを用意
            if budget_item.budget_id != budget_id:
                if relation_budget_no_str != '':
                    relation_budget_no_str += '、'
                relation_budget_no_str += budget_no_str
                # シートをコピー
                ws_new_sheet = wb_new_file.copy_worksheet(wb_new_file[templates_file_sheet])
                ws_new_sheet.title = budget_no_str
            else:
                main_budget_no_str = budget_no_str

        # 各budgetデータ集計+シート出力処理
        for budget_item in budget_list:
            if budget_item.budget_no is not None:
                budget_no_str = budget_item.budget_no
            else:
                budget_no_str = str(budget_item.budget_id)

            if budget_item.budget_id == budget_id:
                # 主予算
                ws_new_sheet = wb_new_file[templates_file_sheet]
                ws_new_sheet.title = budget_no_str
            else:
                # 関連予算
                ws_new_sheet = wb_new_file[budget_no_str]

            # シートにデータを出力
            ws_new_sheet.cell(row=3, column=6).value = budget_item.business_year_id

            department_item = DepartmentMaster.objects.get(
                department_cd=budget_item.budget_main_department_id, lost_flag=0)
            ws_new_sheet.cell(row=14, column=6).value = department_item.department_name

            process_item = ProcessMaster.objects.get(
                process_cd2=budget_item.facility_process_id, lost_flag=0)
            ws_new_sheet.cell(row=14, column=17).value = \
                process_item.process_cd + '(' + process_item.process_name + ')'

            ws_new_sheet.cell(row=15, column=6).value = budget_item.budget_no
            ws_new_sheet.cell(row=15, column=17).value = budget_item.budget_price

            ws_new_sheet.cell(row=16, column=6).value = budget_item.budget_name

            ws_new_sheet.cell(row=19, column=2).value = budget_item.purpose
            ws_new_sheet.cell(row=19, column=14).value = budget_item.detail

            ws_new_sheet.cell(row=30, column=4).value = budget_item.delivery_date
            ws_new_sheet.cell(row=30, column=16).value = budget_item.start_date
            ws_new_sheet.cell(row=30, column=21).value = budget_item.end_date

            # 中止側処理
            if not normal_flag:
                work_cause_item = StopWorkCause.objects.filter(Q(target='probudgetunit') | Q(target='budget'), budget_id=budget_item.budget_id, lost_flag=0).all().order_by('-id')[0]
                ws_new_sheet.cell(row=6, column=14).value = work_cause_item.approval_no
                # 実行件数、累計額出力
                ws_new_sheet.cell(row=33, column=5).value = 0
                ws_new_sheet.cell(row=33, column=17).value = 0
            else:
                # 実行件数、累計額出力
                construction_count, construction_total_price = get_construction_data(budget_item.budget_id)
                ws_new_sheet.cell(row=33, column=5).value = construction_count
                ws_new_sheet.cell(row=33, column=17).value = construction_total_price

            # 年月日を出力
            ws_new_sheet.cell(row=1, column=24).value = \
                '(' + str(d_today.year) + '/' + str(d_today.month) + '/' + str(d_today.day) + ')'

            # 関連予算を出力
            if budget_item.budget_id == budget_id:
                # 主予算
                ws_new_sheet.cell(row=17, column=6).value = relation_budget_no_str
            else:
                # 関連予算
                ws_new_sheet.cell(row=17, column=6).value = main_budget_no_str

        # ダウンロード用ファイルを保存
        new_file_name_full_path = folder_name + new_file_name

        # フォルダが存在しない場合、フォルダを作成する
        if not pathlib.Path(folder_name).exists():
            os.makedirs(folder_name)

        # 指定ファイルが存在する場合、削除
        if pathlib.Path(new_file_name_full_path).exists():
            subprocess.run('del ' + new_file_name_full_path, shell=True)

        # ファイルに保存
        wb_new_file.save(new_file_name_full_path)

        # ファイルアップロード機能データ登録
        attachment_documents_data, created = AttachmentDocuments.objects.get_or_create(file_name=new_file_name, folder=attach_folder_name)
        attachment_documents_data.folder = attach_folder_name
        attachment_documents_data.div_id_name = current_tab
        attachment_documents_data.data = target_folder
        attachment_documents_data.entry_time = now
        attachment_documents_data.entry_user = request.user.username
        attachment_documents_data.lost_flag = 0
        attachment_documents_data.save()

        msg = '出力完了\nファイル一覧から取り出してください'
        data = {
            'msg': msg,
            'div_id_name': current_tab,
        }

        return JsonResponse(data)
    except Exception:
        output_log_exception(request, traceback.format_exc())
        raise
