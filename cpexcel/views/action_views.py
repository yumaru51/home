import os
import openpyxl
from openpyxl.styles import PatternFill
from django.db import connections
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.http.response import JsonResponse
# from isac.models import Orders, Stocks, StockAllocationResult, CompanyReservedItemMaster, OrderStatusMaster, QualityStatusMaster, CompanyMaster, GradeTypeMaster, WarehouseMaster, OrderStatusMaster, IccidentalWorkMaster
from cpexcel.models import Programs, Files, Tasks, ExecuteLogs, TaskEntryLogs, ProgramEntryLogs
import datetime
import calendar
from config.settings.settings_common import BASE_DIR


@login_required
@require_POST
def program_entry(request):

    DIFF_JST_FROM_UTC = 9
    # JST = timezone(timedelta(hours=+9), 'JST')

    now = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)

    operator = request.user.username

    # program_id = int(request.POST['program_id'])
    program_id = int(request.POST['program_id'])
    program_name = request.POST['program_name']
    description = request.POST['description']
    source_file_folder_path = request.POST['source_file_folder_path']
    source_file = request.POST['source_file']
    post_to_file_path = request.POST['post_to_file_path']
    post_to_file = request.POST['post_to_file']

    if program_id == 0:
        Programs(entry_date=now, entry_operator=operator).save()

        program_data = Programs.objects.all().get(entry_date=now, entry_operator=operator)
        program_id = program_data.id

    else:
        program_data = Programs.objects.all().get(id=program_id)
        program_id = program_data.id

    program_data.program_name = program_name
    program_data.description = description
    program_data.save()

    file_data, created = Files.objects.get_or_create(program_id=program_id)
    file_data.source_file_path = source_file_folder_path
    file_data.source_file_name = source_file
    file_data.post_to_file_path = post_to_file_path
    file_data.post_to_file_name = post_to_file
    file_data.lost_flag = 0
    file_data.save()

    ProgramEntryLogs(program_id=program_id, program_name=program_name, description=description, source_file_path=source_file_folder_path, source_file_name=source_file, post_to_file_path=post_to_file_path, post_to_file_name=post_to_file, entry_operator=operator, entry_date=now).save()

    program_data = Programs.objects.all().get(id=program_id)
    program_name = program_data.program_name
    description = program_data.description
    file_data = Files.objects.all().get(program_id=program_id)
    source_file_folder_path = file_data.source_file_path
    source_file = file_data.source_file_name
    post_to_file_path = file_data.post_to_file_path
    post_to_file = file_data.post_to_file_name

    program_id_str = str(program_data.id)
    label = "　更新　"

    data = {
        'program_name': program_name,
        'description': description,
        'source_file_folder_path': source_file_folder_path,
        'source_file': source_file,
        'post_to_file_path': post_to_file_path,
        'post_to_file': post_to_file,
        'program_id_str': program_id_str,
        'label': label
    }

    return TemplateResponse(request, 'cpexcel/parts/program_entry_page.html', data)


@login_required
@require_POST
def program_delete(request):

    DIFF_JST_FROM_UTC = 9
    # JST = timezone(timedelta(hours=+9), 'JST')

    now = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)

    operator = request.user.username

    # program_id = int(request.POST['program_id'])
    program_id = int(request.POST['program_id'])

    program_data = Programs.objects.all().get(id=program_id)
    program_data.lost_flag = 1
    program_data.save()

    file_data, created = Files.objects.get_or_create(program_id=program_id)
    file_data.lost_flag = 1
    file_data.save()

    ProgramEntryLogs(program_id=program_id, lost_flag=1, entry_operator=operator, entry_date=now).save()

    msg = "削除しました！！"

    ary = {
        'msg': msg
    }

    return JsonResponse(ary)


@login_required
@require_POST
def task_entry(request):

    DIFF_JST_FROM_UTC = 9
    # JST = timezone(timedelta(hours=+9), 'JST')

    now = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)

    operator = request.user.username

    program_id = int(request.POST['program_id'])
    task_id = int(request.POST['task_id'])
    source_sheet_name = request.POST['source_sheet_name']
    source_cell_name = request.POST['source_cell_name']
    post_to_sheet_name = request.POST['post_to_sheet_name']
    post_to_cell_name = request.POST['post_to_cell_name']

    if task_id == 0:
        Tasks(entry_date=now, entry_operator=operator).save()

        task_data = Tasks.objects.all().get(entry_date=now, entry_operator=operator)
        task_id = task_data.id

    else:
        task_data = Tasks.objects.all().get(id=task_id)
        task_id = task_data.id

    task_data.program_id = program_id
    task_data.source_sheet_name = source_sheet_name
    task_data.source_cell_name = source_cell_name
    task_data.post_to_sheet_name = post_to_sheet_name
    task_data.post_to_cell_name = post_to_cell_name
    task_data.lost_flag = 0
    task_data.update_date = now
    task_data.update_operator = operator
    task_data.save()

    TaskEntryLogs(task_id=task_id, source_sheet_name=source_sheet_name, source_cell_name=source_cell_name, post_to_sheet_name=post_to_sheet_name, post_to_cell_name=post_to_cell_name, entry_operator=operator, entry_date=now).save()

    msg = "タスクを登録しました！！"

    ary = {
        'msg': msg,
        'task_id': task_id
    }

    return JsonResponse(ary)


@login_required
@require_POST
def task_delete(request):

    DIFF_JST_FROM_UTC = 9
    # JST = timezone(timedelta(hours=+9), 'JST')

    now = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)

    operator = request.user.username

    # program_id = int(request.POST['program_id'])
    task_id = int(request.POST['task_id'])

    task_data = Tasks.objects.all().get(id=task_id)
    task_data.lost_flag = 1
    task_data.save()

    TaskEntryLogs(task_id=task_id, lost_flag=1, entry_operator=operator, entry_date=now).save()

    msg = "タスクを削除しました！！"

    ary = {
        'msg': msg
    }

    return JsonResponse(ary)


@login_required
@require_POST
def program_execute(request):

    DIFF_JST_FROM_UTC = 18
    # JST = timezone(timedelta(hours=+9), 'JST')

    start_datetime = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)

    operator = request.user.username

    # work_folder = 'C:\\\\PythonProjects\\\\isk_tools\\\\static\\\\files\\\\cpexcel\\\\' + operator + '\\\\'
    work_folder = BASE_DIR + '\\static\\files\\cpexcel\\' + operator + '\\'

    # program_id = int(request.GET.get('program_id'))
    program_id = int(request.POST['program_id'])

    file_data = Files.objects.all().get(program_id=program_id)
    source_file_folder_path = file_data.source_file_path
    source_file = file_data.source_file_name
    post_to_file_path = file_data.post_to_file_path
    post_to_file = file_data.post_to_file_name

    task_lists = Tasks.objects.all().filter(program_id=program_id, lost_flag=0).order_by('source_sheet_name', 'post_to_sheet_name')

    # new_source_file_folder_path = source_file_folder_path.replace('\\', '\\\\')

    # source_file_full_path = new_source_file_folder_path + "\\" + source_file
    source_file_full_path = work_folder + source_file
    print(source_file_full_path)
    # new_post_to_file_path = post_to_file_path.replace('\\', '\\\\')

    # post_to_file_full_path = new_post_to_file_path + "\\" + post_to_file
    post_to_file_full_path = work_folder + post_to_file
    print(post_to_file_full_path)
    wb_source = openpyxl.load_workbook(source_file_full_path, data_only=True)

    wb_post_to = openpyxl.load_workbook(post_to_file_full_path)

    i = 0

    for task_lists in task_lists:

        source_sheet_name = task_lists.source_sheet_name
        ws_source = wb_source[source_sheet_name]
        source_cell_name = task_lists.source_cell_name
        source_cell_value = ws_source[source_cell_name].value

        post_to_sheet_name = task_lists.post_to_sheet_name
        ws_post_to = wb_post_to[post_to_sheet_name]
        post_to_cell_name = task_lists.post_to_cell_name
        ws_post_to[post_to_cell_name].value = source_cell_value

        i += 1

    new_file_name_fullpath = work_folder + "作業完了_" + post_to_file

    wb_post_to.save(new_file_name_fullpath)

    # response = HttpResponse(content_type='application/vnd.ms-excel')
    # response['content-Disponsition'] = 'attachment; filename=作業完了_' + post_to_file

    # wb_post_to.save(response)


    end_datetime = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)

    ExecuteLogs(program_id=program_id, operator=operator, start_datetime=start_datetime, end_datetime=end_datetime).save()

    msg = str(i) + "件のタスク、実行完了！！"

    ary = {
        'msg': msg
    }

    return JsonResponse(ary)


@login_required
@require_POST
def file_change_check(request):
    # old_file_path = request.POST['old_file_path']
    # new_file_path = request.POST['new_file_path']
    old_file_name = request.POST['old_file_name']
    new_file_name = request.POST['new_file_name']

    operator = request.user.username

    # work_folder = 'C:\\\\PythonProjects\\\\isk_tools\\\\static\\\\files\\\\cpexcel\\\\' + operator + '\\\\'
    work_folder = BASE_DIR + '\\static\\files\\cpexcel\\' + operator + '\\'

    # new_file_folder_path = new_file_path.replace('\\', '\\\\')

    # new_file = new_file_folder_path + "\\" + new_file_name
    new_file = work_folder + "\\\\" + new_file_name

    # old_file_folder_path = old_file_path.replace('\\', '\\\\')

    # old_file = old_file_folder_path + "\\" + old_file_name
    old_file = work_folder + "\\\\" + old_file_name

    wb_old = openpyxl.load_workbook(old_file, data_only=True)

    wb_new = openpyxl.load_workbook(new_file, data_only=True)

    old_file_sheet_num = len(wb_old.sheetnames)

    new_file_sheet_num = len(wb_new.sheetnames)

    if old_file_sheet_num == new_file_sheet_num:
        change_flag = 0

    else:
        change_flag = 1

    if change_flag == 0:

        for i in range(old_file_sheet_num):

            old_file_sheet_name = wb_old.worksheets[i].title

            new_file_sheet_name = wb_new.worksheets[i].title

            if old_file_sheet_name != new_file_sheet_name:
                change_flag += 1

        if change_flag == 0:
            test = "通りました"
            for i in range(old_file_sheet_num):
                target_sheet_name = wb_old.worksheets[i].title
                old_file_sheet = wb_old[target_sheet_name]
                new_file_sheet = wb_new[target_sheet_name]

                old_file_max_row = wb_old[target_sheet_name].max_row
                new_file_max_row = wb_new[target_sheet_name].max_row

                if old_file_max_row >= new_file_max_row:
                    row_num = old_file_max_row
                else:
                    row_num = new_file_max_row

                old_file_max_column = wb_old[target_sheet_name].max_column
                new_file_max_column = wb_new[target_sheet_name].max_column

                if old_file_max_column >= new_file_max_column:
                    column_num = old_file_max_column
                else:
                    column_num = new_file_max_column

                for j in range(row_num):
                    for k in range(column_num):
                        old_file_cell_value = old_file_sheet.cell(row=j+1, column=k+1).value
                        new_file_cell_value = new_file_sheet.cell(row=j+1, column=k+1).value

                        if old_file_cell_value == new_file_cell_value:
                            no_change_flag = 1
                        else:
                            fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='FF0000', bgColor='FF0000')
                            new_file_sheet.cell(row=j+1, column=k+1).fill = fill
                            change_flag += 1

    new_file_name_fullpath = work_folder + "\\確認完了_" + new_file_name

    wb_new.save(new_file_name_fullpath)

    msg = "確認完了！！"

    ary = {
        'msg': msg
    }

    return JsonResponse(ary)


