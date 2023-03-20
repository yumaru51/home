import openpyxl
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.http.response import JsonResponse
import sys, os
import datetime
from cpexcel.models import Tasks
from config.settings.settings_common import BASE_DIR


@login_required
@require_GET
def file_download(request):

    file_name = request.GET.get('file_name')

    operator = request.user.username

    work_folder = BASE_DIR + '\\static\\files\\cpexcel\\' + operator

    file_full_path = work_folder + file_name

    wb = openpyxl.load_workbook(file_full_path)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['content-Disponsition'] = 'attachment; filename= file_name'

    wb.save(response)

    return response



@login_required
def file_upload(request):

    t_username = request.user.username

    # BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # BASE_DIR = 'C:\\PythonProjects\\isk_tools\\'
    # UPLOADE_DIR = os.path.dirname(os.path.abspath(__file__)) + '/static/files/' + t_username + '/'
    UPLOADE_DIR = os.path.join(BASE_DIR, '\\static\\files\\\cpexcel\\' + t_username + '\\')
    print(UPLOADE_DIR)
    if request.method != 'POST':
        # return render(request, 'upload_form/form.html')
        msg = "アップロードできませんでした！！"

    else:
        file = request.FILES['file']
        path = os.path.join(UPLOADE_DIR, file.name)
        destination = open(path, 'wb')

        for chunk in file.chunks():
            destination.write(chunk)

        destination.close()

        msg = "アップロード完了！！"

    ary = {
        'msg': msg
    }

    return JsonResponse(ary)


@login_required
def file_import(request):

    DIFF_JST_FROM_UTC = 18
    # JST = timezone(timedelta(hours=+9), 'JST')

    now = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)

    t_username = request.user.username

    # BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # BASE_DIR = 'C:\\PythonProjects\\isk_tools\\'
    # UPLOADE_DIR = os.path.dirname(os.path.abspath(__file__)) + '/static/files/' + t_username + '/'
    UPLOADE_DIR = os.path.join(BASE_DIR, '\\static\\files\\cpexcel\\' + t_username + '\\')
    print(UPLOADE_DIR)
    if request.method != 'POST':
        # return render(request, 'upload_form/form.html')
        msg = "アップロードできませんでした！！"

    else:
        file = request.FILES['file']
        path = os.path.join(UPLOADE_DIR, file.name)
        destination = open(path, 'wb')

        for chunk in file.chunks():
            destination.write(chunk)

        destination.close()

        wb = openpyxl.load_workbook(path)

        data_sheet = wb['DATA']

        row_num = data_sheet.max_row

        msg = ""

        if data_sheet.cell(row=1, column=2).value != "":
            program_id = data_sheet.cell(row=1, column=2).value

            j = 0
            k = 0

            for i in range(5, row_num+1):

                if data_sheet.cell(row=i, column=1).value == "record":

                    source_sheet_name = data_sheet.cell(row=i, column=3).value
                    source_cell_name = data_sheet.cell(row=i, column=4).value
                    post_to_sheet_name = data_sheet.cell(row=i, column=5).value
                    post_to_cell_name = data_sheet.cell(row=i, column=6).value
                    lost_flag = data_sheet.cell(row=i, column=7).value

                    if data_sheet.cell(row=i, column=2).value is None:

                        Tasks(program_id=program_id, source_sheet_name=source_sheet_name, source_cell_name=source_cell_name, post_to_sheet_name=post_to_sheet_name, post_to_cell_name=post_to_cell_name, entry_date=now, entry_operator=t_username, lost_flag=0).save()

                        k += 1

                    else:

                        task_id = data_sheet.cell(row=i, column=2).value
                        task_data = Tasks.objects.get(id=task_id)
                        task_data.source_sheet_name = source_sheet_name
                        task_data.source_cell_name = source_cell_name
                        task_data.post_to_sheet_name = post_to_sheet_name
                        task_data.post_to_cell_name = post_to_cell_name
                        task_data.lost_flag = lost_flag
                        task_data.update_date = now
                        task_data.update_operator = t_username
                        task_data.save()

                        j += 1

            msg = str(j) + "件更新登録、" + str(k) + "件新規登録！！"

        else:
            msg = "プログラムidが指定されていないため、登録不可！！"

    ary = {
        'msg': msg
    }

    return JsonResponse(ary)


@login_required
@require_GET
def task_import_file_download(request):

    program_id = int(request.GET.get('program_id'))

    operator = request.user.username

    # file_full_path = 'C:\\\\PythonProjects\\\\isk_tools\\\\static\\\\files\\\\cpexcel\\\\task_import_file.xlsx'
    file_full_path = BASE_DIR + '\\static\\files\\cpexcel\\task_import_file.xlsx'

    # file_full_path = work_folder + file_name

    wb = openpyxl.load_workbook(file_full_path)

    data_sheet = wb['DATA']

    task_lists = Tasks.objects.filter(program_id=program_id)

    data_sheet.cell(row=1, column=2).value = program_id

    i = 5

    for task_lists in task_lists:

        data_sheet.cell(row=i, column=1).value = "record"
        data_sheet.cell(row=i, column=2).value = task_lists.id
        data_sheet.cell(row=i, column=3).value = task_lists.source_sheet_name
        data_sheet.cell(row=i, column=4).value = task_lists.source_cell_name
        data_sheet.cell(row=i, column=5).value = task_lists.post_to_sheet_name
        data_sheet.cell(row=i, column=6).value = task_lists.post_to_cell_name
        data_sheet.cell(row=i, column=7).value = task_lists.lost_flag

        i += 1

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['content-Disponsition'] = 'attachment; filename= file_name'

    wb.save(response)

    return response


@require_GET
def demo_file_download(request):
    file_name = request.GET.get('file_name')

    # file_full_path = 'C:\\\\PythonProjects\\\\isk_tools\\\\static\\\\files\\\\cpexcel\\\\' + file_name
    file_full_path = BASE_DIR + '\\static\\files\\cpexcel\\' + file_name

    wb = openpyxl.load_workbook(file_full_path)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['content-Disponsition'] = 'attachment; filename= file_name'

    wb.save(response)

    return response

