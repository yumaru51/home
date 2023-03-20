import datetime
import mimetypes
import glob
import os
# ログインユーザーを使用するmoduleをインポート
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
# postからの引数を使用できるmoduleをインポート
from django.views.decorators.http import require_POST
# excel操作
import openpyxl
# SQLSERVER接続
import pyodbc


# 削除予定
def quality_daily_report1(request):
    # DB接続
    server = 'YSQLSERV4'
    database = 'OP_ENTRY_M3P'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    data = {
        'test': 'test',
    }
    return render(request, 'reportoutput/機能材/品質日報.html', data)


# 機能材_品質日報
def quality_daily_report2(request, startTime, endTime):
    starttime = startTime + ' 07:00:00'
    endtime = endTime + ' 07:00:00'
    print(starttime)
    print(endtime)
    # DB接続
    server = 'YSQLSERV4'
    database = 'OP_ENTRY_M3P'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # トリガー起動
    cursor.execute(" \
    UPDATE [OP_ENTRY_M3P].[dbo].[W_QualityDailyReportTrigger] \
    SET \
        STARTTIME = ?, \
        ENDTIME = ? \
    WHERE UPDATE_KEY = 1 \
    ", starttime, endtime)
    cnxn.commit()
    print('トリガー起動!!!')

    # 品質日報 77項目
    cursor.execute(" \
SELECT \
    [GRADE_ABBR_NM], \
    CAST([SAMPLING_DATE] AS DATE), \
    FORMAT([SAMPLING_DATE], 'HH:mm') AS 時間, \
    [LOT_SUB_01], \
    [LOT_NO], \
    [検査項目1], \
    [検査項目2], \
    [検査項目3], \
    [検査項目4], \
    [検査項目5], \
    [検査項目6], \
    [検査項目7], \
    [検査項目8], \
    [検査項目9], \
    [検査項目10], \
    [検査項目11], \
    [検査項目12], \
    [検査項目13], \
    [検査項目14], \
    [検査項目15], \
    [検査項目16], \
    [検査項目17], \
    [検査項目18], \
    [検査項目19], \
    [検査項目20], \
    [検査項目21], \
    [検査項目22], \
    [検査項目23], \
    [検査項目24], \
    [検査項目25], \
    [検査項目26], \
    [検査項目27], \
    [検査項目28], \
    [検査項目29], \
    [検査項目30], \
    [検査項目31], \
    [検査項目32], \
    [検査項目33], \
    [検査項目34], \
    [検査項目35], \
    [検査項目36], \
    [検査項目37], \
    [検査項目38], \
    [検査項目39], \
    [検査項目40], \
    [検査項目41], \
    [検査項目42], \
    [検査項目43], \
    [検査項目44], \
    [検査項目45], \
    [検査項目46], \
    [検査項目47], \
    [検査項目48], \
    [検査項目49], \
    [検査項目50], \
    [検査項目51], \
    [検査項目52], \
    [検査項目53], \
    [検査項目54], \
    [検査項目55], \
    [検査項目56], \
    [検査項目57], \
    [検査項目58], \
    [検査項目59], \
    [検査項目60], \
    [検査項目61], \
    [検査項目62], \
    [検査項目63], \
    [検査項目64], \
    [検査項目65], \
    [検査項目66], \
    [検査項目67], \
    [検査項目68], \
    [検査項目69], \
    [検査項目70], \
    [コメント], \
    [データ区分] \
FROM [OP_ENTRY_M3P].[dbo].[W_QualityDailyReport2] \
ORDER BY GRADE, データ区分, SAMPLING_DATE\
")
    tracereport1 = cursor.fetchall()
    query_list1 = tracereport1

    # 空のEXCELを開く
    wb = openpyxl.Workbook()
    ws = wb.active

    # sheet名
    ws.title = "品質日報"
    # sheetTabColor
    ws.sheet_properties.tabColor = "1072BA"

    # フォントの設定--------------------------------------------------------------------------------------------------------------------------------------------
    # font = openpyxl.styles.Font(
    #     name="HGPｺﾞｼｯｸE",
    #     size=15,
    #     bold=False,
    #     italic=False,
    #     underline='none',
    #     strike=False,
    #     color="FFBB00",
    # )

    # セルの塗りつぶし------------------------------------------------------------------------------------------------------------------------------------------
    fill1 = openpyxl.styles.PatternFill(
        fill_type='solid',
        fgColor='adf3f5',
    )
    fill2 = openpyxl.styles.PatternFill(
        fill_type='solid',
        fgColor='f7a287',
    )

    # 枠線------------------------------------------------------------------------------------------------------------------------------------------------------
    # set border (black thin line)
    side = openpyxl.styles.borders.Side(style='thin', color='000000')
    # set border (black thin line)
    border = openpyxl.styles.borders.Border(top=side, bottom=side, left=side, right=side)

    # データ格納------------------------------------------------------------------------------------------------------------------------------------------------
    # カラム
    # ws.cell(row=1, column=1).value = 'SAMPLING_DATE'
    # ws.cell(row=1, column=2).value = 'GRADE'
    # ws.cell(row=1, column=3).value = 'GRADE_ABBR_NM'
    # ws.cell(row=1, column=4).value = 'LOT_SUB_0'
    # ws.cell(row=1, column=5).value = 'LOT_NO'
    # ws.cell(row=1, column=6).value = 'データ区分'

    # データ
    for i in range(0, query_list1.__len__(), 1):

        # データ区分　1:項目名, 2:規格, 3:値, 4:空
        # 塗りつぶし　1:青, 2:赤
        # 枠線        1,2,3
        if query_list1[i][76] == '1':
            for k in range(1, 78, 1):
                ws.cell(row=i + 2, column=k).fill = fill1
                ws.cell(row=i + 2, column=k).border = border
        if query_list1[i][76] == '2':
            for k in range(1, 78, 1):
                ws.cell(row=i + 2, column=k).fill = fill2
                ws.cell(row=i + 2, column=k).border = border
        if query_list1[i][76] == '3':
            for k in range(1, 78, 1):
                ws.cell(row=i + 2, column=k).border = border

        # 書き出し
        for j in range(0, 77, 1):
            # 「値がNone」または「日付型」ならそのまま書き出し
            if query_list1[i][j] is None or type(query_list1[i][j]) == datetime.date:
                ws.cell(row=i + 2, column=j + 1).value = query_list1[i][j]
            # float変換可能なら変換
            elif is_num(query_list1[i][j]):
                ws.cell(row=i + 2, column=j + 1).value = float(query_list1[i][j])
            # float変換不可ならそのまま書き出し
            else:
                ws.cell(row=i + 2, column=j + 1).value = query_list1[i][j]

    # セル幅自動調整--------------------------------------------------------------------------------------------------------------------------------------------
    for col in ws.columns:  # sheet1の列のループを行う
        max_length = 0  # ゼロをmax_lengthに代入
        column = col[0].column_letter  # 列を取得。どこから取っても同じため先頭から取得。

        for cell in col:  # 列に含まれるセルに対してループを行う
            if len(str(cell.value)) > max_length:  # 文字列数が変数max_length（ゼロ）より大きかった場合
                max_length = len(str(cell.value))  # 文字列数に変数max_lengthを代入

        adjusted_width = (max_length + 2) * 1.2  # (max_length + 2) * 1.2の列幅を変数adjusted_widthに代入
        ws.column_dimensions[column].width = adjusted_width  # adjusted_widthの値をsheetの列幅の寸法とする。

    # 枠線------------------------------------------------------------------------------------------------------------------------------------------------------
    # write in sheet
    # for row in ws:
        # for cell in row:
        #     ws[cell.coordinate].border = border

    # 書式適用--------------------------------------------------------------------------------------------------------------------------------------------------
    # # write in sheet
    # for row in ws:
    #     for cell in row:
    #         ws[cell.coordinate].font = font

    # 保存(上書き)
    file_path = '\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\機能材\\品質日報.xlsx'
    file_name = '品質日報.xlsx'
    wb.save(file_path)

    with open(file_path, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(file_name)[0] or 'application/octet-stream')
    return response


# float変換関数(NULLなどはエラーとなるためフォーマットを整えてから送ること)
def is_num(s):
    try:
        float(s)
    except ValueError:
        return False
    else:
        return True

