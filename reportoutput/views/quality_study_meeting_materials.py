import datetime
import mimetypes
import glob
import os
# ログインユーザーを使用するmoduleをインポート
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
# postからの引数を使用できるmoduleをインポート
from django.views.decorators.http import require_POST
# excel操作
import openpyxl
# SQLSERVER接続
import pyodbc
from socket import gethostname


# 酸化チタン_CL法_品質検討会資料
def quality_study_meeting_materials(request):
    # DB接続
    host_name = gethostname()
    if host_name == 'YWEBSERV1':
        print('本番環境起動')
        server = 'YSQLSERV4'
    elif host_name == 'I7161DD6':
        print('テスト環境起動')
        server = 'Y0033OUT,1434'
    else:
        print('開発環境起動')
        server = 'Y0033OUT,1434'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # SQL-トレース履歴銘柄一覧
    cursor.execute("SELECT [銘柄] FROM [T_TRACE_レポート引数_履歴] GROUP BY [銘柄]")
    tracereport1 = cursor.fetchall()

    # SQL-トレース履歴一覧
    cursor.execute("SELECT [銘柄], [ランNO] FROM [T_TRACE_レポート引数_履歴] GROUP BY [銘柄], [ランNO]")
    tracereport2 = cursor.fetchall()

    data = {
        'tracereport1': tracereport1,
        'tracereport2': tracereport2,
    }
    return render(request, 'reportoutput/CL法/品質検討会資料.html', data)


# 酸化チタン_CL法_品質検討会資料
def __select__brand(request):
    # 変数宣言
    value = request.POST['value']

    # DB接続
    host_name = gethostname()
    if host_name == 'YWEBSERV1':
        print('本番環境起動')
        server = 'YSQLSERV4'
    elif host_name == 'I7161DD6':
        print('テスト環境起動')
        server = 'Y0033OUT,1434'
    else:
        print('開発環境起動')
        server = 'Y0033OUT,1434'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # SQL-トレース履歴一覧
    cursor.execute("SELECT [銘柄], [ランNO] FROM [T_TRACE_REPORT_縦持ち_履歴] WHERE [銘柄] LIKE ? GROUP BY [銘柄], [ランNO] ORDER BY [ランNO] DESC", value)
    tracereport2 = cursor.fetchall()
    select = ''
    for tracereport2 in tracereport2:
        select += '<option value = ' + str(tracereport2.ランNO) + '>　' + str(tracereport2.ランNO) + '　</option>'

    data = {
        'select': select,
    }
    return JsonResponse(data)


# 酸化チタン_CL法_品質検討会資料
def trace_report(request, brand, run_no1, run_no2, run_no3):
    # 前回ファイル削除
    file_list = glob.glob("\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\品質検討会資料\\*品質検討会資料(CL法).xlsx")
    for file in file_list:
        # print("remove：{0}".format(file))
        os.remove(file)

    # DB接続
    host_name = gethostname()
    if host_name == 'YWEBSERV1':
        print('本番環境起動')
        server = 'YSQLSERV4'
    elif host_name == 'I7161DD6':
        print('テスト環境起動')
        server = 'Y0033OUT,1434'
    else:
        print('開発環境起動')
        server = 'Y0033OUT,1434'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    cursor.execute(" \
    SELECT \
                    CONCAT(IIF([M_QC-One_品検項目].[表記名] = '', '', CONCAT([M_QC-One_品検項目].[表記名], '_') ), [M_QC-One_品検項目].[出力項目] ) AS 出力項目 \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド2] AS 最小 \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド1] AS 最大 \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] AS 平均 \
    FROM            [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_縦持ち_履歴] \
    INNER JOIN      [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON              [T_TRACE_REPORT_縦持ち_履歴].[ID]		= [M_QC-One_品検項目].[SEQ] \
    AND			    [M_QC-One_品検項目].[工場]			= '0010' \
    AND			    [T_TRACE_REPORT_縦持ち_履歴].[項目名]	= [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	    [OP_ENTRY_INORG].[dbo].[T_TRACE_レポート引数_履歴] \
    ON			    [T_TRACE_レポート引数_履歴].[銘柄]		  = [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
    AND			    [T_TRACE_レポート引数_履歴].[ランNO]	  = [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
    AND 			[T_TRACE_レポート引数_履歴] .[銘柄]       = ? \
    AND			    [T_TRACE_レポート引数_履歴] .[ランNO]     = ? \
    GROUP BY	    CONCAT( IIF ( [M_QC-One_品検項目].[表記名] = '', '',CONCAT ( [M_QC-One_品検項目 ].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド1] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド2] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[ID] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日自] \
                    , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日至] \
                    , [M_QC-One_品検項目].[出力順序] \
    ORDER BY        [M_QC-One_品検項目].[出力順序]", brand, run_no1, )
    tracereport = cursor.fetchall()
    aggregate_list = tracereport
    cursor.execute(" \
    SELECT \
                    CONCAT(IIF([M_QC-One_品検項目].[表記名] = '', '', CONCAT([M_QC-One_品検項目].[表記名], '_') ), [M_QC-One_品検項目].[出力項目] ) AS 出力項目 \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] AS 平均 \
    FROM		    [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_縦持ち_履歴] \
    INNER JOIN	    [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON			    [T_TRACE_REPORT_縦持ち_履歴].[ID] = [M_QC-One_品検項目].[SEQ] \
    AND			    [M_QC-One_品検項目].[工場] = '0010' \
    AND			    [T_TRACE_REPORT_縦持ち_履歴].[項目名 ] = [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	    [OP_ENTRY_INORG].[dbo].[T_TRACE_レポート引数_履歴] \
    ON			    [T_TRACE_レポート引数_履歴].[銘柄]    = [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
    AND			    [T_TRACE_レポート引数_履歴].[ランNO]  = [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
    AND			    [T_TRACE_レポート引数_履歴] .[銘柄]   = ? \
    AND			    [T_TRACE_レポート引数_履歴] .[ランNO] = ? \
    GROUP BY        CONCAT( IIF ( [M_QC-One_品検項目].[表記名] = '', '',CONCAT ( [M_QC-One_品検項目 ].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド1] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド2] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[ID] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日自] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日至] \
                  , [M_QC-One_品検項目].[出力順序] \
    ORDER BY	    [M_QC-One_品検項目].[出力順序]", brand, run_no2, )
    tracereport = cursor.fetchall()
    avg2s_list = tracereport
    cursor.execute(" \
    SELECT \
                    CONCAT(IIF([M_QC-One_品検項目].[表記名] = '', '', CONCAT([M_QC-One_品検項目].[表記名], '_') ), [M_QC-One_品検項目].[出力項目] ) AS 出力項目 \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] AS 平均 \
    FROM		    [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_縦持ち_履歴] \
    INNER JOIN	    [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON			    [T_TRACE_REPORT_縦持ち_履歴].[ID] = [M_QC-One_品検項目].[SEQ] \
    AND			    [M_QC-One_品検項目].[工場] = '0010' \
    AND			    [T_TRACE_REPORT_縦持ち_履歴].[項目名 ] = [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	    [OP_ENTRY_INORG].[dbo].[T_TRACE_レポート引数_履歴] \
    ON			    [T_TRACE_レポート引数_履歴].[銘柄] = [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
    AND			    [T_TRACE_レポート引数_履歴].[ランNO] = [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
    AND			    [T_TRACE_レポート引数_履歴] .[銘柄]		= ? \
    AND			    [T_TRACE_レポート引数_履歴] .[ランNO]		= ? \
    GROUP BY        CONCAT( IIF ( [M_QC-One_品検項目].[表記名] = '', '',CONCAT ( [M_QC-One_品検項目 ].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド1] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド2] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[フィールド3] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[ID] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[銘柄] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[ランNO] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日自] \
                  , [T_TRACE_REPORT_縦持ち_履歴].[操業年月日至] \
                  , [M_QC-One_品検項目].[出力順序] \
    ORDER BY	    [M_QC-One_品検項目].[出力順序]", brand, run_no3, )
    tracereport = cursor.fetchall()
    avg3s_list = tracereport

    cursor.execute(" \
    SELECT \
        TOP 1 \
        操業年月日自, \
        操業年月日至 \
    FROM \
        [OP_ENTRY_INORG].[dbo].[T_TRACE_REPORT_縦持ち_履歴] \
    WHERE \
        銘柄 = ? \
        AND ランNO = ? \
    ", brand, run_no1, )
    tracereport = cursor.fetchall()
    for tracereport in tracereport:
        operation_date_start = tracereport.操業年月日自.strftime('%Y%m%d%H%M')
        operation_date_start = operation_date_start[0:4] + "/" + operation_date_start[4:6] + "/" + operation_date_start[6:8]
        operation_date_end = tracereport.操業年月日至.strftime('%Y%m%d%H%M')
        operation_date_end = operation_date_end[0:4] + "/" + operation_date_end[4:6] + "/" + operation_date_end[6:8]

    print('データ取得完了')

    # 現在日時を取得。ファイル名に加える --- (*1)
    today = datetime.datetime.now().strftime('%Y%m%d%H%M')

    # EXCEL読込 --- (*2)
    template = '\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\テンプレート\\【帳票テンプレート】品質検討会資料(CL法).xlsx'
    print(template)
    wb = openpyxl.load_workbook(template)

    # アクティブなワークシートを得る --- (*3)
    ws = wb["品質検討会資料"]
    ws2 = wb["form"]
    ws3 = wb["pos"]

    # データ書き込み --- (*4)
    ws["F1"] = "XXXX年    X月度    " + brand + "    XX工程"
    ws["A2"] = "（ " + str(run_no1) + " ）前    回    ラ    ン    概    況 [ " + operation_date_start + " ～ " + operation_date_end + " ]"

    for i in range(5, aggregate_list.__len__()+5, 1):
        ws['A' + str(i)] = aggregate_list[i-5][0]
        ws['F' + str(i)] = aggregate_list[i-5][1]
        ws['H' + str(i)] = aggregate_list[i-5][2]
        ws['J' + str(i)] = aggregate_list[i-5][3]
        # for j in range(0, avg1s_list.__len__(), 1):
        #     if aggregate_list[i-5][0] == avg1s_list[j][0]:
        #         ws['J' + str(i)] = avg1s_list[j][1]
        for j in range(0, avg2s_list.__len__(), 1):
            if aggregate_list[i-5][0] == avg2s_list[j][0]:
                ws['L' + str(i)] = avg2s_list[j][1]
        for j in range(0, avg3s_list.__len__(), 1):
            if aggregate_list[i-5][0] == avg3s_list[j][0]:
                ws['N' + str(i)] = avg3s_list[j][1]
        # 自動集計部最大項目数以上は取得不可 テンプレートに収まりきらないので切り捨てる。
        if i == 64:
            break

    print('データ抽出完了')

    # パターンの取得 ["pos"]から銘柄比較してパターン取得
    pattern = ''
    for row in ws3.iter_rows(min_row=51, max_row=451, min_col=2, max_col=2):
        for data in row:
            if data.value == brand:
                pattern = ws3.cell(row=data.row, column=data.column + 1).value

    # パターン毎に処理分岐
    if pattern == 'A':
        for i in range(1, 8):
            for j in range(1, 18):
                copy = ws2.cell(row=i+23, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'B':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+32, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'C':
        for i in range(1, 12):
            for j in range(1, 18):
                copy = ws2.cell(row=i+41, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'D':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+53, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'E':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+62, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'F':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+66, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'G':
        for i in range(1, 11):
            for j in range(1, 18):
                copy = ws2.cell(row=i+70, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'H':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+81, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'I':
        for i in range(1, 14):
            for j in range(1, 18):
                copy = ws2.cell(row=i+85, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'J':
        for i in range(1, 7):
            for j in range(1, 18):
                copy = ws2.cell(row=i+99, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'K':
        for i in range(1, 8):
            for j in range(1, 18):
                copy = ws2.cell(row=i+106, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'L':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+114, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'M':
        for i in range(1, 5):
            for j in range(1, 18):
                copy = ws2.cell(row=i+123, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'N':
        for i in range(1, 5):
            for j in range(1, 18):
                copy = ws2.cell(row=i+128, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)

    # header
    ws.oddHeader.right.text = today[0:4] + "年" + today[4:6] + "月" + today[6:8] + "日\n"

    # 署名
    img = openpyxl.drawing.image.Image("\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\テンプレート\\署名.png")
    ws.add_image(img, 'F92')

    # ファイルを保存 --- (*5)
    file_path = '\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\品質検討会資料\\' + today + '_品質検討会資料(CL法).xlsx'
    file_name = today + '_品質検討会資料(CL法).xlsx'
    wb.save(file_path)

    with open(file_path, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(file_name)[0] or 'application/octet-stream')
    return response


# 酸化チタン_S法_品質検討会資料
def s_quality_study_meeting_materials(request):
    # DB接続
    host_name = gethostname()
    if host_name == 'YWEBSERV1':
        print('本番環境起動')
        server = 'YSQLSERV4'
    elif host_name == 'I7161DD6':
        print('テスト環境起動')
        server = 'Y0033OUT,1434'
    else:
        print('開発環境起動')
        server = 'Y0033OUT,1434'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # SQL-トレース履歴銘柄一覧
    cursor.execute("SELECT [銘柄] FROM [T_TRACE_S法_レポート引数_履歴] GROUP BY [銘柄]")
    tracereport1 = cursor.fetchall()

    # SQL-トレース履歴一覧
    cursor.execute("SELECT [銘柄], [ランNO] FROM [T_TRACE_S法_レポート引数_履歴] GROUP BY [銘柄], [ランNO]")
    tracereport2 = cursor.fetchall()

    data = {
        'tracereport1': tracereport1,
        'tracereport2': tracereport2,
    }
    return render(request, 'reportoutput/S法/品質検討会資料.html', data)


# 酸化チタン_S法_品質検討会資料
def s___select__brand(request):
    # 変数宣言
    value = request.POST['value']

    # DB接続
    host_name = gethostname()
    if host_name == 'YWEBSERV1':
        print('本番環境起動')
        server = 'YSQLSERV4'
    elif host_name == 'I7161DD6':
        print('テスト環境起動')
        server = 'Y0033OUT,1434'
    else:
        print('開発環境起動')
        server = 'Y0033OUT,1434'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # SQL-トレース履歴一覧
    cursor.execute("SELECT [銘柄], [ランNO] FROM [T_TRACE_S法_REPORT_縦持ち_履歴] WHERE [銘柄] LIKE ? GROUP BY [銘柄], [ランNO] ORDER BY [ランNO] DESC", value)
    tracereport2 = cursor.fetchall()
    select = ''
    for tracereport2 in tracereport2:
        select += '<option value = ' + str(tracereport2.ランNO) + '>　' + str(tracereport2.ランNO) + '　</option>'

    data = {
        'select': select,
    }
    return JsonResponse(data)


# 酸化チタン_S法_品質検討会資料
def s_trace_report(request, brand, run_no1, run_no2, run_no3):

    # 前回ファイル削除
    file_list = glob.glob("\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\品質検討会資料\\*品質検討会資料(S法).xlsx")
    for file in file_list:
        # print("remove：{0}".format(file))
        os.remove(file)

    # DB接続
    host_name = gethostname()
    if host_name == 'YWEBSERV1':
        print('本番環境起動')
        server = 'YSQLSERV4'
    elif host_name == 'I7161DD6':
        print('テスト環境起動')
        server = 'Y0033OUT,1434'
    else:
        print('開発環境起動')
        server = 'Y0033OUT,1434'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    cursor.execute(" \
    SELECT \
                  CONCAT( IIF( [M_QC-One_品検項目].[表記名] = '', '', CONCAT( [M_QC-One_品検項目].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] )	AS 出力項目 \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド2]			AS 最小 \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド1]			AS 最大	\
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3]			AS 平均 \
    FROM		  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_REPORT_縦持ち_履歴] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON			  [T_TRACE_S法_REPORT_縦持ち_履歴].[ID]		= [M_QC-One_品検項目].[SEQ] \
    AND			  [M_QC-One_品検項目].[工場]			= '0020' \
    AND			  [T_TRACE_S法_REPORT_縦持ち_履歴].[項目名]	= [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_レポート引数_履歴] \
    ON			  [T_TRACE_S法_レポート引数_履歴].[銘柄]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
    AND			  [T_TRACE_S法_レポート引数_履歴].[ランNO]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
    AND			  [T_TRACE_S法_レポート引数_履歴].[銘柄]		= ? \
    AND			  [T_TRACE_S法_レポート引数_履歴].[ランNO]		= ? \
    GROUP BY	  CONCAT( IIF( [M_QC-One_品検項目].[表記名] = '', '', CONCAT( [M_QC-One_品検項目].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド1] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド2] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ID] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日自] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日至] \
                , [M_QC-One_品検項目].[出力順序] \
    ORDER BY	  [M_QC-One_品検項目].[出力順序]", brand, run_no1, )
    tracereport = cursor.fetchall()
    aggregate_list = tracereport

    cursor.execute(" \
    SELECT \
                  CONCAT(IIF([M_QC-One_品検項目].[表記名] = '', '', CONCAT([M_QC-One_品検項目].[表記名], '_') ), [M_QC-One_品検項目].[出力項目] ) AS 出力項目 \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3] AS 平均 \
    FROM		  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_REPORT_縦持ち_履歴] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON			  [T_TRACE_S法_REPORT_縦持ち_履歴].[ID]		= [M_QC-One_品検項目].[SEQ] \
    AND			  [M_QC-One_品検項目].[工場]			= '0020' \
    AND			  [T_TRACE_S法_REPORT_縦持ち_履歴].[項目名 ]	= [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_レポート引数_履歴] \
    ON			  [T_TRACE_S法_レポート引数_履歴].[銘柄 ]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
    AND			  [T_TRACE_S法_レポート引数_履歴].[ランNO]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
    AND			  [T_TRACE_S法_レポート引数_履歴] .[銘柄]		= ? \
    AND			  [T_TRACE_S法_レポート引数_履歴] .[ランNO]		= ? \
    GROUP BY	  CONCAT( IIF ( [M_QC-One_品検項目].[表記名] = '', '',CONCAT ( [M_QC-One_品検項目].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド1] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド2] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ID] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日自] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日至] \
                , [M_QC-One_品検項目].[出力順序] \
    ORDER BY	  [M_QC-One_品検項目].[出力順序]", brand, run_no2, )
    tracereport = cursor.fetchall()
    avg2s_list = tracereport

    cursor.execute(" \
    SELECT \
                  CONCAT(IIF([M_QC-One_品検項目].[表記名] = '', '', CONCAT([M_QC-One_品検項目].[表記名], '_') ), [M_QC-One_品検項目].[出力項目] ) AS 出力項目 \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3] AS 平均 \
    FROM		  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_REPORT_縦持ち_履歴] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[M_QC-One_品検項目] \
    ON			  [T_TRACE_S法_REPORT_縦持ち_履歴].[ID]		= [M_QC-One_品検項目].[SEQ] \
    AND			  [M_QC-One_品検項目].[工場]			= '0020' \
    AND			  [T_TRACE_S法_REPORT_縦持ち_履歴].[項目名 ]	= [M_QC-One_品検項目].[出力項目] \
    INNER JOIN	  [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_レポート引数_履歴] \
    ON			  [T_TRACE_S法_レポート引数_履歴].[銘柄 ]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
    AND			  [T_TRACE_S法_レポート引数_履歴].[ランNO]		= [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
    AND			  [T_TRACE_S法_レポート引数_履歴] .[銘柄]		= ? \
    AND			  [T_TRACE_S法_レポート引数_履歴] .[ランNO]		= ? \
    GROUP BY	  CONCAT( IIF ( [M_QC-One_品検項目].[表記名] = '', '',CONCAT ( [M_QC-One_品検項目].[表記名],'_' ) ) , [M_QC-One_品検項目].[出力項目] ) \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド1] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド2] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[フィールド3] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ID] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[銘柄] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[ランNO] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日自] \
                , [T_TRACE_S法_REPORT_縦持ち_履歴].[操業年月日至] \
                , [M_QC-One_品検項目].[出力順序] \
    ORDER BY	  [M_QC-One_品検項目].[出力順序]", brand, run_no3, )
    tracereport = cursor.fetchall()
    avg3s_list = tracereport

    cursor.execute(" \
    SELECT \
        TOP 1 \
        操業年月日自, \
        操業年月日至 \
    FROM \
        [OP_ENTRY_INORG].[dbo].[T_TRACE_S法_REPORT_縦持ち_履歴] \
    WHERE \
        銘柄 = ? \
        AND ランNO = ? \
    ", brand, run_no1, )
    tracereport = cursor.fetchall()
    for tracereport in tracereport:
        operation_date_start = tracereport.操業年月日自.strftime('%Y%m%d%H%M')
        operation_date_start = operation_date_start[0:4] + "/" + operation_date_start[4:6] + "/" + operation_date_start[6:8]
        operation_date_end = tracereport.操業年月日至.strftime('%Y%m%d%H%M')
        operation_date_end = operation_date_end[0:4] + "/" + operation_date_end[4:6] + "/" + operation_date_end[6:8]

    print('データ取得完了')

    # 現在日時を取得。ファイル名に加える --- (*1)
    today = datetime.datetime.now().strftime('%Y%m%d%H%M')

    # EXCEL読込 --- (*2)
    template = '\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\テンプレート\\【帳票テンプレート】品質検討会資料(S法).xlsx'
    print(template)
    wb = openpyxl.load_workbook(template)

    # アクティブなワークシートを得る --- (*3)
    ws = wb["品質検討会資料"]
    ws2 = wb["form"]
    ws3 = wb["pos"]

    # データ書き込み --- (*4)
    ws["F1"] = "XXXX年    X月度    " + brand + "    XX工程"
    ws["A2"] = "（ " + str(run_no1) + " ）前    回    ラ    ン    概    況 [ " + operation_date_start + " ～ " + operation_date_end + " ]"

    for i in range(5, aggregate_list.__len__()+5, 1):
        ws['A' + str(i)] = aggregate_list[i-5][0]
        ws['F' + str(i)] = aggregate_list[i-5][1]
        ws['H' + str(i)] = aggregate_list[i-5][2]
        ws['J' + str(i)] = aggregate_list[i-5][3]
        # for j in range(0, avg1s_list.__len__(), 1):
        #     if aggregate_list[i-5][0] == avg1s_list[j][0]:
        #         ws['J' + str(i)] = avg1s_list[j][1]
        for j in range(0, avg2s_list.__len__(), 1):
            if aggregate_list[i-5][0] == avg2s_list[j][0]:
                ws['L' + str(i)] = avg2s_list[j][1]
        for j in range(0, avg3s_list.__len__(), 1):
            if aggregate_list[i-5][0] == avg3s_list[j][0]:
                ws['N' + str(i)] = avg3s_list[j][1]
        # 自動集計部最大項目数以上は取得不可 テンプレートに収まりきらないので切り捨てる
        if i == 64:
            break

    print('データ抽出完了')

    # パターンの取得 ["pos"]から銘柄比較してパターン取得
    pattern = ''
    for row in ws3.iter_rows(min_row=51, max_row=451, min_col=2, max_col=2):
        for data in row:
            if data.value == brand:
                pattern = ws3.cell(row=data.row, column=data.column + 1).value

    # パターン毎に処理分岐
    if pattern == 'A':
        for i in range(1, 8):
            for j in range(1, 18):
                copy = ws2.cell(row=i+23, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'B':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+32, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'C':
        for i in range(1, 12):
            for j in range(1, 18):
                copy = ws2.cell(row=i+41, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'D':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+53, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'E':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+62, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'F':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+66, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'G':
        for i in range(1, 11):
            for j in range(1, 18):
                copy = ws2.cell(row=i+70, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'H':
        for i in range(1, 4):
            for j in range(1, 18):
                copy = ws2.cell(row=i+81, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'I':
        for i in range(1, 14):
            for j in range(1, 18):
                copy = ws2.cell(row=i+85, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'J':
        for i in range(1, 7):
            for j in range(1, 18):
                copy = ws2.cell(row=i+99, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'K':
        for i in range(1, 8):
            for j in range(1, 18):
                copy = ws2.cell(row=i+106, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'L':
        for i in range(1, 9):
            for j in range(1, 18):
                copy = ws2.cell(row=i+114, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'M':
        for i in range(1, 5):
            for j in range(1, 18):
                copy = ws2.cell(row=i+123, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)
    elif pattern == 'N':
        for i in range(1, 5):
            for j in range(1, 18):
                copy = ws2.cell(row=i+128, column=j+1).value
                ws.cell(row=i+65, column=j, value=copy)

    # header
    ws.oddHeader.right.text = today[0:4] + "年" + today[4:6] + "月" + today[6:8] + "日\n"

    # 署名
    img = openpyxl.drawing.image.Image("\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\テンプレート\\署名.png")
    ws.add_image(img, 'F92')

    # ファイルを保存 --- (*5)
    file_path = '\\\\ydomnserv\\common\\部門間フォルダ\\操業データ入力\\0000_取込ファイル\\酸化チタン\\品質検討会資料\\' + today + '_品質検討会資料(S法).xlsx'
    file_name = today + '_品質検討会資料(S法).xlsx'
    wb.save(file_path)

    with open(file_path, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(file_name)[0] or 'application/octet-stream')
    return response


# /***トレースNO採番処理***/
# 「酸化チタン_CL法_トレース採番MAIN処理」「酸化チタン_S法_トレースレポートMAIN処理」
def trace_no_numbering(request):
    group = request.POST['group']
    from_date = request.POST['from_date'] + ' 00:00:00'
    to_date = request.POST['to_date'] + ' 23:59:00'
    # from_date = '2022-04-11 00:00:00'
    # to_date = '2022-04-11 23:59:00'
    print('開始')

    # DB接続
    host_name = gethostname()
    if host_name == 'YWEBSERV1':
        print('本番環境起動')
        server = 'YSQLSERV4'
    elif host_name == 'I7161DD6':
        print('テスト環境起動')
        server = 'Y0033OUT,1434'
    else:
        print('開発環境起動')
        server = 'Y0033OUT,1434'
    database = 'OP_ENTRY_INORG'
    username = 'cmd_user'
    password = 'cmd_user'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # ストアド実行
    cursor.execute(" \
    EXEC [酸化チタン_" + group + "_トレース採番MAIN処理] N'" + from_date + "', N'" + to_date + "' \
    COMMIT")

    print("完了")

    data = {
        'massage': "処理が完了しました！",
    }
    return JsonResponse(data)
